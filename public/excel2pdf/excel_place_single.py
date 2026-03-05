#!F:/projects/flask-app-env/Scripts/python.exe
from __future__ import print_function
from operator import itemgetter
from itertools import groupby
import sys
import os
import shutil
from shutil import make_archive
import simplejson as json
import fitz
import mysql.connector
import pprint
import time
from mysql.connector import Error
import textwrap
from PIL import Image, ImageOps 
import numpy as np
#import cv2.cv2 as cv2
import barcode
from barcode.writer import ImageWriter
import qrcode
import hashlib 
from datetime import datetime
from xml.dom.minidom import parseString
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font 
from openpyxl.styles import Alignment
import openpyxl.utils.cell
from datetime import datetime
import uuid
import pytz
import xlrd
import locale
from getFunctions import *


tz_IND = pytz.timezone('Asia/Calcutta') 
#sys.argv[1] = template id
#sys.argv[2] = data file
#sys.argv[3] = session user id
#sys.argv[4] = entry type (Fresh/Proceed)
#sys.argv[5] = progress file
# python excel_place_single.py 8 marksheet_BArchIX.xlsx 1
# python excel_place_single.py 16 bed1year.xlsx 1
# python excel_place_single.py 20 anu_college.xlsx 1
#print(sys.argv[2])
try:
    connection = mysql.connector.connect(host=host_var, database=database_var, user=user_var, password=password_var)
    if connection.is_connected():
        db_Info = connection.get_server_info()
        cursor = connection.cursor()
        cursor.execute("select ep_details, id, extractor_details, template_name, pdf_page, print_bg_file, print_bg_status, verification_bg_file, verification_bg_status from uploaded_pdfs where id = '%s'" % (sys.argv[1]))
        record = cursor.fetchone()        
        boxes = json.loads(record[0], strict=False)       
        template_id=record[1]
        extractor_details = json.loads(record[2], strict=False)
        template_name=record[3]
        pdf_page=record[4]
        pbg_file=record[5]
        print_bg_status=record[6]
        vbg_file=record[7]
        verification_bg_status=record[8]
        cur=connection.cursor()
        print_setbg=''
        verification_setbg=''
        if pbg_file !=0 and print_bg_status == 'Yes':
            sql_bg="SELECT filename FROM bg_data where id= '%s'" % (pbg_file)
            cur.execute(sql_bg)
            precord = cur.fetchone()
            print_bg_file=print_bg_file_var+precord[0]             
            print_setbg='Yes'
            #print(print_bg_file)
        if vbg_file !=0 and verification_bg_status == 'Yes':
            sql_bg="SELECT filename FROM bg_data where id= '%s'" % (vbg_file)
            cur.execute(sql_bg)
            vrecord = cur.fetchone()
            verification_bg_file=verification_bg_file_var+vrecord[0]            
            verification_setbg='Yes'
            #print(verification_bg_file)

        record_unique_id = datetime.now().strftime('%Y%m%d%H%M%S-') + str(uuid.uuid4()).split('-')[-1]
        sqli="SELECT id FROM file_records ORDER BY id DESC LIMIT 1"
        cur.execute(sqli)
        last_id = cur.fetchone()
        if last_id == None:
            file_records_next_id=1
        else:
            file_records_next_id=last_id[0]+1    
        # get last record's columns
        #for lastID in last_id:
            #print(lastID)     

except Error as e:
    print("Error while connecting to MySQL", e)
"""
finally:
    if (connection.is_connected()):
        #cursor.close()
        #connection.close()
        print("MySQL connection is closed")
"""

#target_folder="symbiosis/BBA_LOGISTIC" 
excelfile="anu_college.xlsx" #marksheet_BArchIX.xlsx bed1year.xlsx anu_college.xlsx 
wb = xlrd.open_workbook(excelfile) # input file name
sheet = wb.sheet_by_index(0)
rows=sheet.nrows-1
cols=sheet.ncols 
#print(cols, rows)
excel_cols = {}
# read header values into the list    
keys = [sheet.cell(0, col_index).value for col_index in range(sheet.ncols)]
#keys = [[col_index+1,sheet.cell(0, col_index).value] for col_index in range(sheet.ncols)]
#col_inx=keys.index('CandidateName')
#print(col_inx)
#print(openpyxl.utils.cell.get_column_letter(1))
#exit()
wb_obj = openpyxl.load_workbook(excelfile, read_only=True, data_only=True)
sheet_obj = wb_obj.active
max_rows=sheet_obj.max_row
max_columns=sheet_obj.max_column
"""
for data in sheet_obj.iter_rows(max_col=max_columns,max_row=max_rows,values_only=True):
    for d in range(len(data)):
        print(data[d],end=' ') # print all data of a row
    print('') # adding line break 
#cell_obj = sheet_obj.cell(row = 2, column = 10)
#print(cell_obj.value)
exit()
"""
my_directory="F:/projects/flask-app-env/marksheet/grade_card_reader/defence_secure_docs/excel_pdfs"
doc_new = fitz.open()
dict_list = []
cntn = 1
cln=0
for row_index in range(1, sheet.nrows):
    #d = {keys[col_index]: sheet.cell(row_index, col_index).value 
        #for col_index in range(sheet.ncols)}
    #dict_list.append(d)
    doc_new.insertPage(-1)    
    cln += 1
    cntn += 1

white = (0, 0, 0, 0)
black = (0, 0, 0, 1)
red = (0, 1, 1, 0)
blue = (1, 1, 0, 0)
pink = (0, 1, 0, 0)
aqua = (1, 0, 0, 0)
green = (1, 0, 1, 0)
purple = (1, 1, 0, 0)
yellow = (0, 0, 1, 0)

locale.setlocale(locale.LC_ALL, '')
l = locale.localeconv()
div1    = l['decimal_point'] or '.'
div1000 = l['thousands_sep'] or ','

doc_new.save(my_directory+"/output.pdf", garbage=4, deflate=True)    
cnt = 1
doc = fitz.open(my_directory+"/output.pdf")
page_count=doc.pageCount

for i in doc:
    page = i
    if not(page._isWrapped):
        page._wrapContents()
    page_data = {cnt:[]}
    words = page.getTextWords()
    page_dict = page.getText('dict')    
    page_rect=page.MediaBox
    row_value = sheet.row_values(cnt)    
    column=1
    #page.drawLine(fitz.Point(32, 522), fitz.Point(526, 522), width = 0.5, color = black)
    for box_line in extractor_details:
        line_coords = box_line['coords']
        lcoords = line_coords.split(",")
        lrect = fitz.Rect(lcoords[0],lcoords[1],lcoords[2],lcoords[3]) 
        #print(line_coords)
        #print(lrect.br,lrect.bl)  
        rect_width=lrect.width
        rect_height=lrect.height
        if rect_width > rect_height:
            p1 = fitz.Point(lcoords[0],lcoords[1])
            p2 = fitz.Point(lcoords[2],lcoords[1])
        else:
            p1 = fitz.Point(lcoords[0],lcoords[1])
            p2 = fitz.Point(lcoords[0],lcoords[3])
        #page.drawRect(lrect, color = (0, 0, 0), width=0.5, overlay=True)
        page.drawLine(p1, p2, width = 0.5, color = black)
    
    for box in boxes:
        placer_coords = box['placer_coords']
        pcoords = placer_coords.split(",")
        prect = fitz.Rect(pcoords[0],pcoords[1],pcoords[2],pcoords[3])
        prect_coords = pcoords[0],pcoords[1],pcoords[2],pcoords[3]
        #print(placer_coords)  
        if box['placer_font_name'] == '':  
            placer_font_name = fitz.Font(fontfile=dirFont+"arial.ttf")   
        else:
            placer_font_name = fitz.Font(fontfile=dirFont+box['placer_font_name'])
        
        placer_font_underline = box['placer_font_underline'] 
        placer_font_size = box['placer_font_size']                   
        placer_type = box['placer_type']      
        placer_display = box['placer_display']
        
        if placer_display == '':
            placer_align = int(0)
        else:
            placer_align = int(placer_display)         
            
        if placer_type == 'Invisible':
            placer_color = yellow
        else:
            placer_color = black
        
        placer_degree_angle = box['degree_angle']        
        placer_opacity = box['opacity_val']
        
        
        if box['font_color'] == '':
            placer_font_color = black
        else:
            placer_font_color = box['font_color']
            placer_font_color=fitz.utils.getColor(placer_font_color)   

                  
        if placer_type == 'Static Text':
            otxt_string = str(box['qr_details'])
            chk_otxt_string = str(box['qr_details'])
        else:
            if placer_type != 'Box':
                col_inx=keys.index(box['source'])
                #c = sheet.cell(cnt,col_inx)
                #print(sheet.cell(cnt,col_inx).value)
                #if box['source'] in skip_cols:
                #print(col_inx+1)
                otxt_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                chk_otxt_string = sheet_obj.cell(row = cnt+1, column = col_inx+1).value
                #print(otxt_string) 
                """
                if type(sheet_obj.cell(row = cnt+1, column = col_inx+1).value) is float:                    
                    otxt_string = repr(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                else:
                    otxt_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                """
            else:
                otxt_string=' '
                chk_otxt_string=' '
        #print(otxt_string, c.ctype) #repr() text into single quotes
        #otxt_string = str(row_value[column]).split(".")[0]   
        #print(box['source'], box['placer'], otxt_string, column)
               
        prects,prects_coords = text_tl_position(pcoords[0], pcoords[1], pcoords[2], pcoords[3])
        m = fitz.Matrix(placer_degree_angle)
        points = fitz.Point(prects.x0, prects.y0)        
        if type(chk_otxt_string) is float:      
            page.insertTextbox(prect, otxt_string, fontfile=str(placer_font_name), fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)
        else:
            wr = fitz.TextWriter(page.rect)
            wr.fillTextbox(prect, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
            wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True, morph=(points,m))  
          
        if placer_type == 'Box':  
            page.drawRect(prect, color=black, fill=None, dashes=None, width=0.5, lineCap=0, lineJoin=0, morph=None, overlay=True, stroke_opacity=1, fill_opacity=1)  
                  
        column += 1
    cnt += 1
doc.save(my_directory+"/output_fill.pdf", garbage=4, deflate=True)
#print(dict_list)
exit()


userid=sys.argv[3]
encrypt_meth = fitz.PDF_ENCRYPT_AES_256  # strongest algorithm
perm = int(
fitz.PDF_PERM_PRINT  # permit printing
)
doc = fitz.open(directory+"uploads/data/"+sys.argv[2])
page_count=doc.pageCount
arr_content = {} #The array for storing the progress.
#Check duplicate ids
if sys.argv[4]=='Fresh':
    get_pdffiles = []
    for pname in doc:
        for boks in boxes:
            if boks['placer_type'] == 'QR Dynamic':
                get_text=[]
                if boks['source'] != '' and boks['source'] != 'Current DateTime':
                    scoords = boks['source_coords'].split(",")
                    srect = fitz.Rect(scoords[0],scoords[1],scoords[2],scoords[3])                
                    get_text.append(pname.getTextbox(srect).replace("\n", ""))
                    extracted_id = get_text[0].replace("/", "").replace("\\", "").replace("-", "").replace(":", "").strip()
                    if connection.is_connected():
                        cursor.execute("select * from individual_records where unique_no = '%s' and publish=1 order by id desc" % (extracted_id))
                        records = cursor.fetchall()  
                        row_count = cursor.rowcount
                        if row_count > 0:    
                            get_pdffiles.append(records[0][7])  

    if not get_pdffiles:
        print("No Duplicates")    #List is empty
    else:
        print("Duplicates")
        print(len(get_pdffiles))
        get_pdffiles_string = ','.join(get_pdffiles)
        print(get_pdffiles_string) #Duplicate unique ids
        """
        for unid in get_pdffiles:
            if connection.is_connected():
                sql = "UPDATE individual_records SET publish = 2 WHERE unique_no = '%s'" % (unid)
                cursor.execute(sql)
                connection.commit()
        """
        exit()


final_data = {'data':[]}
cnt = 1
pp = pprint.PrettyPrinter(indent=4)
output_file = directory+"processed_pdfs/"+sys.argv[2]

white = (0, 0, 0, 0)
black = (0, 0, 0, 1)
red = (0, 1, 1, 0)
blue = (1, 1, 0, 0)
pink = (0, 1, 0, 0)
aqua = (1, 0, 0, 0)
green = (1, 0, 1, 0)
purple = (1, 1, 0, 0)
yellow = (0, 0, 1, 0)

folder=directory+"documents/" + template_name
if not os.path.exists(folder):
    os.makedirs(folder)

inner_folder=folder +"/"+str(record_unique_id)
if not os.path.exists(inner_folder):
    os.makedirs(inner_folder)
    
pdf_folder=inner_folder +"/pdfs"
if not os.path.exists(pdf_folder):
    os.makedirs(pdf_folder)
    
path_pdf_moved=inner_folder+"/" + sys.argv[2]  
#Create excel file to save unique id and QR details
workbook_name=inner_folder+"/" + template_name+".xlsx"
wbc = openpyxl.Workbook()            
Sheet_name = wbc.sheetnames
wbc.save(filename=workbook_name)
wbs = load_workbook(workbook_name)
page_sheet = wbs.active
page_sheet.column_dimensions['A'].width = 30
page_sheet.column_dimensions['B'].width = 50
page_sheet.column_dimensions['C'].width = 120
wrap_alignment = Alignment(wrap_text=True)
page_sheet.append(['Original ID','Unique ID','QR Details'])  
page_sheet.cell(row = 1, column = 1).font = Font(bold = True) 
page_sheet.cell(row = 1, column = 2).font = Font(bold = True) 
page_sheet.cell(row = 1, column = 3).font = Font(bold = True)
datetime_IND = datetime.now(tz_IND) 
beginning_time = datetime_IND.strftime("%H:%M:%S")
start_time = time.time()
for i in doc:
    page = i
    if not(page._isWrapped):
        page._wrapContents()
    page_data = {cnt:[]}
    words = page.getTextWords()
    page_dict = page.getText('dict')    
    page_rect=page.MediaBox
    #print(page_rect)
    #exit()
    dirName = directory+"documents/" + str(sys.argv[1])
    for box in boxes:
        temp = ''
        otxt = ''
        file_path = ''
        #print()
        if box['source'] == '' or box['source'] == 'Current DateTime':
            srect = fitz.Rect(0,0,0,0)            
        else:
            source_coords = box['source_coords']
            scoords = source_coords.split(",")
            srect = fitz.Rect(scoords[0],scoords[1],scoords[2],scoords[3])        
        
        placer_coords = box['placer_coords']
        pcoords = placer_coords.split(",")
        prect = fitz.Rect(pcoords[0],pcoords[1],pcoords[2],pcoords[3])
        prect_coords = pcoords[0],pcoords[1],pcoords[2],pcoords[3]
        #print(placer_coords)  
        if box['placer_font_name'] == '':  
            placer_font_name = fitz.Font(fontfile=dirFont+"arial.ttf")   
        else:
            placer_font_name = fitz.Font(fontfile=dirFont+box['placer_font_name'])
            
        placer_font_underline = box['placer_font_underline'] 
        placer_font_size = box['placer_font_size']                   
        placer_type = box['placer_type']      
        placer_display = box['placer_display']
        if placer_display == '':
            placer_align = int(0)
        else:
            placer_align = int(placer_display)         
            
        if placer_type == 'Invisible':
            placer_color = yellow
        else:
            placer_color = black
        
        placer_degree_angle = box['degree_angle']        
        placer_opacity = box['opacity_val']
        placer_lineHeight = box['line_height']
        
        if box['font_color'] == '':
            placer_font_color = black
        else:
            placer_font_color = box['font_color']
            placer_font_color=fitz.utils.getColor(placer_font_color) 
        
        if placer_type == 'QR Default': 
            if not os.path.exists(dirName):
                os.makedirs(dirName) 
            now = datetime.now()            
            #dt_string = now.strftime("%Y-%m-%d %H:%M:%S")+str(cnt)
            dt_string = now.strftime("%Y%m%d%H%M%S")+str(cnt)
            result = hashlib.md5(dt_string.encode()) 
            barcode_en=result.hexdigest()           
            qr_txt=barcode_en
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=4,
                border=0,
            )
            qr.add_data(qr_txt)
            qr.make(fit=True)
            img = qr.make_image()  # fill_color="black", back_color="white"
            img.save(dirName+"/"+"qr_"+str(barcode_en)+".png")
            qrcode_file=dirName+"/"+"qr_"+str(barcode_en)+".png"            
            page.insertImage(prect, qrcode_file, overlay=True)    
        elif placer_type == 'QR Dynamic': 
            if not os.path.exists(dirName):
                os.makedirs(dirName) 
            qr_txt=''
            now = datetime.now()          
            get_text=[]
            if box['source'] == '' or box['source'] == 'Current DateTime':
                #dt_string = now.strftime("%Y-%m-%d %H:%M:%S")+str(cnt)
                dt_string = now.strftime("%Y%m%d%H%M%S")+str(cnt)
                result = hashlib.md5(dt_string.encode()) 
                barcode_en=result.hexdigest()
            else:
                #dt_string = otxt.replace("/", "")+now.strftime("%Y-%m-%d %H:%M:%S")+str(cnt)                
                #dt_string = otxt.replace("/", "")
                #dt_string = otxt
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                dt_string = get_text[0].replace("/", "").replace("\\", "").replace("-", "").replace(":", "").strip()
                result = hashlib.md5(dt_string.encode()) 
                barcode_en=result.hexdigest()         
                
            qr_details = box['qr_details'].replace("{", "").replace("}", "")
            qr_list=list(filter(bool, qr_details.splitlines()))
            for n in qr_list:
                str_val = n.split("^")
                str_val_count=len(str_val)
                qr_txt = qr_txt + extract_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)
            qr_txt =  qr_txt + "\n"+barcode_en
            #print(qr_txt)
            
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=4,
                border=0,
            )
            qr.add_data(qr_txt)
            qr.make(fit=True)
            img = qr.make_image()  # fill_color="black", back_color="white"
            img.save(dirName+"/"+"qr_"+str(barcode_en)+".png")
            qrcode_file=dirName+"/"+"qr_"+str(barcode_en)+".png"            
            page.insertImage(prect, qrcode_file, overlay=True)
        elif placer_type == 'Barcode': 
            get_text=[]
            get_text.append(page.getTextbox(srect).replace("\n", ""))
            dt_string = get_text[0].strip()
            EAN = barcode.get_barcode_class('code128')
            ean = EAN(dt_string, writer=ImageWriter())
            options = {
                'dpi': 200,
                'write_text': False,
                'module_width': 5,
                'module_height': 100,
                'quiet_zone': 0,
                'text_distance': 0
            }            
            barcode_file = ean.save(dirName, options = options)               
            page.insertImage(prect, barcode_file, overlay=True)             
        elif placer_type == 'Micro Line': 
            mi_txt=''            
            if box['qr_details']=='':
                get_text=[]
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                text_len=len(temp)
                temp = get_text[0].replace(" ", "").strip()
                if isinstance(placer_font_size, float)==True:
                    textwidth = fitz.getTextlength(temp, fontsize=float(placer_font_size))    
                else:
                    textwidth = fitz.getTextlength(temp, fontsize=int(placer_font_size))   
                            
                #print(textwidth)
                chrPerLine=int(float(prect.width))/textwidth
                Totalchrs=str(chrPerLine).split(".")[0]
                #print ("chr = ", chrPerLine)
                #print (Totalchrs)
                repeat_txt=temp * int(Totalchrs)
                if isinstance(placer_font_size, float)==True:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                else:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                
                #print(int(float(prect.width)),"|",int(textwidth),"|",int(float(repeat_textwidth)))
                remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3        
                #print("rs: ",remain_space,"|",text_len)
                if remain_space>0:                
                    RchrPerLine=int(remain_space)/textwidth
                    remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                    remain_chrs = temp[0:int(remain_chrs_count)]
                else:
                    remain_chrs = ''
                    
                #print(remain_chrs) # 0 = left, 1 = center, 2 = right
                page.insertTextbox(prect, (temp * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)
                #wr = fitz.TextWriter(page.rect)
                #wr.fillTextbox(prect, (temp * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align)
                #wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)    
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    #mi_txt = mi_txt + extract_microline_info(extractor_details,inv,words)  
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    mi_txt = mi_txt + extract_microline_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                    
                #print(mi_txt)      
                mi_txt = mi_txt.replace(" ", "")
                text_len=len(mi_txt)
                if isinstance(placer_font_size, float)==True:
                    textwidth = fitz.getTextlength(mi_txt, fontsize=float(placer_font_size))    
                else:
                    textwidth = fitz.getTextlength(mi_txt, fontsize=int(placer_font_size))   
                            
                chrPerLine=int(float(prect.width))/textwidth
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=mi_txt * int(Totalchrs)
                if isinstance(placer_font_size, float)==True:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                else:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                
                remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3        
                if remain_space>0:                
                    RchrPerLine=int(remain_space)/textwidth
                    remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                    remain_chrs = mi_txt[0:int(remain_chrs_count)]
                else:
                    remain_chrs = ''
                    
                page.insertTextbox(prect, (mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)
                #wr = fitz.TextWriter(page.rect)
                #wr.fillTextbox(prect, (mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align)
                #wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)
        elif placer_type == 'Invisible':                        
            inv_txt=''
            get_text=[]
            if box['qr_details']=='':  
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                otxt=get_text[0].strip()
                #print(otxt)
                #page.insertTextbox(prect, otxt, fontfile=placer_font_name, fontsize=placer_font_size, color=placer_color, align=placer_align, overlay=True)
                prects,prects_coords = text_tl_position(pcoords[0], pcoords[1], pcoords[2], pcoords[3])
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prects, otxt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_color, opacity=float(1.0), overlay=True) 
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    inv_txt = inv_txt + extract_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                          
                     
                #page.insertTextbox(prect, inv_txt, fontfile=placer_font_name, fontsize=placer_font_size, color=placer_color, align=placer_align, overlay=True)
                prects,prects_coords = text_tl_position(pcoords[0], pcoords[1], pcoords[2], pcoords[3])
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prects, inv_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_color, opacity=float(1.0), overlay=True) 
        elif placer_type == 'Invisible Image': 
            temp_img = [block for block in page_dict['blocks'] if (fitz.Rect(block['bbox']) in srect and block['type'] == 1)]  
            if len(temp_img) > 0:
                pix = fitz.Pixmap(temp_img[0]['image'])
                if not os.path.exists(dirName):
                    os.makedirs(dirName)
                file_path = dirName + "/"+str(int(round(time.time() * 1000))) + "-" + str(cnt) + "." + temp_img[0]['ext']
                file_path2 = dirName + "/"+str(int(round(time.time() * 1000))) + "-" + str(cnt) + ".png"                        
                pix.writeImage(file_path)
                imgs = Image.open(file_path).convert("L") 
                img1 = ImageOps.colorize(imgs, black ="white", white ="yellow")             
                img1.save(file_path2, 'png')
                page.insertImage(prect, file_path2, overlay=True) 
        elif placer_type == 'Ghost Image': 
            #print(prect.width,"| ",prect.height)
            if placer_degree_angle==90:
                half_width=int(float(prect.height))/2
                top_pos_minus=(int(float(pcoords[0])) - int(float(half_width)))  
                add_right_pos=int(float(pcoords[0])) - top_pos_minus  
                left_pos=int(float(pcoords[0]))-add_right_pos
                top_pos=int(float(pcoords[1]))
                right_pos=int(float(pcoords[2])) + add_right_pos
                bottom_pos=int(float(pcoords[3]))               
                prects = fitz.Rect(left_pos,top_pos,right_pos,bottom_pos) 
                #print(prects)
                #print(prects.width,"| ",prects.height)                         
            else:
                prects=prect
            
            get_text=[]
            get_text.append(page.getTextbox(srect).replace("\n", ""))
            temp = ''.join(c for c in get_text[0].strip().replace("/", "").replace(".", "") if c.isalnum())
            ghost_width = round(box['width'] * 3.7795275591)  # Millimeter to Pixel, 1 mm = 3.7795275591 pixel
            ghost_height = round(box['height'] * 3.7795275591)  
            ghost_words = box['ghost_words']  
            PrintableChars=temp[ 0 : ghost_words ] #extract first chars
            if not os.path.exists(dirName):
                os.makedirs(dirName)            
            ghostImg=CreateGhostImage(dirName, PrintableChars, placer_font_size, ghost_width, ghost_height)
            page.insertImage(prects, ghostImg,overlay=True, keep_proportion=True, rotate=placer_degree_angle)        
        elif placer_type == 'Image':
            image_path=directory+"upload_images/" + box['image_path']
            page.insertImage(prect, image_path,overlay=True)
        elif placer_type == 'Plain Text':           
            inv_txt=''
            get_text=[]
            if box['qr_details']=='':                
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                otxt_string = get_text[0].strip()
                #otxt_string = otxt.replace("/", "")
                #page.insertTextbox(prect, otxt_string, fontsize=placer_font_size, color=placer_font_color, align=placer_align, overlay=True)
                prects,prects_coords = text_tl_position(pcoords[0], pcoords[1], pcoords[2], pcoords[3])
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prects, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)                 
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    inv_txt = inv_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                          
                     
                #page.insertTextbox(prect, inv_txt, fontsize=placer_font_size, color=placer_font_color, align=placer_align, overlay=True) 
                prects,prects_coords = text_tl_position(pcoords[0], pcoords[1], pcoords[2], pcoords[3])
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prects, inv_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)               
        elif placer_type == 'Static Text':     
            otxt_string = box['qr_details']
            prects,prects_coords = text_tl_position(pcoords[0], pcoords[1], pcoords[2], pcoords[3])
            m = fitz.Matrix(placer_degree_angle)
            points = fitz.Point(prects.x0, prects.y0)
            wr = fitz.TextWriter(page.rect)
            wr.fillTextbox(prects, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
            wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))
            if placer_font_underline=="underline":
                rl = page.searchFor(otxt_string)  
                output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")                    
                ucoords = output.split(",")
                urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                shape = page.newShape()
                shape.drawLine(urect.bl, urect.br)
                shape.finish(color=placer_font_color, stroke_opacity=float(placer_opacity))
                shape.commit()            
        elif placer_type == 'Watermark Text':           
            inv_txt=''
            get_text=[]
            if box['qr_details']=='':                
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                otxt_string = get_text[0].strip()
                prects,prects_coords = text_tl_position(pcoords[0], pcoords[1], pcoords[2], pcoords[3])
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(prects_coords)
                new_rect=ir * m   
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(prects.x0, prects.y0)                 
                wr.fillTextbox(prects, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))                
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    inv_txt = inv_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                          
                    
                prects,prects_coords = text_tl_position(pcoords[0], pcoords[1], pcoords[2], pcoords[3])
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(prects_coords)
                new_rect=ir * m  
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(prects.x0, prects.y0)                 
                wr.fillTextbox(prects, inv_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))        
        elif placer_type == 'Watermark Multi Lines':  
            mi_txt=''            
            if box['qr_details']=='':
                get_text=[]
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                text_len=len(temp)
                temp = get_text[0].strip()
                chrPerLine=int(float(page_rect.width)+int(float(page_rect.height)))
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=temp * int(Totalchrs)       
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(page_rect)
                new_page_rect=ir * m         
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(0, 0)
                wr.fillTextbox(new_page_rect, repeat_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    mi_txt = mi_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                    
                #print(mi_txt)      
                text_len=len(mi_txt)
                chrPerLine=int(float(page_rect.width)+int(float(page_rect.height)))
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=mi_txt * int(Totalchrs)
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(page_rect)
                new_page_rect=ir * m         
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(0, 0)  #prect.x0, prect.y0
                wr.fillTextbox(new_page_rect, repeat_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))
                          
        #page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
        #page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_REMOVE)
        doc_new = fitz.open()
        doc_new.insertPDF(doc, from_page=cnt-1, to_page=cnt-1)
        if verification_setbg=='Yes':
            for dpage in doc_new:
                dpage.insertImage(page_rect, verification_bg_file,overlay=False)        
        doc_new.save(pdf_folder+"/"+dt_string+".pdf", garbage=4, deflate=True, owner_pw="owner", encryption=encrypt_meth, permissions=perm) 
        """
        set_background(
            input_pdf=verification_bg_file, 
            output=pdf_folder+"/"+dt_string+".pdf",
            watermark=pdf_folder+"/"+dt_string+".pdf")   
        """       
    if print_setbg=='Yes':
        page.insertImage(page_rect, print_bg_file,overlay=False)      
    if connection.is_connected():
        sql = "INSERT INTO individual_records (file_records_id, template_id, template_name, pdf_page, page_no, encoded_id, unique_no, qr_details, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        val = (file_records_next_id, template_id, template_name, pdf_page, cnt, barcode_en, dt_string, qr_txt, userid, record_unique_id)
        cursor.execute(sql, val)         
        
    page_sheet.append([dt_string,barcode_en,qr_txt])
    page_sheet.cell(row = cnt+1, column = 1).alignment = Alignment(vertical='top')
    page_sheet.cell(row = cnt+1, column = 2).alignment = Alignment(vertical='top')
    page_sheet.cell(row = cnt+1, column = 3).alignment = Alignment(wrapText=True,vertical='top')
    wbs.save(filename=workbook_name) 
    datetime_IND = datetime.now(tz_IND)
    ending_time = datetime_IND.strftime("%H:%M:%S")
    end_time=time.time() - start_time
    seconds_to_hhmmss=time.strftime('%H:%M:%S', time.gmtime(end_time))
    page_seconds_to_hhmmss=time.strftime('%H:%M:%S', time.gmtime(end_time/cnt))
    arr_content['percent'] = int(cnt/page_count * 100)
    arr_content['message'] = "Generating "+str(cnt)+"/"+str(page_count)+" PDF(s)"
    arr_content['beginning_time'] = beginning_time
    arr_content['ending_time'] = ending_time
    arr_content['exec_time'] = end_time
    arr_content['hms_time'] = seconds_to_hhmmss
    arr_content['page_time'] = page_seconds_to_hhmmss
    arr_content['pages_processed'] = cnt
    json_object = json.dumps(arr_content, indent = 4)    
    f = open(directory+"processed_pdfs/"+sys.argv[5], "w")
    f.write(json_object)
    #time.sleep(1)
    cnt += 1

if connection.is_connected():
    sql = "INSERT INTO file_records (template_id, template_name, pdf_page, total_records, source_file, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    val = (template_id, template_name, pdf_page, cnt-1, sys.argv[2], userid, record_unique_id)
    cursor.execute(sql, val)
    connection.commit()
 
doc.save(path_pdf_moved, garbage=4, deflate=True, owner_pw="owner", encryption=encrypt_meth, permissions=perm)
#shutil.copyfile(output_file, path_pdf_moved)
shutil.make_archive(folder+"/"+template_name+"-"+str(record_unique_id), "zip", inner_folder)
print(folder+"/"+template_name+"-"+str(record_unique_id)+".zip")
total_pages=cnt-1
print("Total Records:"+str(total_pages))
shutil.rmtree(dirName, ignore_errors=True)
#shutil.rmtree(inner_folder, ignore_errors=True)
print("documents/"+template_name+"/"+template_name+"-"+str(record_unique_id)+".zip")
