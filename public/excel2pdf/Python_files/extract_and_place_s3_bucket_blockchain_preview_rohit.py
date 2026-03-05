#!C:/Inetpub/vhosts/seqrdoc.com/httpdocs/pdf2pdf/pdf_env/Scripts/python.exe
from __future__ import print_function
from operator import itemgetter
from itertools import groupby
import sys
import os
import shutil
from shutil import make_archive
from turtle import update
import simplejson as json
import fitz
import mysql.connector
import pprint
import time
from mysql.connector import Error
import textwrap
from PIL import Image, ImageOps 
import numpy as np
#import cv2.cv2 as cv2
import barcode
from barcode.writer import ImageWriter
import qrcode
import hashlib 
#from datetime import datetime
from xml.dom.minidom import parseString
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font 
from openpyxl.styles import Alignment
#from datetime import datetime
import uuid
import pytz
#from PyPDF2 import PdfFileWriter, PdfFileReader
import socket
import requests
import datetime
from datetime import date
from PIL import ImageDraw, ImageFont
import math
import subprocess
import tempfile
import base64

tz_IND = pytz.timezone('Asia/Calcutta') 
#sys.argv[1] = template id
#sys.argv[2] = data file
#sys.argv[3] = session user id
#sys.argv[4] = entry type (Fresh/Proceed)
#sys.argv[5] = progress file
#sys.argv[6] = dbName
#sys.argv[7] = subdomain
#sys.argv[8] = directoryUrlForward
#sys.argv[9] = directoryUrlBackward
#sys.argv[10] = servername
#sys.argv[11] = username
#sys.argv[12] = password
#sys.argv[13] = siteid
#sys.argv[14] = username
#sys.argv[15] = printer_name
# python extract_and_place.py 1 FYBAF_1619953609.pdf 1
#print(sys.argv[2])
try:
    directory=sys.argv[8]
    rootDir= directory.replace('pdf2pdf', '')
    connection = mysql.connector.connect(host=sys.argv[10],
                                         database=sys.argv[6],
                                         user=sys.argv[11],
                                         password=sys.argv[12])
    if connection.is_connected():
        db_Info = connection.get_server_info()
        cursor = connection.cursor(buffered=True)
        cursor.execute("select ep_details, id, extractor_details, template_name, pdf_page, print_bg_file, print_bg_status, verification_bg_file, verification_bg_status, IFNULL(bc_contract_address, '') AS bc_contract_address, bc_document_description, bc_document_type from uploaded_pdfs where id = '%s'" % (sys.argv[1]))
        record = cursor.fetchone()        
        boxes = json.loads(record[0], strict=False)        
        template_id=record[1]
        extractor_details = json.loads(record[2], strict=False)
        template_name=record[3]
        pdf_page=record[4]
        pbg_file=record[5]
        print_bg_status=record[6]
        vbg_file=record[7]
        verification_bg_status=record[8]
        bc_contract_address=record[9]
        bc_document_description=record[10]
        bc_document_type=record[11]
        cur=connection.cursor(buffered=True)
        print_setbg=''
        verification_setbg=''
        if pbg_file !=0 and print_bg_status == 'Yes':
            sql_bg="SELECT image_path FROM background_template_master where id= '%s'" % (pbg_file)
            cur.execute(sql_bg)
            precord = cur.fetchone()
            #print_bg_file=sys.argv[8]+"upload_bgs/"+precord[0] 
            print_bg_file=rootDir+sys.argv[7]+"/backend/canvas/bg_images/" + precord[0]
            #print_bg_file="C:/wamp/www/demo/upload_bgs/"+precord[0] 
            print_setbg='Yes'
            #print(print_bg_file)
        if vbg_file !=0 and verification_bg_status == 'Yes':
            sql_bg="SELECT image_path FROM background_template_master where id= '%s'" % (vbg_file)
            cur.execute(sql_bg)
            vrecord = cur.fetchone()
            #verification_bg_file=sys.argv[8]+"upload_bgs/"+vrecord[0]
            verification_bg_file=rootDir+sys.argv[7]+"/backend/canvas/bg_images/" + vrecord[0]
            #verification_bg_file="C:/wamp/www/demo/upload_bgs/"+vrecord[0]
            verification_setbg='Yes'
            #print(verification_bg_file)

        record_unique_id = datetime.datetime.now().strftime('%Y%m%d%H%M%S-') + str(uuid.uuid4()).split('-')[-1]
        
        #sqli="SELECT id FROM file_records ORDER BY id DESC LIMIT 1"
        # Update rohit 01/09/2023
        if sys.argv[18] == '1':
            sqli="SELECT id FROM sb_file_records ORDER BY id DESC LIMIT 1"
        else:
            sqli="SELECT id FROM file_records ORDER BY id DESC LIMIT 1"
        #Update Rohit 01/09/2023
        cur.execute(sqli)
        last_id = cur.fetchone()
        if last_id == None:
            file_records_next_id=1
        else:
            file_records_next_id=last_id[0]+1    
        # get last record's columns
        #for lastID in last_id:
            #print(lastID)     

except Error as e:
    print("Error while connecting to MySQL", e)
"""
finally:
    if (connection.is_connected()):
        #cursor.close()
        #connection.close()
        print("MySQL connection is closed")
"""
#exit()
#print(bc_contract_address)
userid=sys.argv[3]
def extract_microline_info(extractor_detail,search_value,extra_line_first='',extra_line_second='',page=''):
    for keyval in extractor_detail:
        if search_value.lower() == keyval['name'].lower():
            source_coord = keyval['coords']
            scoord = source_coord.split(",")
            s_rect = fitz.Rect(scoord[0],scoord[1],scoord[2],scoord[3])
            get_text=[]            
            get_text.append(page.getTextbox(s_rect).replace("\n", ""))
            otxt=get_text[0].replace(" ", "")
            if extra_line_first == '' and extra_line_second == '':
                return otxt
            elif extra_line_first != '' and extra_line_second != '':
                return extra_line_first + otxt + extra_line_second 
            elif extra_line_first != '' and extra_line_second == '':
                return extra_line_first + otxt 
            elif extra_line_first == '' and extra_line_second != '':
                return otxt + extra_line_second

def extract_plainText(extractor_detail,search_value,extra_line_first='',extra_line_second='',page=''):
    for keyval in extractor_detail:
        if search_value.lower() == keyval['name'].lower():
            source_coord = keyval['coords']
            scoord = source_coord.split(",")
            s_rect = fitz.Rect(scoord[0],scoord[1],scoord[2],scoord[3]) 
            get_text=[]            
            get_text.append(page.getTextbox(s_rect).replace("\n", ""))
            otxt=get_text[0].strip()
            if extra_line_first == '' and extra_line_second == '':
                return otxt +" "
            elif extra_line_first != '' and extra_line_second != '':
                return extra_line_first + otxt + extra_line_second +" " 
            elif extra_line_first != '' and extra_line_second == '':
                return extra_line_first + otxt +" " 
            elif extra_line_first == '' and extra_line_second != '':
                return otxt + extra_line_second +" "                 
                
def extract_info(extractor_detail,search_value,extra_line_first='',extra_line_second='',page=''):
    for keyval in extractor_detail:
        if search_value.lower() == keyval['name'].lower():
            source_coord = keyval['coords']
            scoord = source_coord.split(",")
            s_rect = fitz.Rect(scoord[0],scoord[1],scoord[2],scoord[3])
            get_text=[]            
            get_text.append(page.getTextbox(s_rect).replace("\n", ""))
            otxt=get_text[0].strip()
            if extra_line_first == '' and extra_line_second == '':
                return otxt +"\n"
            elif extra_line_first != '' and extra_line_second != '':
                return extra_line_first + otxt + extra_line_second +"\n" 
            elif extra_line_first != '' and extra_line_second == '':
                return extra_line_first + otxt +"\n" 
            elif extra_line_first == '' and extra_line_second != '':
                return otxt + extra_line_second +"\n" 

def CreateGhostImage(dirName, name, p_font_size, ghost_width, ghost_height):
    #dirChars = "F:/projects/flask-app-env/marksheet/grade_card_reader/defence_secure_docs/chars/"+str(p_font_size)
    #dirChars = "C:/Program Files/Python38/projects/demo/chars/"+str(p_font_size)
    dirChars=sys.argv[8]+'Python_files/chars/'+str(p_font_size);
    name=name.upper()    
    single_char=split(name)       
    my_list = list()
    for c in single_char:
        my_list.append(dirChars +"/"+ c +".png")    
    #print(my_list)
    images = [Image.open(x) for x in my_list]
    widths, heights = zip(*(i.size for i in images))
    total_width = sum(widths)
    max_height = max(heights)    
    new_im = Image.new('RGB', (total_width, max_height))
    x_offset = 0
    for im in images:
      new_im.paste(im, (x_offset,0))
      x_offset += im.size[0]

    new_im.save(dirName +"/"+ name +".png")  
    isize=ghost_width,ghost_height            
    im = Image.open(dirName +"/"+ name +".png")
    im.thumbnail((isize), Image.ANTIALIAS)
    im.save(dirName +"/"+ name +str(p_font_size)+"_th.png", quality=100)    
    #return dirName +"/"+ name +".png" 
    return dirName +"/"+ name +str(p_font_size)+"_th.png"     
    
def split(word): 
    return [char for char in word]

def get_financial_year(datestring):
            date = datetime.datetime.strptime(datestring, "%Y-%m-%d").date()
            #initialize the current year
            year_of_date=date.year
            #initialize the current financial year start date
            financial_year_start_date = datetime.datetime.strptime(str(year_of_date)+"-04-01","%Y-%m-%d").date()
            if date<financial_year_start_date:
                    return str(financial_year_start_date.year-1)[2:]+'-'+ str(financial_year_start_date.year)[2:]
            else:
                    return str(financial_year_start_date.year)[2:]+'-'+ str(financial_year_start_date.year+1)[2:]    

def repeat_to_length(string_to_expand, length):
    return (string_to_expand * (int(length/len(string_to_expand))+1))[:length]

def get_pil_text_size(text, font_size, font_name):
    font = ImageFont.truetype(font_name, font_size)
    size = font.getsize(text)
    return size


def set_background(input_pdf, output, watermark):
    watermark_obj = PdfFileReader(watermark)
    #watermark_obj.decrypt("owner")
    watermark_page = watermark_obj.getPage(0)

    pdf_reader = PdfFileReader(input_pdf)    
    pdf_writer = PdfFileWriter()

    for page in range(pdf_reader.getNumPages()):        
        page = pdf_reader.getPage(page)
        page.mergePage(watermark_page)
        #page.compressContentStreams()
        pdf_writer.addPage(page)

    with open(output, 'wb') as out:
        pdf_writer.write(out)

encrypt_meth = fitz.PDF_ENCRYPT_AES_256  # strongest algorithm
perm = int(
fitz.PDF_PERM_PRINT  # permit printing
)

dirFont = sys.argv[8]+"Python_files/fonts/"
#directory="C:/wamp/www/demo/"
#dirFont = "C:/Program Files/Python38/projects/demo/fonts/"

#doc = fitz.open(rootDir+sys.argv[7]+'/uploads/data/'+sys.argv[2])#directory+"uploads/data/"+
#Updated rohit 01/09/2023
if sys.argv[18] == '1':
    doc = fitz.open(rootDir+sys.argv[7]+'/uploads/data/sandbox/'+sys.argv[2])#directory+"uploads/data/sandbox/"+
else:
    doc = fitz.open(rootDir+sys.argv[7]+'/uploads/data/'+sys.argv[2])#directory+"uploads/data/"+
    
page_count=doc.pageCount
arr_content = {} #The array for storing the progress.

#Check print limit 
if connection.is_connected():
    #cursor.execute("select count(*) AS ts from student_table where site_id = '%s'" % (sys.argv[13]))
    # Update Rohit 01/09/2023
    if sys.argv[18] == '1':
        cursor.execute("select count(*) AS ts from sb_student_table where site_id = '%s'" % (sys.argv[13]))
    else:
        cursor.execute("select count(*) AS ts from student_table where site_id = '%s'" % (sys.argv[13]))
    # Update Rohit 01/09/2023
    rsStudent = cursor.fetchone()
    studentTableCounts=rsStudent[0]

connection2 = mysql.connector.connect(host=sys.argv[10],
                                     database='seqr_demo',
                                     user=sys.argv[11],
                                     password=sys.argv[12])
if connection2.is_connected():
    cursor2 = connection2.cursor(buffered=True)
    cursor2.execute ("UPDATE super_admin SET current_value='%s' WHERE site_id=%s " % (studentTableCounts, sys.argv[13]))    
    cursor2.execute("select value, current_value from super_admin where site_id = '%s'" % (sys.argv[13]))
    rsGenerated = cursor2.fetchone()  
    printLimit=int(rsGenerated[0])
    currentValue=int(rsGenerated[1])
    recordGenerated= currentValue + int(page_count)
    noOfCertificateCanGenerate=printLimit-currentValue

    siteurl_param=sys.argv[7]+".seqrdoc.com"
    #siteurl_param="demo.seqrdoc.com"
    cursor2.execute("select IFNULL(bc_wallet_address, '') AS bc_wallet_address from sites where site_url = '%s'" % (siteurl_param))
    rsSite = cursor2.fetchone()
    bc_wallet_address=rsSite[0]

    connection2.commit()

if currentValue == printLimit:
    print("Over Limit")
    print("You have reached a limit for generating PDF.")
    exit()

if recordGenerated > printLimit:
    print("Over Limit")
    if noOfCertificateCanGenerate > 1:
        print("You can generate "+str(noOfCertificateCanGenerate)+" PDFs.")
    else:
        print("You can generate "+str(noOfCertificateCanGenerate)+" PDF.")
    exit()

get_blank = []
get_extractor_name = []
pno=0
for pagechk in doc:
    pno += 1
    get_extractor_name.append('<br><b>Page '+str(pno)+':</b>')
    for box_chk in extractor_details:
        source_coords_chk = box_chk['coords']
        scoords = source_coords_chk.split(",")
        srect = fitz.Rect(scoords[0],scoords[1],scoords[2],scoords[3])
        blocks = pagechk.getText("blocks", srect)
        blocks.sort(key=lambda block: block[0])  # sort vertically ascending
        try:
            #block_split=str(blocks[0]).split(",")
            block_split=blocks[0]
        except IndexError:
            block_split=''
            extractor_name=box_chk['name']
        
        if block_split=='':
            get_blank.append(extractor_name)
            get_extractor_name.append(extractor_name+', ')

		
if get_blank:
	print("Empty Extractor")
	get_blank_string = ' '.join(get_extractor_name)
	print('Empty Extractor'+get_blank_string) #Blank Source
	exit()


#Check duplicate ids
if sys.argv[4]=='Fresh':
    get_pdffiles = []
    for pname in doc:
        for boks in boxes:
            if boks['placer_type'] == 'QR Dynamic':
                get_text=[]
                if boks['source'] != '' and boks['source'] != 'Current DateTime':
                    scoords = boks['source_coords'].split(",")
                    srect = fitz.Rect(scoords[0],scoords[1],scoords[2],scoords[3])                
                    get_text.append(pname.getTextbox(srect).replace("\n", ""))
                    extracted_id = get_text[0].replace("/", "").replace("\\", "").replace("-", "").strip()
                    if connection.is_connected():
                        #cursor.execute("select * from individual_records where unique_no = '%s' and publish=1 order by id desc" % (extracted_id))
                        # Update Rohit 01/09/2023
                        if sys.argv[18]=='1':
                            cursor.execute("select * from sb_individual_records where unique_no = '%s' and publish=1 order by id desc" % (extracted_id))
                        else:
                            cursor.execute("select * from individual_records where unique_no = '%s' and publish=1 order by id desc" % (extracted_id))
                        # Update Rohit 01/09/2023
                        records = cursor.fetchall()  
                        row_count = cursor.rowcount
                        if row_count > 0:    
                            get_pdffiles.append(records[0][8])  

    if not get_pdffiles:
        print("No Duplicates")    #List is empty
    else:
        print("Duplicates")
        print(len(get_pdffiles))
        get_pdffiles_string = ','.join(get_pdffiles)
        print(get_pdffiles_string) #Duplicate unique ids
        """
        for unid in get_pdffiles:
            if connection.is_connected():
                sql = "UPDATE individual_records SET publish = 2 WHERE unique_no = '%s'" % (unid)
                cursor.execute(sql)
                connection.commit()
        """
        exit()

  

final_data = {'data':[]}
cnt = 1
pp = pprint.PrettyPrinter(indent=4)

#output_file = directory+sys.argv[7]+'/'+"processed_pdfs/"+sys.argv[2]
#Update Rohit 01/09/2023
# if sys.argv[18] == '1':
#     output_folder = rootDir+sys.argv[7]+'/'+"processed_pdfs/sandbox"
# else:
output_folder = rootDir+sys.argv[7]+'/'+"processed_pdfs"

if not os.path.exists(output_folder):
    os.makedirs(output_folder)


# if sys.argv[18] == '1':
#     output_file = directory+sys.argv[7]+'/'+"processed_pdfs/sandbox/"+sys.argv[2]
# else:
output_file = directory+sys.argv[7]+'/'+"processed_pdfs/"+sys.argv[2]
#Update Rohit 01/09/2023

white = (0, 0, 0, 0)
black = (0, 0, 0, 1)
red = (0, 1, 1, 0)
blue = (1, 1, 0, 0)
pink = (0, 1, 0, 0)
aqua = (1, 0, 0, 0)
green = (1, 0, 1, 0)
purple = (1, 1, 0, 0)
yellow = (0, 0, 1, 0)

#folder=rootDir+sys.argv[7]+'/'+"documents/" + template_name
#Update Rohit 01/09/2023
if sys.argv[18] == '1':
    folder=rootDir+sys.argv[7]+'/'+"documents/sandbox/" + template_name
else:
    folder=rootDir+sys.argv[7]+'/'+"documents/" + template_name
#Update Rohit 01/09/2023
if not os.path.exists(folder):
    os.makedirs(folder)

inner_folder=folder +"/"+str(record_unique_id)
if not os.path.exists(inner_folder):
    os.makedirs(inner_folder)
    
pdf_folder=inner_folder +"/pdfs"
if not os.path.exists(pdf_folder):
    os.makedirs(pdf_folder)
    
path_pdf_moved= inner_folder+"/" +sys.argv[2]  #
#print(path_pdf_moved)
#Create excel file to save unique id and QR details
workbook_name=inner_folder+"/" + template_name+".xlsx"
wbc = openpyxl.Workbook()            
Sheet_name = wbc.sheetnames
wbc.save(filename=workbook_name)
wbs = load_workbook(workbook_name)
page_sheet = wbs.active
page_sheet.column_dimensions['A'].width = 30
page_sheet.column_dimensions['B'].width = 50
page_sheet.column_dimensions['C'].width = 120
wrap_alignment = Alignment(wrap_text=True)
page_sheet.append(['Original ID','Unique ID','QR Details'])  
page_sheet.cell(row = 1, column = 1).font = Font(bold = True) 
page_sheet.cell(row = 1, column = 2).font = Font(bold = True) 
page_sheet.cell(row = 1, column = 3).font = Font(bold = True)
datetime_IND = datetime.datetime.now(tz_IND) 
beginning_time = datetime_IND.strftime("%H:%M:%S")
start_time = time.time()
#convert aws string to array
awsS3InstancesArr = sys.argv[17].split('#')
test_pdffiles=[]
for i in doc:
    page = i
    if not(page._isWrapped):
        page._wrapContents()
    page_data = {cnt:[]}
    words = page.getTextWords()
    page_dict = page.getText('dict')    
    page_rect=page.MediaBox
    #print(page_rect)
    #exit()
    #dirName = directory+sys.argv[7]+'/'+"documents/" + str(sys.argv[1])
    # Update Rohit 01/09/2023
    if sys.argv[18] == '1':
        dirName = directory+sys.argv[7]+'/'+"documents/sandbox/" + str(sys.argv[1])
    else:
        dirName = directory+sys.argv[7]+'/'+"documents/" + str(sys.argv[1])
    # Update Rohit 01/09/2023
        
   # print(dirName)
    mintData={}
    mintData["documentType"] = bc_document_type
    mintData["description"] = bc_document_description
    mcount=1
    use_count=0
    for box_blockchain in extractor_details:
        if "blockchain_flag" in box_blockchain:
            if box_blockchain['blockchain_flag'] == 'use':
                box_blockchain_show_flag = 1
                use_count +=1
            else:
                box_blockchain_show_flag = 0
        else:
            box_blockchain_show_flag = 0  

        if box_blockchain_show_flag==1:
            store_metalabel=box_blockchain['metadata_label']
            store_metavalue=box_blockchain['metadata_value']
            store_coords=box_blockchain['coords']
            meta_value=''
            metadata_label = box_blockchain['metadata_label']
            metadata_value = box_blockchain['metadata_value'].replace("{", "").replace("}", "")
            meta_list=list(filter(bool, metadata_value.splitlines()))
            for mdv in meta_list:
                str_val = mdv.split("^")
                meta_value = meta_value + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)  
            mintData['metadata'+str(mcount)]=json.dumps(dict(label = metadata_label, value = meta_value))
            mcount +=1

    if use_count > 0 and use_count < 5:
        for uc in range(use_count+1, 6):
            mintData['metadata'+str(uc)]=json.dumps(dict({}))

    for box in boxes:
        temp = ''
        otxt = ''
        file_path = ''
        #print()
        if box['source'] == '' or box['source'] == 'Current DateTime':
            srect = fitz.Rect(0,0,0,0)            
        else:
            source_coords = box['source_coords']
            scoords = source_coords.split(",")
            srect = fitz.Rect(scoords[0],scoords[1],scoords[2],scoords[3])        
        
        placer_coords = box['placer_coords']
        pcoords = placer_coords.split(",")
        prect = fitz.Rect(pcoords[0],pcoords[1],pcoords[2],pcoords[3])
        prect_coords = pcoords[0],pcoords[1],pcoords[2],pcoords[3]
        #print(placer_coords)  
        if box['placer_font_name'] == '':  
            placer_font_name = fitz.Font(fontfile=dirFont+"arial.ttf")   
        else:
            placer_font_name = fitz.Font(fontfile=dirFont+box['placer_font_name'])
            
        placer_font_underline = box['placer_font_underline'] 
        placer_font_size = box['placer_font_size']                   
        placer_type = box['placer_type']      
        placer_display = box['placer_display']
        if placer_display == '':
            placer_align = int(0)
        else:
            placer_align = int(placer_display)         
            
        if placer_type == 'Invisible':
            placer_color = yellow
        else:
            placer_color = black
        
        placer_degree_angle = box['degree_angle']
        if placer_degree_angle == '':
            placer_degree_angle = int(0)
        else:
            placer_degree_angle = int(placer_degree_angle)        

        placer_opacity = box['opacity_val']
        if box['line_height'] == '':
            placer_lineHeight = 1
        else:
            placer_lineHeight = box['line_height'] 
        
        if "qr_place" in box:
            if box['qr_place'] == 'show':
                qr_show_flag = 1
            else:
                qr_show_flag = 0
        else:
            qr_show_flag = 1         
        
        if "blockchain_flag" in box:
            if box['blockchain_flag'] == 'use':
                blockchain_show_flag = 1
            else:
                blockchain_show_flag = 0
        else:
            blockchain_show_flag = 0         

        if "barcode_content" in box:
            if box['barcode_content'] == 'Source Content':
                barcode_content_flag = 1
            else:
                barcode_content_flag = 0
        else:
            barcode_content_flag = 0           

        if "barcode_content_position" in box:
            if box['barcode_content_position'] == 'Text at Bottom':
                barcode_content_position_flag = 1
            else:
                barcode_content_position_flag = 0
        else:
            barcode_content_position_flag = 0
        
        if box['font_color'] == '':
            placer_font_color = black
        else:
            placer_font_color = box['font_color']
            placer_font_color=fitz.utils.getColor(placer_font_color) 
        
        if placer_type == 'QR Default': 
            if not os.path.exists(dirName):
                os.makedirs(dirName) 
            now = datetime.datetime.now()            
            dt_string = now.strftime("%Y%m%d%H%M%S")+str(cnt)            
            if bc_contract_address == '':  
                result = hashlib.md5(dt_string.encode("latin-1")) 
                barcode_en=result.hexdigest()    
                qr_txt=barcode_en
            else:
                result = hashlib.md5(dt_string.encode("latin-1")) 
                barcode_en=result.hexdigest()  
                arr = bytes(barcode_en, 'latin-1')    
                encryptedData=base64.b64encode(arr)
                qr_txt="https://"+sys.argv[7]+".seqrdoc.com/bverify/" + encryptedData.decode()

            if qr_show_flag==1:
                qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=4, border=0,)
                qr.add_data(qr_txt)
                qr.make(fit=True)
                img = qr.make_image()  # fill_color="black", back_color="white"
                img.save(dirName+"/"+"qr_"+str(barcode_en)+".png")
                qrcode_file=dirName+"/"+"qr_"+str(barcode_en)+".png"            
                page.insertImage(prect, qrcode_file, overlay=True)   
        elif placer_type == 'QR Dynamic': 
            if not os.path.exists(dirName):
                os.makedirs(dirName) 
            qr_txt=''
            now = datetime.datetime.now()          
            get_text=[]
            if box['source'] == '' or box['source'] == 'Current DateTime':
                #dt_string = now.strftime("%Y-%m-%d %H:%M:%S")+str(cnt)
                dt_string = now.strftime("%Y%m%d%H%M%S")+str(cnt)
                result = hashlib.md5(dt_string.encode('latin-1')) 
                barcode_en=result.hexdigest()
            else:
                #dt_string = otxt.replace("/", "")+now.strftime("%Y-%m-%d %H:%M:%S")+str(cnt)                
                #dt_string = otxt.replace("/", "")
                #dt_string = otxt
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                dt_string = get_text[0].replace("/", "").replace("\\", "").replace("-", "").replace(" ", "").strip()
                result = hashlib.md5(dt_string.encode('latin-1')) 
                barcode_en=result.hexdigest()         

            if bc_contract_address == '':
                barcode_enc=barcode_en
            else:
                arr = bytes(barcode_en, 'latin-1')
                encryptedData=base64.b64encode(arr)
                barcode_enc="https://"+sys.argv[7]+".seqrdoc.com/bverify/" + encryptedData.decode()
                
            qr_details = box['qr_details'].replace("{", "").replace("}", "")
            qr_list=list(filter(bool, qr_details.splitlines()))
            for n in qr_list:
                str_val = n.split("^")
                str_val_count=len(str_val)
                qr_txt = qr_txt + extract_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)
            qr_txt =  qr_txt + "\n"+barcode_enc
            #print(qr_txt)
            if qr_show_flag==1:
                qr = qrcode.QRCode(version=1,error_correction=qrcode.constants.ERROR_CORRECT_L,box_size=4,border=0,)
                qr.add_data(qr_txt)
                qr.make(fit=True)
                img = qr.make_image()  # fill_color="black", back_color="white"
                img.save(dirName+"/"+"qr_"+str(barcode_en)+".png")
                qrcode_file=dirName+"/"+"qr_"+str(barcode_en)+".png"            
                page.insertImage(prect, qrcode_file, overlay=True)
        elif placer_type == 'Barcode': 
            #get_text=[]
            #get_text.append(page.getTextbox(srect).replace("\n", ""))
            #dt_strings = get_text[0].strip()
            if connection.is_connected():
                today = date.today()
                current_year=get_financial_year(str(today))
                current_year='PN/'+current_year+'/'
                #cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
                # Update Rohit 01/09/2023
                if sys.argv[18] == '1':
                    cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM sb_printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
                else:
                    cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
                # Update Rohit 01/09/2023    
                record = cursor.fetchone() 
                next_print=record[0]+1
                next_print_serial=current_year+str(next_print)

            if barcode_content_flag==1:
                get_text=[]
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                temp = get_text[0].replace(" ", "").strip()
            else:
                temp = next_print_serial

            if barcode_content_position_flag==1:
                write_text_flag = True
            else:
                write_text_flag = False

            EAN = barcode.get_barcode_class('code128')
            ean = EAN(temp, writer=ImageWriter())
            if isinstance(prect.width, float)==True:
                bcwidth = float(prect.width)
            else:
                bcwidth = int(prect.width)
            
            if isinstance(prect.height, float)==True:
                bcheight = float(prect.height)
            else:
                bcheight = int(prect.height)
            
            rect = fitz.Rect(0, 0.85*bcheight, bcwidth, bcheight)
            #print(w,h) 300 dpi=667 mil
            options = {
                'dpi': 300,
                'write_text': write_text_flag,
                'module_width': bcwidth/667, #0.40
                'module_height': rect.height, #10
                'quiet_zone': 0,
                'text_distance': 1,
                'text_line_distance': 1,
                'font_size': placer_font_size,
				'center_text':True
            }            
            barcode_file = ean.save(dirName, options = options)               
            #page.insertTextbox(prect, str(rect.width))           
            #page.drawRect( prect, color = green, fill = green)           
            im = Image.open(barcode_file)            
            imwidth, imheight = im.size            
            rectwidth = bcwidth*3.7795275591            
            if imwidth>rectwidth:
                page.insertImage(prect, barcode_file, overlay=True)
            else:
                width_in_mm = imwidth*0.2645833333
                newwidth = rectwidth-imwidth
                right_pos = newwidth*0.2645833333
                left_pos=int(float(pcoords[0]))
                top_pos=int(float(pcoords[1]))
                right_pos=int(float(pcoords[2])) - right_pos
                bottom_pos=int(float(pcoords[3]))               
                prects = fitz.Rect(left_pos,top_pos,right_pos,bottom_pos)          
                page.insertImage(prects, barcode_file, overlay=True)             
        elif placer_type == 'Micro Line': 
            mi_txt=''            
            if box['qr_details']=='':
                get_text=[]
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                #text_len=len(temp)                
                temp = get_text[0].replace(" ", "").strip()
                search_text = get_text[0].strip()
                #print(type(search_text))
                if placer_font_underline=="underline":
                    rl = doc.loadPage(cnt-1).searchFor(search_text, hit_max=1)
                    #print(rl[0])
                    #print(rl)
                    
                    if page_count==1:
                        output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                    else:
                        if len(rl) <=2:
                            output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                        else:
                            output = str(rl[len(rl)-1]).replace("Rect", "").replace("(", "").replace(")", "")                    
                    
                    #output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")                   
                    ucoords = output.split(",")
                    urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                    #print (urect) 
                    #print (urect.height) 
                    #urect.bl, urect.br
                    #prect_coords = pcoords[0],pcoords[1],pcoords[2],pcoords[3]
                    urect_coords = int(float(ucoords[1]))+int(float(urect.height))-1                    
                    new_urect=fitz.Rect(ucoords[0],urect_coords,ucoords[2],ucoords[3])
                    text_len=len(temp)
                    if isinstance(placer_font_size, float)==True:
                        textwidth = fitz.getTextlength(search_text, fontsize=float(placer_font_size))    
                    else:
                        textwidth = fitz.getTextlength(search_text, fontsize=int(placer_font_size))   
                                
                    #print(textwidth)
                    chrPerLine=int(float(new_urect.width))/textwidth
                    Totalchrs=str(chrPerLine).split(".")[0]
                    #print ("chr = ", chrPerLine)
                    #print (Totalchrs)
                    repeat_txt=temp * int(Totalchrs)
                    #print (repeat_txt)
                    if isinstance(placer_font_size, float)==True:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                    else:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                    
                    #print(int(float(new_urect.width)),"|",int(textwidth),"|",int(float(repeat_textwidth)))
                    remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))        
                    #print("rs: ",remain_space,"|",text_len)
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = temp[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''
                     
                    #print(remain_chrs)
                    wr = fitz.TextWriter(page.rect)
                    wr.fillTextbox(new_urect, (temp * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True) 
                    #page.drawRect(new_urect)
                else:
                    text_len=len(temp)
                    if isinstance(placer_font_size, float)==True:
                        textwidth = fitz.getTextlength(temp, fontsize=float(placer_font_size))    
                    else:
                        textwidth = fitz.getTextlength(temp, fontsize=int(placer_font_size))   
                                
                    #print(textwidth)
                    chrPerLine=int(float(prect.width))/textwidth
                    Totalchrs=str(chrPerLine).split(".")[0]
                    #print ("chr = ", chrPerLine)
                    #print (Totalchrs)
                    repeat_txt=temp * int(Totalchrs)
                    if isinstance(placer_font_size, float)==True:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                    else:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                    
                    #print(int(float(prect.width)),"|",int(textwidth),"|",int(float(repeat_textwidth)))
                    remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3        
                    #print("rs: ",remain_space,"|",text_len)
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = temp[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''
                        
                    #print(remain_chrs) # 0 = left, 1 = center, 2 = right
                    #page.insertTextbox(prect, (temp * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)
                    wr = fitz.TextWriter(page.rect)
                    wr.fillTextbox(prect, (temp * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)    
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    #mi_txt = mi_txt + extract_microline_info(extractor_details,inv,words)  
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    mi_txt = mi_txt + extract_microline_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                    
                #print(mi_txt)      
                
                #print(placer_font_name)
                if "Kruti" in box['placer_font_name']:
                    temp=mi_txt.replace(" ", "")
                    get_text=[]
                    get_text.append(page.getTextbox(srect).replace("\n", ""))
                    search_text = get_text[0].strip()
                    text_len=len(temp)                                   
                    text_to_cal=temp                                   
                else:
                    get_text=[]
                    get_text.append(page.getTextbox(srect).replace("\n", ""))
                    temp = get_text[0].replace(" ", "").strip()              
                    search_text = get_text[0].strip()
                    text_len=len(temp)
                    text_to_cal=search_text

                mi_txt = mi_txt.replace(" ", "")
                
                if placer_font_underline=="underline":
                    #rl = page.searchFor(search_text)  
                    rl = doc.loadPage(cnt-1).searchFor(search_text, hit_max=1) 
                    
                    if page_count==1:
                        output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                    else:
                        if len(rl) > 2:
                            output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                        else:
                            output = str(rl[len(rl)-1]).replace("Rect", "").replace("(", "").replace(")", "")
                    
                    #output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")                  
                    ucoords = output.split(",")
                    urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                    #print (urect, prect) 
                    #urect.bl, urect.br
                    #prect_coords = pcoords[0],pcoords[1],pcoords[2],pcoords[3]
                    urect_coords = int(float(ucoords[1]))+int(float(urect.height))-2                    
                    new_urect=fitz.Rect(ucoords[0],urect_coords,ucoords[2],ucoords[3])
                    
                    if isinstance(placer_font_size, float)==True:
                        textwidth = fitz.getTextlength(text_to_cal, fontsize=float(placer_font_size))    
                    else:
                        textwidth = fitz.getTextlength(text_to_cal, fontsize=int(placer_font_size))
                                
                    #print(textwidth,mi_txt)
                    chrPerLine=(int(float(new_urect.width))/textwidth)
                    Totalchrs=str(chrPerLine).split(".")[0]
                    #print ("chr = ", chrPerLine)
                    #print (Totalchrs)                    
                    repeat_txt=mi_txt * int(Totalchrs)
                    if isinstance(placer_font_size, float)==True:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))-10   
                    else:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                    
                    #print(int(float(new_urect.width)),"|",int(textwidth),"|",int(float(repeat_textwidth)))
                    remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))-3
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = mi_txt[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''                    
                    #exit()  
                    
                    wl_lst = fitz.getTextlength((mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size)+1
                    wl = sum([wl_lst])
                    if wl>new_urect.width:      
                        remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))-10     
                    else:       
                        remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))-3
                    
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = mi_txt[0:int(remain_chrs_count)]
                        #print("space|", text_len, remain_space, remain_chrs_count)
                    else:
                        remain_chrs = ''
                    
                    #print((mi_txt * int(Totalchrs))+remain_chrs)
                    wr = fitz.TextWriter(page.rect)
                    if "Kruti" in box['placer_font_name']:
                        wr.fillTextbox(new_urect, (mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align)
                    else:
                        wr.fillTextbox(new_urect, (mi_txt * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True) 
                    #page.drawRect(new_urect)
                else:
                    text_len=len(mi_txt)
                    if isinstance(placer_font_size, float)==True:
                        textwidth = fitz.getTextlength(mi_txt, fontsize=float(placer_font_size))    
                    else:
                        textwidth = fitz.getTextlength(mi_txt, fontsize=int(placer_font_size))   
                                
                    chrPerLine=(int(float(prect.width))/textwidth)
                    Totalchrs=str(chrPerLine).split(".")[0]
                    repeat_txt=mi_txt * int(Totalchrs)
                    if isinstance(placer_font_size, float)==True:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                    else:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                    
                          
                    remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3 
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = mi_txt[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''                    
                    #exit()                    
                    wl_lst = fitz.getTextlength((mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size)+1
                    wl = sum([wl_lst])
                    if wl>prect.width:      
                        remain_space=int(float(prect.width))-int(float(repeat_textwidth))-10     
                    else:       
                        remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3
                    
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = mi_txt[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''
                        
                    #page.insertTextbox(prect, (mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)
                    wr = fitz.TextWriter(page.rect)
                    wr.fillTextbox(prect, (mi_txt * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)
        elif placer_type == 'Invisible':                        
            invisible_font_color=fitz.utils.getColor("YELLOW")
            inv_txt=''
            get_text=[]
            if box['qr_details']=='':  
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                otxt=get_text[0].strip()
                #print(otxt)
                #page.insertTextbox(prect, otxt, fontsize=placer_font_size, color=placer_color, align=placer_align, overlay=True)
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prect, otxt,  font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                #wr.writeText(page, color=placer_color, opacity=float(1.0), overlay=True)
                wr.writeText(page, color=invisible_font_color, opacity=float(1.0), overlay=True) 
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    inv_txt = inv_txt + extract_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                          
                     
                #page.insertTextbox(prect, inv_txt, fontsize=placer_font_size, color=placer_color, align=placer_align, overlay=True)
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prect, inv_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                #wr.writeText(page, color=placer_color, opacity=float(1.0), overlay=True)
                wr.writeText(page, color=invisible_font_color, opacity=float(1.0), overlay=True) 
        elif placer_type == 'Invisible Image': 
            temp_img = [block for block in page_dict['blocks'] if (fitz.Rect(block['bbox']) in srect and block['type'] == 1)]  
            if len(temp_img) > 0:
                pix = fitz.Pixmap(temp_img[0]['image'])
                if not os.path.exists(dirName):
                    os.makedirs(dirName)
                file_path = dirName + "/"+str(int(round(time.time() * 1000))) + "-" + str(cnt) + "." + temp_img[0]['ext']
                file_path2 = dirName + "/"+str(int(round(time.time() * 1000))) + "-" + str(cnt) + ".png"                        
                pix.writeImage(file_path)
                imgs = Image.open(file_path).convert("L") 
                img1 = ImageOps.colorize(imgs, black ="white", white ="yellow")             
                img1.save(file_path2, 'png')
                page.insertImage(prect, file_path2, overlay=True)                       
        elif placer_type == 'Ghost Image': 
            #print(prect.width,"| ",prect.height)
            if placer_degree_angle==90:
                half_width=int(float(prect.height))/2
                top_pos_minus=(int(float(pcoords[0])) - int(float(half_width)))  
                add_right_pos=int(float(pcoords[0])) - top_pos_minus  
                left_pos=int(float(pcoords[0]))-add_right_pos
                top_pos=int(float(pcoords[1]))
                right_pos=int(float(pcoords[2])) + add_right_pos
                bottom_pos=int(float(pcoords[3]))               
                prects = fitz.Rect(left_pos,top_pos,right_pos,bottom_pos) 
                #print(prects)
                #print(prects.width,"| ",prects.height)                         
            else:
                prects=prect         

            get_text=[]
            get_text.append(page.getTextbox(srect).replace("\n", ""))
            temp = ''.join(c for c in get_text[0].strip().replace("/", "").replace(".", "") if c.isalnum())
            ghost_width = round(box['width'] * 3.7795275591)  # Millimeter to Pixel, 1 mm = 3.7795275591 pixel
            ghost_height = round(box['height'] * 3.7795275591)  
            ghost_words = box['ghost_words']  
            PrintableChars=temp[ 0 : ghost_words ] #extract first chars
            #print(ghost_words)
            if not os.path.exists(dirName):
                os.makedirs(dirName)            
            ghostImg=CreateGhostImage(dirName, PrintableChars, placer_font_size, ghost_width, ghost_height)
            page.insertImage(prects, ghostImg,overlay=True, rotate=placer_degree_angle)        
        elif placer_type == 'Image':
            #image_path=rootDir+sys.argv[7]+"/upload_images/" + box['image_path']
            image_path=rootDir+sys.argv[7]+"/backend/templates/pdf2pdf_images/" + box['image_path']
            #print(image_path)
            page.insertImage(prect, image_path,overlay=True)
        elif placer_type == 'Plain Text':           
            inv_txt=''
            get_text=[]
            if box['qr_details']=='':                
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                otxt_string = get_text[0].strip()
                #otxt_string = otxt.replace("/", "")
                #page.insertTextbox(prect, otxt_string, fontsize=placer_font_size, color=placer_font_color, align=placer_align, overlay=True)
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prect, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)                 
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    inv_txt = inv_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                          
                     
                #page.insertTextbox(prect, inv_txt, fontsize=placer_font_size, color=placer_font_color, align=placer_align, overlay=True) 
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prect, inv_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)               
        elif placer_type == 'Static Text':     
            otxt_string = box['qr_details']
            m = fitz.Matrix(0) #placer_degree_angle
            points = fitz.Point(prect.x0, prect.y0)
            wr = fitz.TextWriter(page.rect)
            wr.fillTextbox(prect, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
            wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))
            if placer_font_underline=="underline":
                rl = page.searchFor(otxt_string)  
                output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")                    
                ucoords = output.split(",")
                urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                shape = page.newShape()
                shape.drawLine(urect.bl, urect.br)
                shape.finish(color=placer_font_color, stroke_opacity=float(placer_opacity))
                shape.commit()            
        elif placer_type == 'Watermark Text':           
            inv_txt=''
            get_text=[]
            if box['qr_details']=='':                
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                otxt_string = get_text[0].strip()
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(prect_coords)
                new_rect=ir * m         
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(prect.x0, prect.y0)                 
                wr.fillTextbox(prect, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))                
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    inv_txt = inv_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                          
                    
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(prect_coords)
                new_rect=ir * m         
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(prect.x0, prect.y0)                 
                wr.fillTextbox(prect, inv_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))        
        elif placer_type == 'Watermark Multi Lines':  
            mi_txt=''            
            if box['qr_details']=='':
                get_text=[]
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                text_len=len(temp)
                temp = get_text[0].strip()+' '
                #chrPerLine=int(float(page_rect.width)+int(float(page_rect.height)))
                chrPerLine=int(390)
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=temp * int(Totalchrs)       
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(page_rect)
                new_page_rect=ir * m 
                #new_page_rect.y0=new_page_rect.y0-int(10)
                #new_page_rect.x0=new_page_rect.x0+int(200)
                new_page_rect.x1=new_page_rect.x1+int(100) 
                new_page_rect.y1=new_page_rect.y1+int(100)     
                wr = fitz.TextWriter(page.rect) #wr = fitz.TextWriter(page.rect)
                points = fitz.Point(0, 0)
                wr.fillTextbox(new_page_rect, repeat_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align, lineheight=placer_lineHeight)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=False, morph=(points,m))

                     
                   # wC=textsize(self, 'ABCD', font=None, *args, **kwargs)
                    #print(wC)
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    mi_txt = mi_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                    
                #print(mi_txt)      
                text_len=len(mi_txt)

                
                if isinstance(placer_font_size, float)==True:
                    textwidth = fitz.getTextlength(mi_txt, fontsize=float(placer_font_size))    
                else:
                    textwidth = fitz.getTextlength(mi_txt, fontsize=int(placer_font_size))   
                
                chrPerLine=(int(float(page_rect.width))/textwidth)
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=mi_txt * int(Totalchrs)
                if isinstance(placer_font_size, float)==True:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                else:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                
                      
                remain_space=int(float(page_rect.width))-int(float(repeat_textwidth))-3 
                if remain_space>0:                
                    RchrPerLine=int(remain_space)/textwidth
                    remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                    remain_chrs = mi_txt[0:int(remain_chrs_count)]
                else:
                    remain_chrs = ''                    
                #exit()                    
                wl_lst = fitz.getTextlength((mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size)+1
                wl = sum([wl_lst])

                if placer_lineHeight<3:
                    yOffset=600
                else:
                    yOffset=100

                pageW=float(page_rect.width)+int(60)
                pageH=float(page_rect.height)+int(yOffset)
                areaPage=int(pageW)*int(pageH)
                diagonal = round(math.sqrt((pageW**2) + (pageH**2)), 4)
                font = ImageFont.truetype('arial.ttf', placer_font_size)
                size = font.getsize(mi_txt)
                diagonal=float(diagonal) 
                sizeWInPoints=float(size[0])*float(1.3333)
                charPerSingleLine=int(math.ceil(diagonal/sizeWInPoints))
                sizeHInPoints=float(size[1])*float(1.3333)
                areaPerLine=int(diagonal)* int(math.ceil(sizeHInPoints))
                totalRepeatText=int(math.ceil(areaPage/areaPerLine))
                
                # areaPage=int(float(page_rect.width)*int(float(page_rect.height)))
                # #font = ImageFont.truetype('arial.ttf', placer_font_size)
                # font = ImageFont.truetype('arial.ttf', 10)
                # size = font.getsize(mi_txt)
                # mmWidth= int(math.ceil(float(size[0])*float(0.26458333))) 
                # #charPerSingleLine=int(math.ceil(364/mmWidth))
                # charPerSingleLine=int(math.ceil(wl/mmWidth))
                # areaPerLine=int(mmWidth)*int(charPerSingleLine)*int(size[1])
                # totalRepeatText=int(math.ceil(areaPage/areaPerLine))

                #chrPerLine=int(364)
                #Totalchrs=str(chrPerLine).split(".")[0]
                #repeat_txt=mi_txt * int(Totalchrs)
                no_of_repeat=int(totalRepeatText) * int(charPerSingleLine)
                #no_of_repeat= int(no_of_repeat) + int(10) 
                repeat_txt=mi_txt * int(no_of_repeat)
                repeat_txt=repeat_txt
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(page_rect)
                new_page_rect=ir * m
                new_page_rect.x1=new_page_rect.x1 #+int(100) 
                new_page_rect.y1=new_page_rect.y1 #+int(100) 
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(0, 0)  #prect.x0, prect.y0
                wr.fillTextbox(new_page_rect, repeat_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align, lineheight=placer_lineHeight)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=False, morph=(points,m))
                          
                #if sys.argv[7]=='demo':
                    #print(mi_txt)
                    #print(totalRepeatText)
                    #print(Totalchrs)
                #     font = ImageFont.truetype('arial.ttf', 12)
                #     size = font.getsize(mi_txt)
                #     print(mi_txt)
        """
        if blockchain_show_flag==1 and bc_contract_address != '':
            meta_value=''
            metadata_label = box['metadata_label']
            metadata_value = box['metadata_value'].replace("{", "").replace("}", "")
            meta_list=list(filter(bool, metadata_value.splitlines()))
            for mdv in meta_list:
                str_val = mdv.split("^")
                meta_value = meta_value + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)  
            mintData['metadata'+str(mcount)]=json.dumps(dict(label = metadata_label, value = meta_value))
            mcount +=1
        """            
        #page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
        #page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_REMOVE)
        doc_new = fitz.open()
        doc_new.insertPDF(doc, from_page=cnt-1, to_page=cnt-1)
        if verification_setbg=='Yes':
            for dpage in doc_new:
                dpage.insertImage(page_rect, verification_bg_file,overlay=False)                
        #doc_new.save(pdf_folder+"/"+dt_string+".pdf", garbage=4, deflate=True)
        #doc_new.save(rootDir+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf", garbage=4, deflate=True)
        #Upload file to secdoc directory
        """
        if sys.argv[13] :
            URL = "https://"+sys.argv[7]+"."+sys.argv[10]+"/admin/store-file"
            pdf_path=pdf_folder+"/"+dt_string+".pdf"
            PARAMS = {'pdf_path':pdf_path,'site_id':sys.argv[13]}
            response =requests.get(url = URL, params = PARAMS, timeout=None)        
            response.close()     
        #doc_new.save(pdf_folder+"/"+dt_string+".pdf", garbage=4, deflate=True, owner_pw="owner", encryption=encrypt_meth, permissions=perm) 
        """ 
        
        # doc_new.save(rootDir+"/backend/pdf_file/" + dt_string+".pdf", garbage=4, deflate=True)
        # pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
        # shutil.move(rootDir+"/backend/pdf_file/" + dt_string+".pdf", pdf_file_path)
        # Update Rohit 01/09/2023
        if sys.argv[18] == '1':
            if not os.path.exists(rootDir+"/backend/pdf_file/sandbox"):
                os.makedirs(rootDir+"/backend/pdf_file/sandbox") 

            doc_new.save(rootDir+"/backend/pdf_file/sandbox/" + dt_string+".pdf", garbage=4, deflate=True)
            pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/sandbox/" + dt_string+".pdf"
            shutil.move(rootDir+"/backend/pdf_file/sandbox/" + dt_string+".pdf", pdf_file_path)
        else:
            if not os.path.exists(rootDir+"/backend/pdf_file"):
                os.makedirs(rootDir+"/backend/pdf_file") 
            doc_new.save(rootDir+"/backend/pdf_file/" + dt_string+".pdf", garbage=4, deflate=True)
            pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
            shutil.move(rootDir+"/backend/pdf_file/" + dt_string+".pdf", pdf_file_path)
        # Update Rohit 01/09/2023

        """
        doc_new.save(pdf_folder+"/"+dt_string+".pdf", garbage=4, deflate=True)
        pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
        
        #shutil.move(pdf_folder+"/"+dt_string+".pdf", pdf_file_path)
        arg2= pdf_folder+"/"+dt_string+".pdf"
        arg1= '-sOutputFile='+pdf_file_path
        #on local
        #subprocess.Popen(['C:/Program Files/gs/gs9.53.3/bin/gswin64c.exe', '-sDEVICE=pdfwrite', '-dCompatibilityLevel=1.4', '-dPDFSETTINGS=/ebook', '-dNOPAUSE', '-dBATCH',  '-dQUIET', str(arg1), arg2], stdout=subprocess.PIPE)
        #on server
        subprocess.Popen(['C:/Program Files/gs/gs9.56.1/bin/gswin64c.exe', '-sDEVICE=pdfwrite', '-dCompatibilityLevel=1.4', '-dPDFSETTINGS=/ebook', '-dNOPAUSE', '-dBATCH',  '-dQUIET', str(arg1), arg2], stdout=subprocess.PIPE)
        """
        """ 
        set_background(
            input_pdf=verification_bg_file, 
            output=pdf_folder+"/"+dt_string+".pdf",
            watermark=pdf_folder+"/"+dt_string+".pdf")   
        """       

    # sys.argv[18] == sandboxing param  -- [0 -> off || 1 -> On]
    # Update Rohit 01/09/2023
    #if use_count > 0 and bc_contract_address != '':
    if use_count > 0 and bc_contract_address != '' and sys.argv[18] != '1':      
        pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
        pdf_file_path = pdf_file_path.replace("//", "/")
        pdf_file_path = pdf_file_path.replace("/", "\\")
        #print(pdf_file_path)
        """
        temp_dir=tempfile.gettempdir()
        r = requests.post("https://"+sys.argv[7]+".seqrdoc.com/"+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf", allow_redirects=True)
        open(temp_dir+'\/'+dt_string+'.pdf', 'wb').write(r.content)
        files=[
            ('document',(dt_string+'.pdf',open(temp_dir+'\/'+dt_string+'.pdf','rb'),'application/pdf'))
        ]
        """        
        files=[
            ('document',(dt_string+'.pdf',open(pdf_file_path,'rb'),'application/pdf'))
        ]
        mintData["walletID"] = bc_wallet_address 
        mintData["smartContractAddress"] = bc_contract_address 
        mintData["uniqueHash"] = barcode_en
        mintData["pdf_file"] = "https://"+sys.argv[7]+".seqrdoc.com/"+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
        #mintData["template_id"] = template_id
        blockchain_response = requests.post('https://veraciousapis.herokuapp.com/v1/mint', data=mintData, files=files)
        #print(mintData)
        #print(blockchain_response)
    else:
        blockchain_response = requests.post('https://veraciousapis.herokuapp.com/v1/mint') #"<Response [201]>"
    #print(blockchain_response)

    if print_setbg=='Yes':
        page.insertImage(page_rect, print_bg_file,overlay=False)      
    if connection.is_connected():
        #sql = "INSERT INTO individual_records (file_records_id, template_id, template_name, pdf_page, page_no, encoded_id, unique_no, qr_details, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        # Update Rohit 01/09/2023
        if sys.argv[19]=='0':
            if sys.argv[18]=='1':
                sql = "INSERT INTO sb_individual_records (file_records_id, template_id, template_name, pdf_page, page_no, encoded_id, unique_no, qr_details, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            else:
                sql = "INSERT INTO individual_records (file_records_id, template_id, template_name, pdf_page, page_no, encoded_id, unique_no, qr_details, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            # Update Rohit 01/09/2023
            val = (file_records_next_id, template_id, template_name, pdf_page, cnt, barcode_en, dt_string, qr_txt, userid, record_unique_id)
            cursor.execute(sql, val)         
        
    current_date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S');
    if connection.is_connected():

        #cursor.execute("select * from student_table where serial_no = '%s' and status=1 order by id desc" % (dt_string))
        # Update Rohit 01/09/2023
        if sys.argv[18] == '1':
            cursor.execute("select * from sb_student_table where serial_no = '%s' and status=1 order by id desc" % (dt_string))
        else:
            cursor.execute("select * from student_table where serial_no = '%s' and status=1 order by id desc" % (dt_string))
        # Update Rohit 01/09/2023
        records = cursor.fetchall()  
        row_count = cursor.rowcount
        if row_count > 0: 
            #sql = "UPDATE student_table SET status = %s, updated_by = %s, updated_at = %s WHERE serial_no = %s"
            # Update Rohit 01/09/2023
            if sys.argv[19]=='0':
                if sys.argv[18] == '1':
                    sql = "UPDATE sb_student_table SET status = %s, updated_by = %s, updated_at = %s WHERE serial_no = %s"
                else:
                    sql = "UPDATE student_table SET status = %s, updated_by = %s, updated_at = %s WHERE serial_no = %s"
                # Update Rohit 01/09/2023

                val = ("0", userid, current_date_time, dt_string)
                cursor.execute(sql, val)   

            #cursor.execute ("UPDATE student_table SET status = '%s', updated_by = '%s', updated_at = '%s' WHERE serial_no = '%s' ", ('0', userid, current_date_time, dt_string))

    # site_url='demo.seqrdoc.com'
    # if connection.is_connected():
    #     cursor.execute("SELECT site_id FROM sites WHERE site_url = '%s' " % (site_url))
    #     siteData = cursor.fetchone()
        path="path"
        key="key"
        certificate_filename=dt_string+'.pdf'
        qr_path='qr/'+barcode_en+'.png'
        if blockchain_response.status_code == 200:
            mint_json = json.loads(blockchain_response.text)
            #sql2 = """INSERT INTO student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`, `bc_txn_hash`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            # Update Rohit 01/09/2023
            if sys.argv[19]=='0':
                if sys.argv[18] == '1':
                    sql2 = """INSERT INTO sb_student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`, `bc_txn_hash`,`bc_ipfs_hash`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                else:
                    sql2 = """INSERT INTO student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`, `bc_txn_hash`,`bc_ipfs_hash`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                # Update Rohit 01/09/2023
                val2 = (dt_string, certificate_filename, template_id, barcode_en, qr_path, userid, userid, 1, 1, current_date_time,sys.argv[13],1,mint_json['txnHash'],mint_json['ipfsHash'])
        else:
            
            #sql2 = """INSERT INTO student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            # Update Rohit 01/09/2023
            if sys.argv[19]=='0':
                if sys.argv[18] == '1':
                    sql2 = """INSERT INTO sb_student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                else:
                    sql2 = """INSERT INTO student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                # Update Rohit 01/09/2023    
                val2 = (dt_string, certificate_filename, template_id, barcode_en, qr_path, userid, userid, 1, 1, current_date_time,sys.argv[13],1)
                cursor.execute(sql2, val2)  
                student_table_id = cursor.lastrowid

        if blockchain_response.status_code == 200:
            mint_json = json.loads(blockchain_response.text)
            #print(mint_json['txnHash'], mint_json['tokenID'], mint_json['gasPrice'])
            #sql = "UPDATE student_table SET bc_txn_hash = %s WHERE serial_no = %s AND status = %s"
            # Update Rohit 01/09/2023
            if sys.argv[19]=='0':

                if sys.argv[18] == '1':
                    sql = "UPDATE sb_student_table SET bc_txn_hash = %s WHERE serial_no = %s AND status = %s"
                else:
                    sql = "UPDATE student_table SET bc_txn_hash = %s WHERE serial_no = %s AND status = %s"
                # Update Rohit 01/09/2023
                val = (mint_json['txnHash'], dt_string, '1')
                cursor.execute(sql, val)
            # Update Rohit 01/09/2023
            if sys.argv[19]=='0':
                if sys.argv[18] != '1':
                # Update Rohit 01/09/2023
                    sql2 = """INSERT INTO bc_mint_data (`txn_hash`, `gas_fees`, `token_id`, `key`, `created_at`) VALUES (%s, %s, %s, %s, %s)"""
                    val2 = (mint_json['txnHash'], mint_json['gasPrice'], mint_json['tokenID'], barcode_en, current_date_time)
                    cursor.execute(sql2, val2) 

        cursor.execute("select printer_name from system_config where site_id = '%s'" % (sys.argv[13]))
        recordSystem = cursor.fetchone()  
        printer_name=recordSystem[0]

        # cursor.execute("SELECT * FROM printing_details WHERE sr_no = '%s' " % (dt_string))
        # Update Rohit 01/09/2023
        if sys.argv[18] == '1':
            cursor.execute("SELECT * FROM sb_printing_details WHERE sr_no = '%s' " % (dt_string))
        else:
            cursor.execute("SELECT * FROM printing_details WHERE sr_no = '%s' " % (dt_string))
        # Update Rohit 01/09/2023
        records = cursor.fetchall()  
        print_count = cursor.rowcount

        today = date.today()
        current_year=get_financial_year(str(today))
        current_year='PN/'+current_year+'/'
        #cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
        # Update Rohit 01/09/2023
        if sys.argv[18] == '1':
            cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM sb_printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
        else:
            cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
        # Update Rohit 01/09/2023
        record = cursor.fetchone() 
        next_print=record[0]+1
        next_print_serial=current_year+str(next_print)
        #sql3 = """INSERT INTO printing_details (`username`, `print_datetime`, `printer_name`, `print_count`, `print_serial_no`, `sr_no`, `template_name`,`created_at`, `created_by`, `updated_at`, `updated_by`,`status`, `site_id`, `publish`, `student_table_id`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        # Update Rohit 01/09/2023
        if sys.argv[19]=='0':
            if sys.argv[18] == '1':
                sql3 = """INSERT INTO sb_printing_details (`username`, `print_datetime`, `printer_name`, `print_count`, `print_serial_no`, `sr_no`, `template_name`,`created_at`, `created_by`, `updated_at`, `updated_by`,`status`, `site_id`, `publish`, `student_table_id`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            else:
                sql3 = """INSERT INTO printing_details (`username`, `print_datetime`, `printer_name`, `print_count`, `print_serial_no`, `sr_no`, `template_name`,`created_at`, `created_by`, `updated_at`, `updated_by`,`status`, `site_id`, `publish`, `student_table_id`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            # Update Rohit 01/09/2023
            val3 = (sys.argv[14], current_date_time, printer_name, print_count, next_print_serial, dt_string, template_name, current_date_time, userid, current_date_time,userid, 1, sys.argv[13], 1,student_table_id)
            cursor.execute(sql3, val3) 
    if (sys.argv[7] in awsS3InstancesArr):    
        source_file=pdf_file_path
        test_pdffiles.append(source_file)
    page_sheet.append([dt_string,barcode_en,qr_txt])
    page_sheet.cell(row = cnt+1, column = 1).alignment = Alignment(vertical='top')
    page_sheet.cell(row = cnt+1, column = 2).alignment = Alignment(vertical='top')
    page_sheet.cell(row = cnt+1, column = 3).alignment = Alignment(wrapText=True,vertical='top')
    wbs.save(filename=workbook_name) 
    datetime_IND = datetime.datetime.now(tz_IND)
    ending_time = datetime_IND.strftime("%H:%M:%S")
    end_time=time.time() - start_time
    seconds_to_hhmmss=time.strftime('%H:%M:%S', time.gmtime(end_time))
    page_seconds_to_hhmmss=time.strftime('%H:%M:%S', time.gmtime(end_time/cnt))
    arr_content['percent'] = int(cnt/page_count * 100)
    arr_content['message'] = "Generating "+str(cnt)+"/"+str(page_count)+" PDF(s)"
    arr_content['beginning_time'] = beginning_time
    arr_content['ending_time'] = ending_time
    arr_content['exec_time'] = end_time
    arr_content['hms_time'] = seconds_to_hhmmss
    arr_content['page_time'] = page_seconds_to_hhmmss
    arr_content['pages_processed'] = cnt
    json_object = json.dumps(arr_content, indent = 4)

    #f = open(rootDir+sys.argv[7]+'/'+"processed_pdfs/"+sys.argv[5], "w")
    # Update Rohit 01/09/2023
    # if sys.argv[18] == '1':
    #     f = open(rootDir+sys.argv[7]+'/'+"processed_pdfs/sandbox/"+sys.argv[5], "w")
    # else:
    f = open(rootDir+sys.argv[7]+'/'+"processed_pdfs/"+sys.argv[5], "w")
    # Update Rohit 01/09/2023
        
    f.write(json_object)
    #time.sleep(1)
    cnt += 1

if connection.is_connected():
    #sql = "INSERT INTO file_records (template_id, template_name, pdf_page, total_records, source_file, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    # Update Rohit 01/09/2023
    if sys.argv[19]=='0':
        if sys.argv[18]=='1':
            sql = "INSERT INTO sb_file_records (template_id, template_name, pdf_page, total_records, source_file, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s)"
        else:
            sql = "INSERT INTO file_records (template_id, template_name, pdf_page, total_records, source_file, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s)"
        # Update Rohit 01/09/2023
        val = (template_id, template_name, pdf_page, cnt-1, sys.argv[2], userid, record_unique_id)
        cursor.execute(sql, val)

    connection.commit()
print(path_pdf_moved)

#Upload pdf to AWS s3
if (sys.argv[7] in awsS3InstancesArr):
            URL = "https://"+sys.argv[7]+"."+sys.argv[10]+"/admin/upload-file-s3"
            test_pdffilesStr = ', '.join(test_pdffiles)
            PARAMS = {'pdf_paths':test_pdffilesStr}
            response =requests.get(url = URL, params = PARAMS, timeout=None)        
            response.close()  
            print("mmk"+test_pdffilesStr)

removePath=rootDir+sys.argv[7]+'/'
#print(removePath)
rewisePath = path_pdf_moved.replace(removePath,'')

doc.save(path_pdf_moved, garbage=4, deflate=True)
#doc.save(path_pdf_moved, garbage=4, deflate=True, owner_pw="owner", encryption=encrypt_meth, permissions=perm)
#doc.save(path_pdf_moved, incremental=True)
#shutil.copyfile(output_file, path_pdf_moved)
print(rewisePath)
#shutil.make_archive(folder+"/"+template_name+"-"+str(record_unique_id), "zip", inner_folder)
#print(folder+"/"+template_name+"-"+str(record_unique_id)+".zip")
total_pages=cnt-1
#print("Total Records:"+str(total_pages))
shutil.rmtree(dirName, ignore_errors=True)
#shutil.rmtree(inner_folder, ignore_errors=True)
#print("documents/"+template_name+"/"+template_name+"-"+str(record_unique_id)+".zip")
#print(pdf_folder+"/"+dt_string+".pdf");
#print(rewisePath)
#Upload file to secdoc directory
"""
if sys.argv[13] :
    URL = "https://"+sys.argv[7]+"."+sys.argv[10]+"/admin/store-file"
    pdf_path=pdf_folder+"/"+dt_string+".pdf"
    PARAMS = {'pdf_path':pdf_path,'site_id':sys.argv[13]}
    response =requests.get(url = URL, params = PARAMS)
"""


