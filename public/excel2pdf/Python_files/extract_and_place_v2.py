#!D:/omkar_python/QR/venv/Scripts/python.exe
import os.path
import subprocess
import sys
sys.path.append("F:\\pdf2pdf_env\\Lib\\site-packages")
import os
import shutil
import simplejson as json
import fitz
import mysql.connector
import time
from mysql.connector import Error
from PIL import Image, ImageOps 
import barcode
from barcode.writer import ImageWriter
import qrcode
import hashlib 
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font 
from openpyxl.styles import Alignment
import uuid
import pytz
from PyPDF2 import PdfFileWriter, PdfFileReader
import datetime
from datetime import date
from PIL import ImageFont
import math
import base64
import requests

'''
#sys.argv[1] = template id
#sys.argv[2] = data file
#sys.argv[3] = session user id
#sys.argv[4] = entry type (Fresh/Proceed)
#sys.argv[5] = progress file
#sys.argv[6] = dbName
#sys.argv[7] = subdomain
#sys.argv[8] = directoryUrlForward
#sys.argv[9] = directoryUrlBackward
#sys.argv[10] = servername
#sys.argv[11] = username
#sys.argv[12] = password
#sys.argv[13] = siteid
#sys.argv[14] = username
#sys.argv[15] = printer_name
# python extract_and_place.py 1 FYBAF_1619953609.pdf 1
#print(sys.argv[2])
'''

try:
    directory=sys.argv[8]
    rootDir= directory.replace('pdf2pdf', '')
    connection = mysql.connector.connect(host=sys.argv[10], database=sys.argv[6], user=sys.argv[11], password=sys.argv[12])

    if connection.is_connected():
        db_Info = connection.get_server_info()
        cursor = connection.cursor(buffered=True)
        cursor.execute("select ep_details, id, extractor_details, template_name, pdf_page, print_bg_file, print_bg_status, verification_bg_file, verification_bg_status, IFNULL(bc_contract_address, '') AS bc_contract_address, bc_document_description, bc_document_type from uploaded_pdfs where id = '%s'" % (sys.argv[1]))
        record = cursor.fetchone()        
        boxes = json.loads(record[0], strict=False)        
        template_id=record[1]
        extractor_details = json.loads(record[2], strict=False)
        template_name=record[3]
        pdf_page=record[4]
        pbg_file=record[5]
        print_bg_status=record[6]
        vbg_file=record[7]
        verification_bg_status=record[8]

        # ---- 23 feb 2024
        bc_contract_address=record[9]
        bc_document_description=record[10]
        bc_document_type=record[11]
        # ---- 23 feb 2024

        cur=connection.cursor(buffered=True)
        print_setbg=''
        verification_setbg=''

        if pbg_file !=0 and print_bg_status == 'Yes':
            sql_bg="SELECT image_path FROM background_template_master where id= '%s'" % (pbg_file)
            cur.execute(sql_bg)
            precord = cur.fetchone()
            #print_bg_file=sys.argv[8]+"upload_bgs/"+precord[0]
            print_bg_file=rootDir+sys.argv[7]+"/backend/canvas/bg_images/" + precord[0]
            #print_bg_file="C:/wamp/www/demo/upload_bgs/"+precord[0] 
            print_setbg='Yes'

        if vbg_file !=0 and verification_bg_status == 'Yes':
            sql_bg="SELECT image_path FROM background_template_master where id= '%s'" % (vbg_file)
            cur.execute(sql_bg)
            vrecord = cur.fetchone()
            #verification_bg_file=sys.argv[8]+"upload_bgs/"+vrecord[0]
            verification_bg_file=rootDir+sys.argv[7]+"/backend/canvas/bg_images/" + vrecord[0]
            #verification_bg_file="C:/wamp/www/demo/upload_bgs/"+vrecord[0]
            verification_setbg='Yes'

        record_unique_id = datetime.datetime.now().strftime('%Y%m%d%H%M%S-') + str(uuid.uuid4()).split('-')[-1]

        # --------- start 23 feb
        # sqli will return the last id(latest) from the table
        if sys.argv[18]:
            if sys.argv[18] == '1':
                sqli="SELECT id FROM sb_file_records ORDER BY id DESC LIMIT 1"
            else:
                sqli="SELECT id FROM file_records ORDER BY id DESC LIMIT 1"
            # --------- end 23 feb
            
            cur.execute(sqli)
            last_id = cur.fetchone()
            
            if last_id == None:
                file_records_next_id=1
            else:
                file_records_next_id=last_id[0]+1     
        

except Error as e:
    print("Error while connecting to MySQL", e)


# colors
white = (0, 0, 0, 0)
black = (0, 0, 0, 1)
red = (0, 1, 1, 0)
blue = (1, 1, 0, 0)
pink = (0, 1, 0, 0)
aqua = (1, 0, 0, 0)
green = (1, 0, 1, 0)
purple = (1, 1, 0, 0)
yellow = (0, 0, 1, 0)

is_compressed_enabled = 1
userid=sys.argv[3]
tz_IND = pytz.timezone('Asia/Calcutta') 

encrypt_meth = fitz.PDF_ENCRYPT_AES_256  # strongest algorithm
perm = int(fitz.PDF_PERM_PRINT)

dirFont = sys.argv[8]+"Python_files/fonts/"


folder=rootDir+sys.argv[7]+"/documents/" + template_name
inner_folder=folder +"/"+str(record_unique_id)
    
pdf_folder=inner_folder +"/pdfs"
if not os.path.exists(pdf_folder):
    os.makedirs(pdf_folder)

preview_folder = rootDir+"/processed_pdfs/preview"
if not os.path.exists(preview_folder):
	os.makedirs(preview_folder)

# split a word in a char and return in list for eg split("test") => ['t', 'e', 's', 't']
def split(word): 
    return [char for char in word]

# based on the search_value parameter, this func will extract coords from extractor_detail col from uploaded_pdfs and based on coords return the plaintext from given pdf. 
# plaintext eg studentname, serialno 
def extract_microline_info(extractor_detail,search_value,extra_line_first='',extra_line_second='',page=''):
    for keyval in extractor_detail:
        if search_value.lower() == keyval['name'].lower():
            source_coord = keyval['coords']
            scoord = source_coord.split(",")
            s_rect = fitz.Rect(scoord[0],scoord[1],scoord[2],scoord[3])
            get_text=[]            
            get_text.append(page.getTextbox(s_rect).replace("\n", ""))
            otxt=get_text[0].replace(" ", "")

            if extra_line_first == '' and extra_line_second == '':
                return otxt
            elif extra_line_first != '' and extra_line_second != '':
                return extra_line_first + otxt + extra_line_second 
            elif extra_line_first != '' and extra_line_second == '':
                return extra_line_first + otxt 
            elif extra_line_first == '' and extra_line_second != '':
                return otxt + extra_line_second

# based on the search_value parameter, this func will extract coords from extractor_detail col from uploaded_pdfs and based on coords return the plaintext from given pdf. 
# plaintext eg studentname + " ", serialno + " " 
def extract_plainText(extractor_detail,search_value,extra_line_first='',extra_line_second='',page=''):
    # extractor_detail is a list of dict
    for keyval in extractor_detail:  # here we iterate over each dict
        if search_value.lower() == keyval['name'].lower():  # check the search_value in a dict, if found then extract text from coords
            source_coord = keyval['coords']
            scoord = source_coord.split(",")
            s_rect = fitz.Rect(scoord[0],scoord[1],scoord[2],scoord[3]) 
            get_text=[]            
            get_text.append(page.getTextbox(s_rect).replace("\n", ""))
            otxt=get_text[0].strip()

            if extra_line_first == '' and extra_line_second == '':
                return otxt +" "
            elif extra_line_first != '' and extra_line_second != '':
                return extra_line_first + otxt + extra_line_second +" " 
            elif extra_line_first != '' and extra_line_second == '':
                return extra_line_first + otxt +" " 
            elif extra_line_first == '' and extra_line_second != '':
                return otxt + extra_line_second +" "                 
                
# based on the search_value parameter, this func will extract coords from extractor_detail col from uploaded_pdfs and based on coords return the plaintext from given pdf. 
# plaintext eg studentname + "\n", serialno + "\n" 
def extract_info(extractor_detail,search_value,extra_line_first='',extra_line_second='',page=''):
    for keyval in extractor_detail:
        if search_value.lower() == keyval['name'].lower():
            source_coord = keyval['coords']
            scoord = source_coord.split(",")
            s_rect = fitz.Rect(scoord[0],scoord[1],scoord[2],scoord[3])
            get_text=[]            
            get_text.append(page.getTextbox(s_rect).replace("\n", ""))
            otxt=get_text[0].strip()
            if extra_line_first == '' and extra_line_second == '':
                return otxt +"\n"
            elif extra_line_first != '' and extra_line_second != '':
                return extra_line_first + otxt + extra_line_second +"\n" 
            elif extra_line_first != '' and extra_line_second == '':
                return extra_line_first + otxt +"\n" 
            elif extra_line_first == '' and extra_line_second != '':
                return otxt + extra_line_second +"\n" 

# this func take the name parameter, then split it into single chars, then based on chars select images from dir of ghostimage and then combine them
def CreateGhostImage(dirName, name, p_font_size, ghost_width, ghost_height):  
    # D:/wamp64/www/demo/public/pdf2pdf/secura/documents/3 PawarB 12 529 87
    # print(rootDir)

    dirChars = rootDir+"pdf2pdf/Python_files/chars/"+str(p_font_size)                           
    name=name.upper()    
    single_char=split(name)                                                           
    
    my_list = list()                                                                  
    for c in single_char:
        my_list.append(dirChars +"/"+ c +".png")    

    images = [Image.open(x) for x in my_list]                                         
    widths, heights = zip(*(i.size for i in images))                  # (1167, 1167, 1167, 1167, 1167, 1167) (1129, 1129, 1129, 1129, 1129, 1129)           
    total_width = sum(widths)                                         # calculate total width by combining all img                  
    max_height = max(heights)                                                         
    new_im = Image.new('RGB', (total_width, max_height))                              
    x_offset = 0

    for im in images:
      new_im.paste(im, (x_offset,0))
      x_offset += im.size[0]                                           # placed each char image one beside other

    new_im.save(dirName +"/"+ name +".png")  
    isize=ghost_width,ghost_height            
    im = Image.open(dirName +"/"+ name +".png")
    im.thumbnail((isize), Image.ANTIALIAS)
    im.save(dirName +"/"+ name +str(p_font_size)+"_th.png", quality=100)    
    return dirName +"/"+ name +str(p_font_size)+"_th.png"  

def get_financial_year(datestring):
    date = datetime.datetime.strptime(datestring, "%Y-%m-%d").date()
    year_of_date=date.year
    financial_year_start_date = datetime.datetime.strptime(str(year_of_date)+"-04-01","%Y-%m-%d").date()

    if date<financial_year_start_date:
        return str(financial_year_start_date.year-1)[2:]+'-'+ str(financial_year_start_date.year)[2:]
    else:
        return str(financial_year_start_date.year)[2:]+'-'+ str(financial_year_start_date.year+1)[2:]    

def repeat_to_length(string_to_expand, length):
    return (string_to_expand * (int(length/len(string_to_expand))+1))[:length]

def get_pil_text_size(text, font_size, font_name):
    font = ImageFont.truetype(font_name, font_size)
    size = font.getsize(text)
    return size

def set_background(input_pdf, output, watermark):
    watermark_obj = PdfFileReader(watermark)
    watermark_page = watermark_obj.getPage(0)

    pdf_reader = PdfFileReader(input_pdf)    
    pdf_writer = PdfFileWriter()

    for page in range(pdf_reader.getNumPages()):        
        page = pdf_reader.getPage(page)
        page.mergePage(watermark_page)
        pdf_writer.addPage(page)

    with open(output, 'wb') as out:
        pdf_writer.write(out)

def compress(input_file_path, output_file_path):
    """Function to compress PDF via Ghostscript command line interface"""

    dpi=120
    gs_path = r"E:\wamp64\www\uneb\public\pdf2pdf\Python_files\bin\gswin64.exe"

    startupinfo = None

    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    
    process = subprocess.Popen(
        [
            gs_path,
            "-sDEVICE=pdfwrite",
            "-dCompatibilityLevel=1.4",
            "-dDownsampleColorImages=true",
            "-dDownsampleGrayImages=true",
            "-dDownsampleMonoImages=true",
            f"-dColorImageResolution={dpi}",
            f"-dGrayImageResolution={dpi}",
            f"-dMonoImageResolution={dpi}",
            "-r{}".format(dpi),
            "-dNOPAUSE",
            "-dBATCH",
            "-sOutputFile={}".format(output_file_path),
            input_file_path,
        ],
        shell=False,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        startupinfo=startupinfo,
    )

    try:
        output, error = process.communicate()
    except Exception as e:
        print(e)

    if process.returncode != 0:
        print("Error during compression:")
        print(error.decode('utf-8'))

class SecurityFeatures():

    def extract_from_coords_else_exit(doc):
        '''
          It extract the source coords from extractor_details.
          Based on the coords, it extract the text. If it won't find the text in those coords, it will stop the process 
          at pre stage before applying the features and returns an error.
 
        '''   
        get_blank = []
        get_extractor_name = []
        pno=0

        for pagechk in doc:
            pno += 1
            get_extractor_name.append('<br><b>Page '+str(pno)+':</b>')
            for box_chk in extractor_details:
                source_coords_chk = box_chk['coords']
                scoords = source_coords_chk.split(",")
                srect = fitz.Rect(scoords[0],scoords[1],scoords[2],scoords[3])
                blocks = pagechk.getText("blocks", srect)
                blocks.sort(key=lambda block: block[0])  # sort vertically ascending
                try:
                    block_split=blocks[0]
                except IndexError:
                    block_split=''
                    extractor_name=box_chk['name']
                
                if block_split=='':
                    get_blank.append(extractor_name)
                    get_extractor_name.append(extractor_name+', ')

                
        if get_blank:
            print("Empty Extractor")
            get_blank_string = ' '.join(get_extractor_name)
            print('Empty Extractor'+get_blank_string) #Blank Source
            exit()

    def qr_default(qr_show_flag, page, prect, cnt, dirName):
        now = datetime.datetime.now()            
        dt_string = now.strftime("%Y%m%d%H%M%S") + str(cnt)

        result = hashlib.md5(dt_string.encode("latin-1")) 
        barcode_en=result.hexdigest()           
        # qr_txt=barcode_en

        # ----23 FEB start
        if bc_contract_address == '':      
            qr_txt=barcode_en
        else:  
            arr = bytes(barcode_en, 'latin-1')    
            encryptedData=base64.b64encode(arr)
            qr_txt="https://"+sys.argv[7]+".seqrdoclocal.com/bverify/" + encryptedData.decode() + "\n\n" + barcode_en

        # ----23 FEB END

        if qr_show_flag==1:
            qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=4, border=0)
            qr.add_data(qr_txt)
            qr.make(fit=True)
            img = qr.make_image()  # fill_color="black", back_color="white"
            img.save(dirName+"/"+"qr_"+str(barcode_en)+".png")
            qrcode_file=dirName+"/"+"qr_"+str(barcode_en)+".png"            
            page.insertImage(prect, qrcode_file, overlay=True)        

        return barcode_en, qr_txt

    def qr_dynamic(box, page, qr_show_flag, srect, prect, cnt, dirName):
        '''extract text from coord then apply enc, then create qr and save as image, then place the img on placer coords'''
        qr_txt=''
        now = datetime.datetime.now()          
        get_text=[]

        if box['source'] == '' or box['source'] == 'Current DateTime':
            dt_string = now.strftime("%Y%m%d%H%M%S")+str(cnt)
        else:
            get_text.append(page.getTextbox(srect).replace("\n", ""))
            dt_string = get_text[0].replace("/", "").replace("\\", "").replace("-", "").replace(" ", "").strip()
        
        result = hashlib.md5(dt_string.encode('latin-1')) 
        barcode_en=result.hexdigest()    

        # ---- 23 FEB start
        if bc_contract_address == '':
            barcode_enc=barcode_en
        else:
            arr = bytes(barcode_en, 'latin-1')
            encryptedData=base64.b64encode(arr)
            barcode_enc="https://"+sys.argv[7]+".seqrdoclocal.com/bverify/" + encryptedData.decode() + "\n\n" + barcode_en
        # ----23 FEB END
            
        qr_details = box['qr_details'].replace("{", "").replace("}", "")
        qr_list=list(filter(bool, qr_details.splitlines()))

        for n in qr_list:
            str_val = n.split("^")
            qr_txt = qr_txt + extract_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)
        qr_txt =  qr_txt + "\n"+barcode_enc

        if qr_show_flag==1:
            qr = qrcode.QRCode(version=1,error_correction=qrcode.constants.ERROR_CORRECT_L,box_size=4,border=0,)
            qr.add_data(qr_txt)
            qr.make(fit=True)
            img = qr.make_image()  # fill_color="black", back_color="white"
            img.save(dirName+"/"+"qr_"+str(barcode_en)+".png")
            qrcode_file=dirName+"/"+"qr_"+str(barcode_en)+".png"            
            page.insertImage(prect, qrcode_file, overlay=True)

        return barcode_en, qr_txt, dt_string

    def place_barcode(barcode_content_flag, barcode_content_position_flag, page, srect, prect, placer_font_size, pcoords, dirName):
        # Based on the current year it creates the barcode_serial no
        if connection.is_connected():
            today = date.today()                                               
            current_year=get_financial_year(str(today))                       
            current_year='PN/'+current_year+'/'                               
            
            # return us the latest serial no that has been used, so we add +1 in it and used it for current pdf
            # ----23 FEB start
            if sys.argv[18] == '1':
                cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM sb_printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
            else:
                cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
            # ----23 FEB END

            record = cursor.fetchone() 
            next_print=record[0]+1                                             
            next_print_serial=current_year+str(next_print)                     # PN/23-24/331
        
        if barcode_content_flag==1:
            get_text=[]
            get_text.append(page.getTextbox(srect).replace("\n", ""))
            temp = get_text[0].replace(" ", "").strip()
        else:
            temp = next_print_serial

        write_text_flag = True if barcode_content_position_flag==1 else False

        EAN = barcode.get_barcode_class('code128')
        ean = EAN(temp, writer=ImageWriter())

        bcwidth = float(prect.width) if isinstance(prect.width, float)==True else int(prect.width)
        bcheight = float(prect.height) if isinstance(prect.height, float)==True else int(prect.height)
        rect = fitz.Rect(0, 0.85*bcheight, bcwidth, bcheight)
        
        options = {
            'dpi': 300,
            'write_text': write_text_flag,
            'module_width': bcwidth/667, 
            'module_height': rect.height,
            'quiet_zone': 0,
            'text_distance': 1,
            'text_line_distance': 1,
            'font_size': placer_font_size,
            'center_text':True
        }            
        barcode_file = ean.save(dirName, options = options)               
        im = Image.open(barcode_file)            
        imwidth, imheight = im.size            
        rectwidth = bcwidth*3.7795275591   

        if imwidth>rectwidth:
            page.insertImage(prect, barcode_file, overlay=True)
        else:
            width_in_mm = imwidth*0.2645833333
            newwidth = rectwidth-imwidth
            right_pos = newwidth*0.2645833333

            left_pos=int(float(pcoords[0]))
            top_pos=int(float(pcoords[1]))
            right_pos=int(float(pcoords[2])) - right_pos
            bottom_pos=int(float(pcoords[3]))           
            prects = fitz.Rect(left_pos,top_pos,right_pos,bottom_pos)  
            page.insertImage(prects, barcode_file, overlay=True)      

    def plain_text(box, page, srect, prect, placer_font_name, placer_font_size, placer_align, placer_font_color):
        '''extract name and place it above the barcode'''
        inv_txt=''
        get_text=[]
        if box['qr_details']=='':                
            get_text.append(page.getTextbox(srect).replace("\n", ""))
            otxt_string = get_text[0].strip()                              
            txt = otxt_string           
        else:
            qr_details = box['qr_details'].replace("{", "").replace("}", "")
            qr_list=list(filter(bool, qr_details.splitlines()))
            for inv in qr_list:
                str_val = inv.split("^")
                inv_txt = inv_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                          
            txt = inv_txt
        wr = fitz.TextWriter(page.rect)
        wr.fillTextbox(prect, txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
        wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)  

    def placer_invisible_text(box, page, srect, prect, placer_font_name, placer_font_size, placer_align):
        ''' Based on the qr_details it will add the invisible text'''
        invisible_font_color=fitz.utils.getColor("YELLOW")
        inv_txt=''
        get_text=[]
        if box['qr_details']=='':  
            get_text.append(page.getTextbox(srect).replace("\n", ""))
            otxt=get_text[0].strip()
            txt = otxt
        else:
            qr_details = box['qr_details'].replace("{", "").replace("}", "")
            qr_list=list(filter(bool, qr_details.splitlines()))
            for inv in qr_list:
                str_val = inv.split("^")
                str_val_count=len(str_val)
                inv_txt = inv_txt + extract_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)   
            txt = inv_txt                    
                
        wr = fitz.TextWriter(page.rect)
        wr.fillTextbox(prect, txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
        wr.writeText(page, color=invisible_font_color, opacity=float(1.0), overlay=True) 

    def place_ghost_image(placer_degree_angle, srect, prect, pcoords, page, box, placer_font_size, dirName):
        if placer_degree_angle==90:
            half_width=int(float(prect.height))/2
            top_pos_minus=(int(float(pcoords[0])) - int(float(half_width)))  
            add_right_pos=int(float(pcoords[0])) - top_pos_minus  
            left_pos=int(float(pcoords[0]))-add_right_pos
            top_pos=int(float(pcoords[1]))
            right_pos=int(float(pcoords[2])) + add_right_pos
            bottom_pos=int(float(pcoords[3]))               
            prects = fitz.Rect(left_pos,top_pos,right_pos,bottom_pos)                          
        else:
            prects=prect         

        get_text=[]
        get_text.append(page.getTextbox(srect).replace("\n", ""))                                          
        temp = ''.join(c for c in get_text[0].strip().replace("/", "").replace(".", "") if c.isalnum())    
        ghost_width = round(box['width'] * 3.7795275591)   # Millimeter to Pixel, 1 mm = 3.7795275591 pixel
        ghost_height = round(box['height'] * 3.7795275591)                                                 
        ghost_words = box['ghost_words']                                                                   
        PrintableChars=temp[ 0 : ghost_words ] #extract first chars
       
        ghostImg=CreateGhostImage(dirName, PrintableChars, placer_font_size, ghost_width, ghost_height)
        page.insertImage(prects, ghostImg,overlay=True, rotate=placer_degree_angle) 
                    
    def place_watermark_text(box, page, srect, prect, placer_degree_angle, placer_font_name, placer_font_size, placer_align, placer_font_color, placer_opacity):
        inv_txt=''
        get_text=[]

        if box['qr_details']=='':                
            get_text.append(page.getTextbox(srect).replace("\n", ""))
            otxt_string = get_text[0].strip()
            txt = otxt_string                
        else:
            qr_details = box['qr_details'].replace("{", "").replace("}", "")
            qr_list=list(filter(bool, qr_details.splitlines()))
            for inv in qr_list:
                str_val = inv.split("^")
                inv_txt = inv_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                          
            txt = inv_txt

        m = fitz.Matrix(placer_degree_angle)
        wr = fitz.TextWriter(page.rect)
        points = fitz.Point(prect.x0, prect.y0)                 
        wr.fillTextbox(prect, txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
        wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))        

    def place_watermark_multi_lines(box, srect, page, placer_degree_angle, page_rect, placer_font_name, placer_font_size, placer_align, placer_lineHeight, placer_font_color, placer_opacity):
        mi_txt=''            
        if box['qr_details']=='':
            get_text=[]
            get_text.append(page.getTextbox(srect).replace("\n", ""))
            temp = get_text[0].strip()+' '
            chrPerLine=int(390)
            Totalchrs=str(chrPerLine).split(".")[0]
            repeat_txt=temp * int(Totalchrs)       
            m = fitz.Matrix(placer_degree_angle)
            ir = fitz.IRect(page_rect)
            new_page_rect=ir * m 
            new_page_rect.x1=new_page_rect.x1+int(100) 
            new_page_rect.y1=new_page_rect.y1+int(100)     
            
        else:
            # Based on the set extractor coords it will extract serial_no, Student_Name  and Degree
            qr_details = box['qr_details'].replace("{", "").replace("}", "")
            qr_list=list(filter(bool, qr_details.splitlines()))
            for inv in qr_list:
                str_val = inv.split("^")
                mi_txt = mi_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                    
            text_len=len(mi_txt)
            
            if isinstance(placer_font_size, float)==True:
                textwidth = fitz.getTextlength(mi_txt, fontsize=float(placer_font_size))    
            else:
                textwidth = fitz.getTextlength(mi_txt, fontsize=int(placer_font_size))   

            
            chrPerLine=(int(float(page_rect.width))/textwidth)
            Totalchrs=str(chrPerLine).split(".")[0]
            repeat_txt=mi_txt * int(Totalchrs)

            if isinstance(placer_font_size, float)==True:
                repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
            else:
                repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
            
                
            remain_space=int(float(page_rect.width))-int(float(repeat_textwidth))-3 
            if remain_space>0:                
                RchrPerLine=int(remain_space)/textwidth
                remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                remain_chrs = mi_txt[0:int(remain_chrs_count)]
            else:
                remain_chrs = ''                    
                  
            wl_lst = fitz.getTextlength((mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size)+1
            wl = sum([wl_lst])

            yOffset = 600 if placer_lineHeight<3 else 100
                
            # it calculate all the page height, width, space between each line, what font to use, size and all. Place the watermark lines
            pageW=float(page_rect.width)+int(60)
            pageH=float(page_rect.height)+int(yOffset)
            areaPage=int(pageW)*int(pageH)
            diagonal = round(math.sqrt((pageW**2) + (pageH**2)), 4)
            font = ImageFont.truetype('arial.ttf', placer_font_size)
            size = font.getsize(mi_txt)
            
            diagonal=float(diagonal) 
            sizeWInPoints=float(size[0])*float(1.3333)
            charPerSingleLine=int(math.ceil(diagonal/sizeWInPoints))
            sizeHInPoints=float(size[1])*float(1.3333)
            areaPerLine=int(diagonal)* int(math.ceil(sizeHInPoints))
            totalRepeatText=int(math.ceil(areaPage/areaPerLine))

            no_of_repeat=int(totalRepeatText) * int(charPerSingleLine)
            repeat_txt=mi_txt * int(no_of_repeat)
            m = fitz.Matrix(placer_degree_angle)
            ir = fitz.IRect(page_rect)
            new_page_rect=ir * m
            new_page_rect.x1=new_page_rect.x1 
            new_page_rect.y1=new_page_rect.y1 

        wr = fitz.TextWriter(page.rect)
        points = fitz.Point(0, 0)  
        wr.fillTextbox(new_page_rect, repeat_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align, lineheight=placer_lineHeight)
        wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=False, morph=(points,m))

    def place_micro_line(box, page, srect, prect, placer_font_underline, cnt, placer_font_size, placer_font_name, placer_align, placer_font_color, page_count, doc):
        mi_txt=''            

        if box['qr_details']=='':
            get_text=[]
            get_text.append(page.getTextbox(srect).replace("\n", ""))
            temp = get_text[0].replace(" ", "").strip()
            search_text = get_text[0].strip()

            if placer_font_underline=="underline":
                rl = doc.loadPage(cnt-1).searchFor(search_text, hit_max=1)

                if page_count==1:
                    output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                else:
                    if len(rl) <=2:
                        output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                    else:
                        output = str(rl[len(rl)-1]).replace("Rect", "").replace("(", "").replace(")", "")                    

                ucoords = output.split(",")
                urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                urect_coords = int(float(ucoords[1]))+int(float(urect.height))-1                    
                new_urect=fitz.Rect(ucoords[0],urect_coords,ucoords[2],ucoords[3])
                text_len=len(temp)

                if isinstance(placer_font_size, float)==True:
                    textwidth = fitz.getTextlength(search_text, fontsize=float(placer_font_size))    
                else:
                    textwidth = fitz.getTextlength(search_text, fontsize=int(placer_font_size))   
                            
                chrPerLine=int(float(new_urect.width))/textwidth
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=temp * int(Totalchrs)
                
                if isinstance(placer_font_size, float)==True:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                else:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                
                remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))        
                if remain_space>0:                
                    RchrPerLine=int(remain_space)/textwidth
                    remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                    remain_chrs = temp[0:int(remain_chrs_count)]
                else:
                    remain_chrs = ''
                
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(new_urect, (temp * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True) 
                
            else:
                text_len=len(temp)

                if isinstance(placer_font_size, float)==True:
                    textwidth = fitz.getTextlength(temp, fontsize=float(placer_font_size))    
                else:
                    textwidth = fitz.getTextlength(temp, fontsize=int(placer_font_size))   
                            
                chrPerLine=int(float(prect.width))/textwidth
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=temp * int(Totalchrs)

                if isinstance(placer_font_size, float)==True:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                else:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                
                remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3        

                if remain_space>0:                
                    RchrPerLine=int(remain_space)/textwidth
                    remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                    remain_chrs = temp[0:int(remain_chrs_count)]
                else:
                    remain_chrs = ''
                    
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prect, (temp * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)    
        else:
            qr_details = box['qr_details'].replace("{", "").replace("}", "")
            qr_list=list(filter(bool, qr_details.splitlines()))
            for inv in qr_list:
                str_val = inv.split("^")
                str_val_count=len(str_val)
                mi_txt = mi_txt + extract_microline_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                    
            
            if "Kruti" in box['placer_font_name']:
                temp=mi_txt.replace(" ", "")
                get_text=[]
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                search_text = get_text[0].strip()
                text_len=len(temp)                                   
                text_to_cal=temp                                   
            else:
                get_text=[]
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                temp = get_text[0].replace(" ", "").strip()              
                search_text = get_text[0].strip()
                text_len=len(temp)
                text_to_cal=search_text

            mi_txt = mi_txt.replace(" ", "")
            
            if placer_font_underline=="underline":
                rl = doc.loadPage(cnt-1).searchFor(search_text, hit_max=1)

                if page_count==1:
                    output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                else:
                    if len(rl) > 2:
                        output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                    else:
                        output = str(rl[len(rl)-1]).replace("Rect", "").replace("(", "").replace(")", "")
                
                ucoords = output.split(",")
                urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                urect_coords = int(float(ucoords[1]))+int(float(urect.height))-2                    
                new_urect=fitz.Rect(ucoords[0],urect_coords,ucoords[2],ucoords[3])
                
                if isinstance(placer_font_size, float)==True:
                    textwidth = fitz.getTextlength(text_to_cal, fontsize=float(placer_font_size))    
                else:
                    textwidth = fitz.getTextlength(text_to_cal, fontsize=int(placer_font_size))
                            
                chrPerLine=(int(float(new_urect.width))/textwidth)
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=mi_txt * int(Totalchrs)

                if isinstance(placer_font_size, float)==True:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))-10   
                else:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                
                remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))-3

                if remain_space>0:                
                    RchrPerLine=int(remain_space)/textwidth
                    remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                    remain_chrs = mi_txt[0:int(remain_chrs_count)]
                else:
                    remain_chrs = ''                      
                
                wl_lst = fitz.getTextlength((mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size)+1
                wl = sum([wl_lst])

                if wl>new_urect.width:      
                    remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))-10     
                else:       
                    remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))-3
                
                if remain_space>0:                
                    RchrPerLine=int(remain_space)/textwidth
                    remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                    remain_chrs = mi_txt[0:int(remain_chrs_count)]
                else:
                    remain_chrs = ''
                
                wr = fitz.TextWriter(page.rect)
                if "Kruti" in box['placer_font_name']:
                    wr.fillTextbox(new_urect, (mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align)
                else:
                    wr.fillTextbox(new_urect, (mi_txt * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                
                wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True) 
            else:
                text_len=len(mi_txt)

                if isinstance(placer_font_size, float)==True:
                    textwidth = fitz.getTextlength(mi_txt, fontsize=float(placer_font_size))    
                else:
                    textwidth = fitz.getTextlength(mi_txt, fontsize=int(placer_font_size))   
                            
                chrPerLine=(int(float(prect.width))/textwidth)
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=mi_txt * int(Totalchrs)

                if isinstance(placer_font_size, float)==True:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                else:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                
                    
                remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3 

                if remain_space>0:                
                    RchrPerLine=int(remain_space)/textwidth
                    remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                    remain_chrs = mi_txt[0:int(remain_chrs_count)]
                else:
                    remain_chrs = ''                    
                                    
                wl_lst = fitz.getTextlength((mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size)+1
                wl = sum([wl_lst])

                if wl>prect.width:      
                    remain_space=int(float(prect.width))-int(float(repeat_textwidth))-10     
                else:       
                    remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3
                
                if remain_space>0:                
                    RchrPerLine=int(remain_space)/textwidth
                    remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                    remain_chrs = mi_txt[0:int(remain_chrs_count)]
                else:
                    remain_chrs = ''
                    
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prect, (mi_txt * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)

    def exit_if_already_created(doc):
        if sys.argv[4]=='Fresh':
            get_pdffiles = []
            for pname in doc:
                for boks in boxes:
                    if boks['placer_type'] == 'QR Dynamic':
                        get_text=[]
                        if boks['source'] != '' and boks['source'] != 'Current DateTime':
                            scoords = boks['source_coords'].split(",")
                            srect = fitz.Rect(scoords[0],scoords[1],scoords[2],scoords[3])                
                            get_text.append(pname.getTextbox(srect).replace("\n", ""))
                            extracted_id = get_text[0].replace("/", "").replace("\\", "").replace("-", "").strip()
                            if connection.is_connected():

                                # --------- start 23 feb
                                if sys.argv[18]=='1':
                                    cursor.execute("select * from sb_individual_records where unique_no = '%s' and publish=1 order by id desc" % (extracted_id))
                                else:
                                    cursor.execute("select * from individual_records where unique_no = '%s' and publish=1 order by id desc" % (extracted_id))        
                                # --------- end 23 feb
                                
                                records = cursor.fetchall()  
                                row_count = cursor.rowcount
                                if row_count > 0:    
                                    get_pdffiles.append(records[0][8]) 

            if not get_pdffiles:
                print("No Duplicates")    #List is empty
            else:
                print("Duplicates")
                print(len(get_pdffiles))
                get_pdffiles_string = ','.join(get_pdffiles)
                print(get_pdffiles_string) #Duplicate unique ids
                """
                for unid in get_pdffiles:
                    if connection.is_connected():
                        sql = "UPDATE individual_records SET publish = 2 WHERE unique_no = '%s'" % (unid)
                        cursor.execute(sql)
                        connection.commit()
                """
                exit()

    def check_print_limit(page_count):
        # store the count of pdf generated by particular site_id
        if connection.is_connected():
            # --------- start 23 feb
            if sys.argv[18] == '1':
                cursor.execute("select count(*) AS ts from sb_student_table where site_id = '%s'" % (sys.argv[13]))
            else:
                cursor.execute("select count(*) AS ts from student_table where site_id = '%s'" % (sys.argv[13]))
            
            rsStudent = cursor.fetchone()
            studentTableCounts=rsStudent[0]

        connection2 = mysql.connector.connect(host=sys.argv[10], database='seqr_demo', user=sys.argv[11], password=sys.argv[12])

        # BASED ON THE SITE ID, UPDATE THE VALUE OF PDF GENERATED EACH TIME in super_admin table.Also FETCH THE PRINT LIMIT AND CURRENT VALUE FROM super_admin
        if connection2.is_connected():
            cursor2 = connection2.cursor(buffered=True)
            cursor2.execute ("UPDATE super_admin SET current_value='%s' WHERE site_id=%s " % (studentTableCounts, sys.argv[13]))    
            cursor2.execute("select value, current_value from super_admin where site_id = '%s'" % (sys.argv[13]))
            rsGenerated = cursor2.fetchone()  
            printLimit=int(rsGenerated[0])
            currentValue=int(rsGenerated[1])
            recordGenerated= currentValue + int(page_count)
            noOfCertificateCanGenerate=printLimit-currentValue

            # --------- start 23 feb
            siteurl_param=sys.argv[7]+".seqrdoclocal.com"
            #siteurl_param="demo.seqrdoclocal.com"
            cursor2.execute("select IFNULL(bc_wallet_address, '') AS bc_wallet_address from sites where site_url = '%s'" % (siteurl_param))
            rsSite = cursor2.fetchone()
            # --------- end 23 feb

            connection2.commit()

        # If you reached the print limit, it will stop 
        if currentValue == printLimit:
            print("Over Limit")
            print("You have reached a limit for generating PDF.")
            exit()

        if recordGenerated > printLimit:
            print("Over Limit")
            if noOfCertificateCanGenerate > 1:
                print("You can generate "+str(noOfCertificateCanGenerate)+" PDFs.")
            else:
                print("You can generate "+str(noOfCertificateCanGenerate)+" PDF.")
            exit()

        return rsSite[0]


# def main():

# print(sys.argv[17]);

# --------- start 23 feb
if sys.argv[18] == '1':
    doc = fitz.open(rootDir+sys.argv[7]+'/uploads/data/sandbox/'+sys.argv[2]) #directory+"uploads/data/sandbox/"+
else:
    doc = fitz.open(rootDir+sys.argv[7]+'/uploads/data/'+sys.argv[2]) #directory+"uploads/data/"+
# --------- end 23 feb
    
page_count=doc.pageCount
arr_content = {}  #The array for storing the progress.
cnt = 1

sec_features = SecurityFeatures

# CHECK PRINT LIMIT EXCEED OR NOT
# Also get the value of bc_wallet_address
bc_wallet_address = sec_features.check_print_limit(page_count)

if sys.argv[7]!='mpkv':
    # extract the text from extract coords, Else exit if image is there or nothing is present  
    sec_features.extract_from_coords_else_exit(doc)

if sys.argv[4]=='Fresh':
    #Check duplicate ids
    sec_features.exit_if_already_created(doc)

path_pdf_moved= inner_folder+"/" +sys.argv[2]  

# ------- EXCEL FILE SETUP (start)------------------
# Create excel file to save unique id and QR details
workbook_name=inner_folder+"/" + template_name+".xlsx"

wbc = openpyxl.Workbook()            
wbc.save(filename=workbook_name)
wbs = load_workbook(workbook_name)

page_sheet = wbs.active
page_sheet.column_dimensions['A'].width = 30
page_sheet.column_dimensions['B'].width = 50
page_sheet.column_dimensions['C'].width = 120

page_sheet.append(['Original ID','Unique ID','QR Details'])  
page_sheet.cell(row = 1, column = 1).font = Font(bold = True) 
page_sheet.cell(row = 1, column = 2).font = Font(bold = True) 
page_sheet.cell(row = 1, column = 3).font = Font(bold = True)

# ------- EXCEL FILE SETUP (end)------------------

datetime_IND = datetime.datetime.now(tz_IND) 
beginning_time = datetime_IND.strftime("%H:%M:%S")
start_time = time.time()

# ----23 FEB start
#convert aws string to array
awsS3InstancesArr = sys.argv[17].split('#')
test_pdffiles=[]
# ----23 FEB END

# process start
for i in doc:
    page = i                                                                      
    if not(page._isWrapped):
        page._wrapContents()
                                                
    page_dict = page.getText('dict')                                             
    page_rect=page.MediaBox                                                        

    # ----23 FEB start
    if sys.argv[18] == '1':
        dirName = directory+sys.argv[7]+'/'+"documents/sandbox/" + str(sys.argv[1])
    else:
        dirName = directory+sys.argv[7]+'/'+"documents/" + str(sys.argv[1])

    if not os.path.exists(dirName):
        os.makedirs(dirName)

    mintData={}
    mintData["documentType"] = bc_document_type
    mintData["description"] = bc_document_description
    mcount=1
    use_count=0

    for box_blockchain in extractor_details:
        if "blockchain_flag" in box_blockchain:
            if box_blockchain['blockchain_flag'] == 'use':
                box_blockchain_show_flag = 1
                use_count +=1
            else:
                box_blockchain_show_flag = 0
        else:
            box_blockchain_show_flag = 0  

        if box_blockchain_show_flag==1:
            store_metalabel=box_blockchain['metadata_label']
            store_metavalue=box_blockchain['metadata_value']
            store_coords=box_blockchain['coords']
            meta_value=''
            metadata_label = box_blockchain['metadata_label']
            metadata_value = box_blockchain['metadata_value'].replace("{", "").replace("}", "")
            meta_list=list(filter(bool, metadata_value.splitlines()))

            for mdv in meta_list:
                str_val = mdv.split("^")
                meta_value = meta_value + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)  
            
            mintData['metadata'+str(mcount)]=json.dumps(dict(label = metadata_label, value = meta_value))
            mcount +=1

    if use_count > 0 and use_count < 5:
        for uc in range(use_count+1, 6):
            mintData['metadata'+str(uc)]=json.dumps(dict({}))
    
    # ----23 FEB END
    
    for box in boxes:
        temp = ''
        otxt = ''
        file_path = ''

        placer_coords = box['placer_coords']                                    
        pcoords = placer_coords.split(",")
        prect = fitz.Rect(pcoords[0],pcoords[1],pcoords[2],pcoords[3])         
        prect_coords = pcoords[0],pcoords[1],pcoords[2],pcoords[3] 

        placer_font_underline = box['placer_font_underline'] 
        placer_font_size = box['placer_font_size']                                      
        placer_type = box['placer_type']                                        
        placer_display = box['placer_display']

        # srect based on "source" key
        if box['source'] == '' or box['source'] == 'Current DateTime':
            srect = fitz.Rect(0,0,0,0)            
        else:
            source_coords = box['source_coords']
            scoords = source_coords.split(",")
            srect = fitz.Rect(scoords[0],scoords[1],scoords[2],scoords[3])       
        
        # sets placer_font_name to arial if box['placer_font_name'] is "" otherwise set the font which is mentioned
        if box['placer_font_name'] == '':  
            placer_font_name = fitz.Font(fontfile=dirFont+"arial.ttf")   
        else:
            placer_font_name = fitz.Font(fontfile=dirFont+box['placer_font_name'])
            
        # placer_font_color based on "font_color" key
        if box['font_color'] == '':
            placer_font_color = black
        else:
            placer_font_color = box['font_color']
            placer_font_color=fitz.utils.getColor(placer_font_color) 

        # placer_align to int(0) if placer_display is an empty string, otherwise, it converts placer_display to an integer.
        placer_align = int(0) if placer_display == '' else int(placer_display)        

        placer_degree_angle = box['degree_angle']
        # sets 0 if placer_degree_angle is null else set to int
        placer_degree_angle = int(0) if placer_degree_angle == '' else int(placer_degree_angle)

        placer_opacity = box['opacity_val']
        # sets placer_lineHeight to 1 if box['line_height'] is an empty string, otherwise, it sets it to the value of box['line_height'].
        placer_lineHeight = 1 if box['line_height'] == '' else box['line_height']
        
        # sets qr_show_flag to 1 if the value is "show", otherwise it set to 0.
        qr_show_flag = 1 if box.get("qr_place", "") == "show" else 0 

        # sets barcode_content_flag to 1 if the value is "Source Content", otherwise, it sets it to 0. 
        barcode_content_flag = 1 if box.get("barcode_content", "") == "Source Content" else 0
    
        # sets barcode_content_position_flag to 1 if the value is "Text at Bottom", otherwise, it sets it to 0.
        barcode_content_position_flag = 1 if box.get("barcode_content_position", "") == "Text at Bottom" else 0

        # checks if the key 'blockchain_flag' is in the dictionary box, and if its value is 'use', it sets blockchain_show_flag to 1; otherwise, it sets it to 0
        blockchain_show_flag = 1 if box.get('blockchain_flag') == 'use' else 0

        #    ******  Below code is on placer_type ********
        if placer_type == 'QR Default': 
            barcode_en, qr_txt = sec_features.qr_default(qr_show_flag, page, prect, cnt, dirName)
            
        elif placer_type == 'QR Dynamic': 
            barcode_en, qr_txt, dt_string = sec_features.qr_dynamic(box, page, qr_show_flag, srect, prect, cnt, dirName)

        elif placer_type == 'Barcode': 
            sec_features.place_barcode(barcode_content_flag, barcode_content_position_flag, page, srect, prect, placer_font_size, pcoords, dirName)
        
        elif placer_type == 'Micro Line': 
            sec_features.place_micro_line(box, page, srect, prect, placer_font_underline, cnt, placer_font_size, placer_font_name, placer_align, placer_font_color, page_count, doc)
        
        elif placer_type == 'Invisible':    
            sec_features.placer_invisible_text(box, page, srect, prect, placer_font_name, placer_font_size, placer_align)
        
        elif placer_type == 'Invisible Image': 
            temp_img = [block for block in page_dict['blocks'] if (fitz.Rect(block['bbox']) in srect and block['type'] == 1)]  
            if len(temp_img) > 0:
                pix = fitz.Pixmap(temp_img[0]['image'])
                file_path = dirName + "/"+str(int(round(time.time() * 1000))) + "-" + str(cnt) + "." + temp_img[0]['ext']
                file_path2 = dirName + "/"+str(int(round(time.time() * 1000))) + "-" + str(cnt) + ".png"                        
                pix.writeImage(file_path)
                imgs = Image.open(file_path).convert("L") 
                img1 = ImageOps.colorize(imgs, black ="white", white ="yellow")             
                img1.save(file_path2, 'png')
                page.insertImage(prect, file_path2, overlay=True)                       
        
        elif placer_type == 'Ghost Image': 
            sec_features.place_ghost_image(placer_degree_angle, srect, prect, pcoords, page, box, placer_font_size, dirName)
        
        elif placer_type == 'Image':
            image_path=rootDir+sys.argv[7]+"/backend/templates/pdf2pdf_images/" + box['image_path']
            page.insertImage(prect, image_path,overlay=True)
        
        elif placer_type == 'Plain Text':    
            sec_features.plain_text(box, page, srect, prect, placer_font_name, placer_font_size, placer_align, placer_font_color)
        
        elif placer_type == 'Static Text':     
            otxt_string = box['qr_details']
            m = fitz.Matrix(0) 
            points = fitz.Point(prect.x0, prect.y0)
            wr = fitz.TextWriter(page.rect)
            wr.fillTextbox(prect, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
            wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))

            if placer_font_underline=="underline":
                rl = page.searchFor(otxt_string)  
                output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")                    
                ucoords = output.split(",")
                urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                shape = page.newShape()
                shape.drawLine(urect.bl, urect.br)
                shape.finish(color=placer_font_color, stroke_opacity=float(placer_opacity))
                shape.commit()            
            
        elif placer_type == 'Watermark Text':     
            sec_features.place_watermark_text(box, page, srect, prect, placer_degree_angle, placer_font_name, placer_font_size, placer_align, placer_font_color, placer_opacity)
        
        elif placer_type == 'Watermark Multi Lines':
            sec_features.place_watermark_multi_lines(box, srect, page, placer_degree_angle, page_rect, placer_font_name, placer_font_size, placer_align, placer_lineHeight, placer_font_color, placer_opacity)

        # --------- start 23 feb
        """
        if blockchain_show_flag==1 and bc_contract_address != '':
            meta_value=''
            metadata_label = box['metadata_label']
            metadata_value = box['metadata_value'].replace("{", "").replace("}", "")
            meta_list=list(filter(bool, metadata_value.splitlines()))
            for mdv in meta_list:
                str_val = mdv.split("^")
                meta_value = meta_value + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)  
            mintData['metadata'+str(mcount)]=json.dumps(dict(label = metadata_label, value = meta_value))
            mcount +=1
        """ 
        
        #page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
        #page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_REMOVE)     
        # --------- end 23 feb

        doc_new = fitz.open()
        doc_new.insertPDF(doc, from_page=cnt-1, to_page=cnt-1)

        if verification_setbg=='Yes':
            for dpage in doc_new:
                dpage.insertImage(page_rect, verification_bg_file,overlay=False)  

        # --------- start 23 feb        
        """
        if sys.argv[13] :
            URL = "https://"+sys.argv[7]+"."+sys.argv[10]+"/admin/store-file"
            pdf_path=pdf_folder+"/"+dt_string+".pdf"
            PARAMS = {'pdf_path':pdf_path,'site_id':sys.argv[13]}
            response =requests.get(url = URL, params = PARAMS, timeout=None)        
            response.close()     
        #doc_new.save(pdf_folder+"/"+dt_string+".pdf", garbage=4, deflate=True, owner_pw="owner", encryption=encrypt_meth, permissions=perm) 
        """ 

        if sys.argv[18] == '1':
            if not os.path.exists(rootDir+"/backend/pdf_file/sandbox"):
                os.makedirs(rootDir+"/backend/pdf_file/sandbox") 

            pdf_loc = rootDir+"/backend/pdf_file/sandbox/" + dt_string+".pdf"
            doc_new.save(pdf_loc, garbage=4, deflate=True)
            pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/sandbox/" + dt_string+".pdf"
            
        else:
            if not os.path.exists(rootDir+"/backend/pdf_file"):
                os.makedirs(rootDir+"/backend/pdf_file") 

            pdf_loc = rootDir+"/backend/pdf_file/" + dt_string+".pdf"
            doc_new.save(pdf_loc, garbage=4, deflate=True)
            pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
            


        if is_compressed_enabled == 1:
            # compressed & save to new loc and delete previous loc
            compress(pdf_loc, pdf_file_path)
            os.remove(pdf_loc)
        else:
            shutil.move(pdf_loc, pdf_file_path)

        """ 
        set_background(
            input_pdf=verification_bg_file, 
            output=pdf_folder+"/"+dt_string+".pdf",
            watermark=pdf_folder+"/"+dt_string+".pdf")   
        """ 
        
    if use_count > 0 and bc_contract_address != '' and sys.argv[18] != '1':      
        pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
        pdf_file_path = pdf_file_path.replace("//", "/")
        pdf_file_path = pdf_file_path.replace("/", "\\")
        #print(pdf_file_path)
        """
        temp_dir=tempfile.gettempdir()
        r = requests.post("https://"+sys.argv[7]+".seqrdoclocal.com/"+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf", allow_redirects=True)
        open(temp_dir+'\/'+dt_string+'.pdf', 'wb').write(r.content)
        files=[
            ('document',(dt_string+'.pdf',open(temp_dir+'\/'+dt_string+'.pdf','rb'),'application/pdf'))
        ]
        """        
        files=[
            ('document',(dt_string+'.pdf',open(pdf_file_path,'rb'),'application/pdf'))
        ]

        if sys.argv[7]=='mpkv':
            mintData["description"]=mintData["description"] + " "+dt_string

            mintData["walletID"] = bc_wallet_address 
            mintData["smartContractAddress"] = bc_contract_address 
            mintData["uniqueHash"] = barcode_en
            mintData["pdf_file"] = "https://"+sys.argv[7]+".seqrdoclocal.com/"+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
            #mintData["template_id"] = template_id

        if sys.argv[7]=='mpkv' or sys.argv[7]=='ksu' or (sys.argv[7]=='demo' and template_id==137):
            if(sys.argv[7]=='demo'):
                mintData["walletID"] = "0xB509AF6532Af95eE59286A8235f2A290c26b5730" 
            #blockchain_response = requests.post('https://veraciousapis.herokuapp.com/v1/mint', data=mintData, files=files)
            blockchain_response = requests.post('https://mainnet-apis.herokuapp.com/v1/mainnet/mint', data=mintData, files=files)
        else:
            blockchain_response = requests.post('https://veraciousapis.herokuapp.com/v1/mint', data=mintData, files=files)
            #print(mintData)
            #print(blockchain_response)
    else:
        if sys.argv[7]=='mpkv':
            blockchain_response = requests.post('https://veraciousapis.herokuapp.com/v1/mint') #"<Response [201]>"
            #blockchain_response = requests.post('https://mainnet-apis.herokuapp.com/v1/mainnet/mint') #"<Response [201]>"
        else:
            blockchain_response = requests.post('https://veraciousapis.herokuapp.com/v1/mint') #"<Response [201]>"
        
            
    # --------- end 23 feb
            
        

    if print_setbg=='Yes':
        page.insertImage(page_rect, print_bg_file,overlay=False)   

    # if pdf file successfully gets create, a record is generated inside individual_records table to avoid creation of same file again  
    if connection.is_connected():    
        # --------23 feb start
        if sys.argv[18]=='1':
            sql = "INSERT INTO sb_individual_records (file_records_id, template_id, template_name, pdf_page, page_no, encoded_id, unique_no, qr_details, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        else:
            sql = "INSERT INTO individual_records (file_records_id, template_id, template_name, pdf_page, page_no, encoded_id, unique_no, qr_details, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        # --------23 feb end

        # 134 3 MPKV_Certificate_2 Single 1 7e83f3861a6ae3b1c4357e2863c1ea10 MPKV000006720 7e83f3861a6ae3b1c4357e2863c1ea10 6 20240226160821-d465e4c27623
        val = (file_records_next_id, template_id, template_name, pdf_page, cnt, barcode_en, dt_string, qr_txt, userid, record_unique_id)
        cursor.execute(sql, val)         
        
    # it will store the record count of latest pdf generated, acc to status and same serial no
    current_date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if connection.is_connected():

        # --------23 feb start
        if sys.argv[18] == '1':
            cursor.execute("select * from sb_student_table where serial_no = '%s' and status=1 order by id desc" % (dt_string))
        else:
            cursor.execute("select * from student_table where serial_no = '%s' and status=1 order by id desc" % (dt_string))
        # --------23 feb end
        
        records = cursor.fetchall()  
        row_count = cursor.rowcount

        # on adding new pdf of same serial no, the status of old pdf is set to 0
        if row_count > 0: 
            # --------23 feb start
            if sys.argv[18] == '1':
                sql = "UPDATE sb_student_table SET status = %s, updated_by = %s, updated_at = %s WHERE serial_no = %s"
            else:
                sql = "UPDATE student_table SET status = %s, updated_by = %s, updated_at = %s WHERE serial_no = %s"
            # --------23 feb end

            # 0 6 2024-02-12 10:03:36 MPKV000006720
            val = ("0", userid, current_date_time, dt_string)            
            cursor.execute(sql, val)   

        
        certificate_filename=dt_string+'.pdf'
        qr_path='qr/'+barcode_en+'.png'

        # --------23 feb start
        # On successfully pdf generation a records is inserted inside student_table
        if blockchain_response.status_code == 200:
            mint_json = json.loads(blockchain_response.text)

            if sys.argv[18] == '1':
                sql2 = """INSERT INTO sb_student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`, `bc_txn_hash`,`bc_ipfs_hash`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            else:
                sql2 = """INSERT INTO student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`, `bc_txn_hash`,`bc_ipfs_hash`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

            val2 = (dt_string, certificate_filename, template_id, barcode_en, qr_path, userid, userid, 1, 1, current_date_time,sys.argv[13],1,mint_json['txnHash'],mint_json['ipfsHash'])
        else:
            
            if sys.argv[18] == '1':
                sql2 = """INSERT INTO sb_student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            else:
                sql2 = """INSERT INTO student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            
            val2 = (dt_string, certificate_filename, template_id, barcode_en, qr_path, userid, userid, 1, 1, current_date_time,sys.argv[13],1)

        # --------23 feb end

        cursor.execute(sql2, val2) 
        student_table_id = cursor.lastrowid

        # --------23 feb start
        
        if blockchain_response.status_code == 200:
            mint_json = json.loads(blockchain_response.text)
            #print(mint_json['txnHash'], mint_json['tokenID'], mint_json['gasPrice'])

            if sys.argv[18] == '1':
                sql = "UPDATE sb_student_table SET bc_txn_hash = %s WHERE serial_no = %s AND status = %s"
            else:
                sql = "UPDATE student_table SET bc_txn_hash = %s WHERE serial_no = %s AND status = %s"

            val = (mint_json['txnHash'], dt_string, '1')
            cursor.execute(sql, val)

            if sys.argv[18] != '1':
                sql2 = """INSERT INTO bc_mint_data (`txn_hash`, `gas_fees`, `token_id`, `key`, `created_at`) VALUES (%s, %s, %s, %s, %s)"""
                val2 = (mint_json['txnHash'], mint_json['gasPrice'], mint_json['tokenID'], barcode_en, current_date_time)
                cursor.execute(sql2, val2) 
        # --------23 feb end

        # Based on the site_id of a system_config table, we fetch the printer name
        cursor.execute("select printer_name from system_config where site_id = '%s'" % (sys.argv[13]))
        recordSystem = cursor.fetchone()  
        printer_name=recordSystem[0]
        
        # --------23 feb start
        # fetch all those record of same serial no from printing_details Table and store it's count
        if sys.argv[18] == '1':
            cursor.execute("SELECT * FROM sb_printing_details WHERE sr_no = '%s' " % (dt_string))
        else:
            cursor.execute("SELECT * FROM printing_details WHERE sr_no = '%s' " % (dt_string))
        # --------23 feb end
        
        records = cursor.fetchall()  
        print_count = cursor.rowcount

        today = date.today()
        current_year=get_financial_year(str(today))
        current_year='PN/'+current_year+'/'
        
        # --------23 feb start
        # from the printing_details table, it return us the latest serial no that has been used, so we add +1 in it and used it for current pdf
        if sys.argv[18] == '1':
            cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM sb_printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
        else:
            cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
        # --------23 feb end
        
        record = cursor.fetchone() 
        next_print=record[0]+1
        next_print_serial=current_year+str(next_print)
        
        # --------23 feb start
            # It will store the record in printing_details Table
        if sys.argv[18] == '1':
            sql3 = """INSERT INTO sb_printing_details (`username`, `print_datetime`, `printer_name`, `print_count`, `print_serial_no`, `sr_no`, `template_name`,`created_at`, `created_by`, `updated_at`, `updated_by`,`status`, `site_id`, `publish`, `student_table_id`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        else:
            sql3 = """INSERT INTO printing_details (`username`, `print_datetime`, `printer_name`, `print_count`, `print_serial_no`, `sr_no`, `template_name`,`created_at`, `created_by`, `updated_at`, `updated_by`,`status`, `site_id`, `publish`, `student_table_id`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        # --------23 feb end

        #  admin 2024-02-12 10:03:36 Printer 1 8 PN/23-24/354 MPKV000006720 MPKV_Certificate_2 2024-02-12 10:03:36 6 2024-02-12 10:03:36 6 1 233 1 0
        val3 = (sys.argv[14], current_date_time, printer_name, print_count, next_print_serial, dt_string, template_name, current_date_time, userid, current_date_time,userid, 1, sys.argv[13], 1,student_table_id)
        cursor.execute(sql3, val3) 

    # --------23 feb start
    if (sys.argv[7] in awsS3InstancesArr):    
        source_file=pdf_file_path
        test_pdffiles.append(source_file)
    # --------23 feb end

    # ------- excel data ------------------
    page_sheet.append([dt_string,barcode_en,qr_txt])
    page_sheet.cell(row = cnt+1, column = 1).alignment = Alignment(vertical='top')
    page_sheet.cell(row = cnt+1, column = 2).alignment = Alignment(vertical='top')
    page_sheet.cell(row = cnt+1, column = 3).alignment = Alignment(wrapText=True,vertical='top')
    wbs.save(filename=workbook_name) 
    # ------- excel data (end)------------------

    # ------- json file (start)------------
    datetime_IND = datetime.datetime.now(tz_IND)
    ending_time = datetime_IND.strftime("%H:%M:%S")
    end_time=time.time() - start_time
    seconds_to_hhmmss=time.strftime('%H:%M:%S', time.gmtime(end_time))
    page_seconds_to_hhmmss=time.strftime('%H:%M:%S', time.gmtime(end_time/cnt))

    arr_content['percent'] = int(cnt/page_count * 100)
    arr_content['message'] = "Generating "+str(cnt)+"/"+str(page_count)+" PDF(s)"
    arr_content['beginning_time'] = beginning_time
    arr_content['ending_time'] = ending_time
    arr_content['exec_time'] = end_time
    arr_content['hms_time'] = seconds_to_hhmmss
    arr_content['page_time'] = page_seconds_to_hhmmss
    arr_content['pages_processed'] = cnt
    json_object = json.dumps(arr_content, indent = 4)

    f = open(rootDir+sys.argv[7]+'/'+"processed_pdfs/"+sys.argv[5], "w")
    f.write(json_object)
    # ------- json file (end)------------

    cnt += 1

# Store's the single record after each execution in file_records
if connection.is_connected():        
    # --------23 feb start
    if sys.argv[18]=='1':
        sql = "INSERT INTO sb_file_records (template_id, template_name, pdf_page, total_records, source_file, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    else:
        sql = "INSERT INTO file_records (template_id, template_name, pdf_page, total_records, source_file, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    # --------23 feb end

    val = (template_id, template_name, pdf_page, cnt-1, sys.argv[2], userid, record_unique_id)              # 3 MPKV_Certificate_2 Single 1 MPKV_with_out_9_line_1_process_1707455271.pdf 6 20240212100333-385168314574
    cursor.execute(sql, val)
    connection.commit()

print('path_pdf_moved----', path_pdf_moved)

# --------23 feb start
#Upload pdf to AWS s3
if (sys.argv[7] in awsS3InstancesArr):
    URL = "https://"+sys.argv[7]+"."+sys.argv[10]+"/admin/upload-file-s3"
    test_pdffilesStr = ', '.join(test_pdffiles)
    PARAMS = {'pdf_paths':test_pdffilesStr}
    response =requests.get(url = URL, params = PARAMS, timeout=None)        
    response.close()  
    print("mmk"+test_pdffilesStr)
# --------23 feb end

removePath=rootDir+sys.argv[7]+'/'
rewisePath = path_pdf_moved.replace(removePath,'')
doc.save(path_pdf_moved, garbage=4, deflate=True)
print(rewisePath)
# print('rewisePath----', rewisePath)
shutil.rmtree(dirName, ignore_errors=True)
cursor.close()
connection.close()


"""
if sys.argv[13] :
    URL = "https://"+sys.argv[7]+"."+sys.argv[10]+"/admin/store-file"
    pdf_path=pdf_folder+"/"+dt_string+".pdf"
    PARAMS = {'pdf_path':pdf_path,'site_id':sys.argv[13]}
    response =requests.get(url = URL, params = PARAMS)
"""


# if __name__ == "__main__":
#     main()