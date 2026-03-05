#!C:/Users/Administrator/AppData/Local/Programs/Python/Python38/python.exe
from __future__ import print_function
from operator import itemgetter
from itertools import groupby
import sys
import os
import shutil
from shutil import make_archive
import simplejson as json
import fitz
import mysql.connector
import pprint
import time
from mysql.connector import Error
import textwrap
from PIL import Image, ImageOps 
import numpy as np
#import cv2.cv2 as cv2
import barcode
from barcode.writer import ImageWriter
import qrcode
import hashlib 
# from datetime import datetime
from xml.dom.minidom import parseString
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font 
from openpyxl.styles import Alignment
import openpyxl.utils.cell
# from datetime import datetime

import datetime
from datetime import date
import uuid
import pytz
import xlrd
import locale
import re
from PIL import ImageDraw, ImageFont
# from getFunctions import *
import math

tz_IND = pytz.timezone('Asia/Calcutta')
#sys.argv[1] = template id
#sys.argv[2] = data file
#sys.argv[3] = session user id
#sys.argv[4] = entry type (Fresh/Proceed)
#sys.argv[5] = progress file
#sys.argv[6] = dbName
#sys.argv[7] = subdomain
#sys.argv[8] = directoryUrlForward
#sys.argv[9] = directoryUrlBackward
#sys.argv[10] = servername
#sys.argv[11] = username
#sys.argv[12] = password
#sys.argv[13] = siteid
#sys.argv[14] = username
#sys.argv[15] = printer_name
# python extract_and_place.py 1 FYBAF_1619953609.pdf 1
#print(sys.argv[2])



try:
    directory=sys.argv[8]
    rootDir= directory.replace('excel2pdf', '')

    
    # connection = mysql.connector.connect(host=host_var, database=database_var, user=user_var, password=password_var)
    connection = mysql.connector.connect(host=sys.argv[10],
                                         database=sys.argv[6],
                                         user=sys.argv[11],
                                         password=sys.argv[12])

    
    if connection.is_connected():
        db_Info = connection.get_server_info()
        cursor = connection.cursor()
        cursor.execute("select ep_details, id, extractor_details, template_name, pdf_page, print_bg_file, print_bg_status, verification_bg_file, verification_bg_status, file_name from uploaded_pdfs where id = '%s'" % (sys.argv[1]))
        record = cursor.fetchone()        
        boxes = json.loads(record[0], strict=False)       
        template_id=record[1]
        extractor_details = json.loads(record[2], strict=False)
        template_name=record[3]
        pdf_page=record[4]
        pbg_file=record[5]
        print_bg_status=record[6]
        vbg_file=record[7]
        verification_bg_status=record[8]
        template_file_name=record[9]
        cur=connection.cursor()
        print_setbg=''
        verification_setbg=''
        if pbg_file !=0 and print_bg_status == 'Yes':
            sql_bg="SELECT image_path FROM background_template_master where id= '%s'" % (pbg_file)
            cur.execute(sql_bg)
            precord = cur.fetchone()
            #print_bg_file=sys.argv[8]+"upload_bgs/"+precord[0] 
            print_bg_file=rootDir+sys.argv[7]+"/backend/canvas/bg_images/" + precord[0]
            #print_bg_file="C:/wamp/www/demo/upload_bgs/"+precord[0] 
            print_setbg='Yes'
            #print(print_bg_file)
        if vbg_file !=0 and verification_bg_status == 'Yes':
            sql_bg="SELECT image_path FROM background_template_master where id= '%s'" % (vbg_file)
            cur.execute(sql_bg)
            vrecord = cur.fetchone()
            #verification_bg_file=sys.argv[8]+"upload_bgs/"+vrecord[0]
            verification_bg_file=rootDir+sys.argv[7]+"/backend/canvas/bg_images/" + vrecord[0]
            #verification_bg_file="C:/wamp/www/demo/upload_bgs/"+vrecord[0]
            verification_setbg='Yes'
            #print(verification_bg_file)

        record_unique_id = datetime.datetime.now().strftime('%Y%m%d%H%M%S-') + str(uuid.uuid4()).split('-')[-1]
        sqli="SELECT id FROM file_records ORDER BY id DESC LIMIT 1"
        cur.execute(sqli)
        last_id = cur.fetchone()
        if last_id == None:
            file_records_next_id=1
        else:
            file_records_next_id=last_id[0]+1    
        # get last record's columns
        #for lastID in last_id:
            #print(lastID)     

except Error as e:
    print("Error while connecting to MySQL", e)
"""
finally:
    if (connection.is_connected()):
        #cursor.close()
        #connection.close()
        print("MySQL connection is closed")
"""
userid=sys.argv[3]

def text_tl_position(L, T, R, B):
    left_pos=L                  
    top_pos=int(float(T))-4
    right_pos=R
    bottom_pos=int(float(B))               
    prects = fitz.Rect(left_pos,top_pos,right_pos,bottom_pos)   
    prects_coords = left_pos,top_pos,right_pos,bottom_pos
    return prects,prects_coords

def calculate_coordinates(L, T, text_width, text_height):
    left_pos=L                  
    top_pos=int(float(T))-4
    right_pos=text_width
    bottom_pos=int(float(text_height))               
    prects = fitz.Rect(left_pos,top_pos,right_pos,bottom_pos)   
    prects_coords = left_pos,top_pos,right_pos,bottom_pos
    return prects,prects_coords

def getTextlength_custom(text, font_path, font_size):
    # Load the font
    font = ImageFont.truetype(font_path, font_size)

    # Create a blank image (required for ImageDraw)
    dummy_image = Image.new("RGB", (1, 1))
    draw = ImageDraw.Draw(dummy_image)

    # Use textbbox to get the bounding box of the text
    bbox = draw.textbbox((0, 0), text, font=font)

    # Calculate width and height from the bounding box
    width = bbox[2] - bbox[0]
    height = bbox[3] - bbox[1]
    return width, height

def extract_microline_info(extractor_detail,search_value,extra_line_first='',extra_line_second='',page=''):
    for keyval in extractor_detail:
        if search_value.lower() == keyval['name'].lower():
            source_coord = keyval['coords']
            scoord = source_coord.split(",")
            s_rect = fitz.Rect(scoord[0],scoord[1],scoord[2],scoord[3])
            get_text=[]            
            get_text.append(page.getTextbox(s_rect).replace("\n", ""))
            otxt=get_text[0].replace(" ", "")
            if extra_line_first == '' and extra_line_second == '':
                return otxt
            elif extra_line_first != '' and extra_line_second != '':
                return extra_line_first + otxt + extra_line_second 
            elif extra_line_first != '' and extra_line_second == '':
                return extra_line_first + otxt 
            elif extra_line_first == '' and extra_line_second != '':
                return otxt + extra_line_second

def extract_plainText(extractor_detail,search_value,extra_line_first='',extra_line_second='',page=''):
    for keyval in extractor_detail:
        if search_value.lower() == keyval['name'].lower():
            source_coord = keyval['coords']
            scoord = source_coord.split(",")
            s_rect = fitz.Rect(scoord[0],scoord[1],scoord[2],scoord[3]) 
            get_text=[]            
            get_text.append(page.getTextbox(s_rect).replace("\n", ""))
            otxt=get_text[0].strip()
            if extra_line_first == '' and extra_line_second == '':
                return otxt +" "
            elif extra_line_first != '' and extra_line_second != '':
                return extra_line_first + otxt + extra_line_second +" " 
            elif extra_line_first != '' and extra_line_second == '':
                return extra_line_first + otxt +" " 
            elif extra_line_first == '' and extra_line_second != '':
                return otxt + extra_line_second +" "                 
                
def extract_info(extractor_detail,search_value,extra_line_first='',extra_line_second='',page=''):
    for keyval in extractor_detail:
        if search_value.lower() == keyval['name'].lower():
            source_coord = keyval['coords']
            scoord = source_coord.split(",")
            s_rect = fitz.Rect(scoord[0],scoord[1],scoord[2],scoord[3])
            get_text=[]            
            get_text.append(page.getTextbox(s_rect).replace("\n", ""))
            otxt=get_text[0].strip()
            if extra_line_first == '' and extra_line_second == '':
                return otxt +"\n"
            elif extra_line_first != '' and extra_line_second != '':
                return extra_line_first + otxt + extra_line_second +"\n" 
            elif extra_line_first != '' and extra_line_second == '':
                return extra_line_first + otxt +"\n" 
            elif extra_line_first == '' and extra_line_second != '':
                return otxt + extra_line_second +"\n" 

def CreateGhostImage(dirName, name, p_font_size, ghost_width, ghost_height):
    #dirChars = "F:/projects/flask-app-env/marksheet/grade_card_reader/defence_secure_docs/chars/"+str(p_font_size)
    #dirChars = "C:/Program Files/Python38/projects/demo/chars/"+str(p_font_size)
    # dirChars=sys.argv[8]+'Python_files/chars/'+str(p_font_size);
    dirChars=sys.argv[8]+'Python_files/chars/'+str(int(p_font_size))
    
    name=name.upper()    
    
    single_char=split(name)       
    my_list = list()
    for c in single_char:
        my_list.append(dirChars +"/"+ c +".png")    
    # print(my_list)
    images = [Image.open(x) for x in my_list]
    widths, heights = zip(*(i.size for i in images))
    total_width = sum(widths)
    max_height = max(heights)    
    new_im = Image.new('RGB', (total_width, max_height))
    x_offset = 0
    for im in images:
      new_im.paste(im, (x_offset,0))
      x_offset += im.size[0]
    new_im.save(dirName +"/"+ name +".png")  
    isize=ghost_width,ghost_height            
    im = Image.open(dirName +"/"+ name +".png")
    im.thumbnail((isize), Image.LANCZOS)
    im.save(dirName +"/"+ name +str(p_font_size)+"_th.png", quality=100)    
    #return dirName +"/"+ name +".png" 
    return dirName +"/"+ name +str(p_font_size)+"_th.png"     
    
def split(word): 
    return [char for char in word]

def get_financial_year(datestring):
            date = datetime.datetime.strptime(datestring, "%Y-%m-%d").date()
            #initialize the current year
            year_of_date=date.year
            #initialize the current financial year start date
            financial_year_start_date = datetime.datetime.strptime(str(year_of_date)+"-04-01","%Y-%m-%d").date()
            if date<financial_year_start_date:
                    return str(financial_year_start_date.year-1)[2:]+'-'+ str(financial_year_start_date.year)[2:]
            else:
                    return str(financial_year_start_date.year)[2:]+'-'+ str(financial_year_start_date.year+1)[2:]    

def repeat_to_length(string_to_expand, length):
    return (string_to_expand * (int(length/len(string_to_expand))+1))[:length]

def get_pil_text_size(text, font_size, font_name):
    font = ImageFont.truetype(font_name, font_size)
    size = font.getsize(text)
    return size


def hex_to_rgb( hex_color):    
    hex_color = hex_color.lstrip('#')

    if len(hex_color) != 6:
        raise ValueError("Invalid hex color code length (must be 6 digits).")
    
    rgb_tuple = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))  # Convert to RGB tuple
    return normalize_rgb(rgb_tuple)


def normalize_rgb(rgb):
    """Convert an RGB tuple to normalized values between 0 and 1."""
    return tuple(x / 255 for x in rgb)



def set_background(input_pdf, output, watermark):
    watermark_obj = PdfFileReader(watermark)
    #watermark_obj.decrypt("owner")
    watermark_page = watermark_obj.getPage(0)

    pdf_reader = PdfFileReader(input_pdf)    
    pdf_writer = PdfFileWriter()

    for page in range(pdf_reader.getNumPages()):        
        page = pdf_reader.getPage(page)
        page.mergePage(watermark_page)
        #page.compressContentStreams()
        pdf_writer.addPage(page)

    with open(output, 'wb') as out:
        pdf_writer.write(out)

encrypt_meth = fitz.PDF_ENCRYPT_AES_256  # strongest algorithm
perm = int(
fitz.PDF_PERM_PRINT  # permit printing
)


# dirFont = sys.argv[8]+"Python_files/fonts/"
dirFont = rootDir+sys.argv[7]+"/backend/canvas/fonts/"
#target_folder="symbiosis/BBA_LOGISTIC" 
# excelfile="anu_college.xlsx" #marksheet_BArchIX.xlsx bed1year.xlsx anu_college.xlsx 
excelfile=rootDir+sys.argv[7]+'/uploads/data/'+sys.argv[2]
# print(excelfile)
wb = xlrd.open_workbook(excelfile) # input file name
sheet = wb.sheet_by_index(0)
rows=sheet.nrows-1
cols=sheet.ncols 
#print(cols, rows)
excel_cols = {}
# read header values into the list    
keys = [sheet.cell(0, col_index).value for col_index in range(sheet.ncols)]
#keys = [[col_index+1,sheet.cell(0, col_index).value] for col_index in range(sheet.ncols)]
#col_inx=keys.index('CandidateName')
#print(col_inx)
#print(openpyxl.utils.cell.get_column_letter(1))
#exit()
wb_obj = openpyxl.load_workbook(excelfile, read_only=True, data_only=True)
sheet_obj = wb_obj.active
max_rows=sheet_obj.max_row
max_columns=sheet_obj.max_column
"""
for data in sheet_obj.iter_rows(max_col=max_columns,max_row=max_rows,values_only=True):
    for d in range(len(data)):
        print(data[d],end=' ') # print all data of a row
    print('') # adding line break 
#cell_obj = sheet_obj.cell(row = 2, column = 10)
#print(cell_obj.value)
exit()
"""
# my_directory="F:/projects/flask-app-env/marksheet/grade_card_reader/defence_secure_docs/excel_pdfs"
my_directory=rootDir+sys.argv[7]+'/uploads/data/'

# doc = fitz.open(rootDir+sys.argv[7]+'/uploads/data/'+sys.argv[2])

# add width and image
template_pdf_path=rootDir+sys.argv[7]+'/'+template_file_name
# # print(template_pdf_path)
template_doc = fitz.open(template_pdf_path)
template_page = template_doc[0]
template_width = template_page.rect.width
template_height = template_page.rect.height

doc_new = fitz.open()
dict_list = []
cntn = 1
cln=0



for row_index in range(1, sheet.nrows):
    #d = {keys[col_index]: sheet.cell(row_index, col_index).value 
        #for col_index in range(sheet.ncols)}
    #dict_list.append(d)
    # A4 size in points

    # print(row_index)
    doc_new.newPage(width=template_width, height=template_height)

    # doc_new.newPage(width=template_width, height=template_height)  
    # doc_new.insertPage(-1)
    cln += 1
    cntn += 1

white = (0, 0, 0, 0)
black = (0, 0, 0, 1)
red = (0, 1, 1, 0)
blue = (1, 1, 0, 0)
pink = (0, 1, 0, 0)
aqua = (1, 0, 0, 0)
green = (1, 0, 1, 0)
purple = (1, 1, 0, 0)
yellow = (0, 0, 1, 0)

locale.setlocale(locale.LC_ALL, '')
l = locale.localeconv()
div1    = l['decimal_point'] or '.'
div1000 = l['thousands_sep'] or ','

# print(my_directory)
if not os.path.exists(rootDir+sys.argv[7]+"/processed_pdfs/preview/"+sys.argv[1]):
    os.makedirs(rootDir+sys.argv[7]+"/processed_pdfs/preview/"+sys.argv[1]) 




dt_string_file_name = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
previewFile = rootDir+sys.argv[7]+"/processed_pdfs/preview/"+sys.argv[1]+"/output_"+ dt_string_file_name+".pdf"

 
doc_new.save(previewFile, garbage=4, deflate=True)
# doc_new.save(my_directory+"/output.pdf", garbage=4, deflate=True)    
cnt = 1

# print(previewFile)
# doc = fitz.open(my_directory+"/output.pdf")

# my_directory

doc = fitz.open(previewFile)




# print(template_width)
# print(template_height)

# add width and image

page_count=doc.pageCount
arr_content = {} #The array for storing the progress.
# Iterate over the rows
#Check print limit 
if connection.is_connected():
    cursor.execute("select count(*) AS ts from student_table where site_id = '%s'" % (sys.argv[13]))
    rsStudent = cursor.fetchone()
    studentTableCounts=rsStudent[0]

connection2 = mysql.connector.connect(host=sys.argv[10],
                                     database='seqr_demo',
                                     user=sys.argv[11],
                                     password=sys.argv[12])
if connection2.is_connected():
    cursor2 = connection2.cursor(buffered=True)
    cursor2.execute ("UPDATE super_admin SET current_value='%s' WHERE site_id=%s " % (studentTableCounts, sys.argv[13]))    
    cursor2.execute("select value, current_value from super_admin where site_id = '%s'" % (sys.argv[13]))
    rsGenerated = cursor2.fetchone()  
    printLimit=int(rsGenerated[0])
    currentValue=int(rsGenerated[1])
    recordGenerated= currentValue + int(page_count)
    noOfCertificateCanGenerate=printLimit-currentValue
    connection2.commit()

if currentValue == printLimit:
    print("Over Limit")
    print("You have reached a limit for generating PDF.")
    exit()

if recordGenerated > printLimit:
    print("Over Limit")
    if noOfCertificateCanGenerate > 1:
        print("You can generate "+str(noOfCertificateCanGenerate)+" PDFs.")
    else:
        print("You can generate "+str(noOfCertificateCanGenerate)+" PDF.")
    exit()


output_file = rootDir+sys.argv[7]+'/'+"processed_pdfs/"+sys.argv[2]

folder=rootDir+sys.argv[7]+'/'+"documents/" + template_name
if not os.path.exists(folder):
    os.makedirs(folder)

inner_folder=folder +"/"+str(record_unique_id)
if not os.path.exists(inner_folder):
    os.makedirs(inner_folder)
    
# pdf_folder=inner_folder +"/pdfs"
# if not os.path.exists(pdf_folder):
#     os.makedirs(pdf_folder)
    
path_pdf_moved= inner_folder+"/" +sys.argv[2]  #



datetime_IND = datetime.datetime.now(tz_IND) 
beginning_time = datetime_IND.strftime("%H:%M:%S")
start_time = time.time()


for i in range(len(doc)):
# for i in doc:

    # Create a new PDF document for the current page
    single_page_doc = fitz.open()

    # Get the current page from the original document
    page = doc[i]

    # page = i
    if not(page._isWrapped):
        page._wrapContents()
    page_data = {cnt:[]}
    words = page.getTextWords()
    page_dict = page.getText('dict')    
    page_rect=page.MediaBox
    row_value = sheet.row_values(cnt)    
    column=1

    dt_string = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    # single_page_doc = fitz.open()

    # Set the page size to match the template
    # page.setMediaBox(fitz.Rect(0, 0, template_width, template_height))
    
    # # Optionally scale content to fit the new size
    # scale_x = template_width / page.rect.width
    # scale_y = template_height / page.rect.height
    # transform = fitz.Matrix(scale_x, scale_y)
    # page.apply_transformation(transform)

    # page(width=template_width, height=template_height)
    # page_data = {cnt:[]}
    # words = page.getTextWords()
    # page_dict = page.getText('dict')    
    # page_rect=page.MediaBox
    # row_value = sheet.row_values(cnt)    
    # column=1
    #page.drawLine(fitz.Point(32, 522), fitz.Point(526, 522), width = 0.5, color = black)
    for box_line in extractor_details:
        line_coords = box_line['coords']
        lcoords = line_coords.split(",")
        lrect = fitz.Rect(lcoords[0],lcoords[1],lcoords[2],lcoords[3]) 
        #print(line_coords)
        #print(lrect.br,lrect.bl)  
        rect_width=lrect.width
        rect_height=lrect.height
        if rect_width > rect_height:
            p1 = fitz.Point(lcoords[0],lcoords[1])
            p2 = fitz.Point(lcoords[2],lcoords[1])
        else:
            p1 = fitz.Point(lcoords[0],lcoords[1])
            p2 = fitz.Point(lcoords[0],lcoords[3])
        #page.drawRect(lrect, color = (0, 0, 0), width=0.5, overlay=True)
        page.drawLine(p1, p2, width = 0.5, color = black)
    
    dirName = directory+sys.argv[7]+'/'+"documents/" + str(sys.argv[1])
    for box in boxes:

        

        placer_coords = box['placer_coords']
        pcoords = placer_coords.split(",")
        prect = fitz.Rect(pcoords[0],pcoords[1],pcoords[2],pcoords[3])
        prect_coords = pcoords[0],pcoords[1],pcoords[2],pcoords[3]
        #print(placer_coords)  
        if box['placer_font_name'] == '':  
            placer_font_name = fitz.Font(fontfile=dirFont+"arial.ttf") 
            # print(placer_font_name)
        else:
            placer_font_name = fitz.Font(fontfile=dirFont+box['placer_font_name'])


        if box['placer_font_name'] == '':  
            placer_font_name = "arial.ttf" 
        else:
            placer_font_name = box['placer_font_name']
        
        
        font_file_path = dirFont
        if placer_font_name == "Crashnumberinggothic_N.ttf":
            placer_font_name = fitz.Font(fontfile=font_file_path+"CrashNumberingGothic_N.otf")
        else:
            placer_font_name = fitz.Font(fontfile=font_file_path+placer_font_name)
            
        placer_font_underline = box['placer_font_underline'] 
        placer_font_size = box['placer_font_size'] 
        placer_font_size = float(placer_font_size) 
        placer_font_sizev1 = box['placer_font_size']              
        placer_type = box['placer_type']      
        placer_display = box['placer_display']
        if placer_display == '':
            placer_align = int(0)
        else:
            placer_align = int(placer_display)         
            
        if placer_type == 'Invisible':
            placer_color = yellow
        else:
            placer_color = black
        
        placer_degree_angle = box['degree_angle']
        if placer_degree_angle == '':
            placer_degree_angle = int(0)
        else:
            placer_degree_angle = int(placer_degree_angle)        

        placer_opacity = box['opacity_val']
        
        if placer_opacity is None:
            placer_opacity = 1.0

        if box['line_height'] == '':
            placer_lineHeight = 1
        else:
            placer_lineHeight = box['line_height']
        
        if "qr_place" in box:
            if box['qr_place'] == 'show':
                qr_show_flag = 1
            else:
                qr_show_flag = 0
        else:
            qr_show_flag = 1         
        
        if "blockchain_flag" in box:
            if box['blockchain_flag'] == 'use':
                blockchain_show_flag = 1
            else:
                blockchain_show_flag = 0
        else:
            blockchain_show_flag = 0         

        if "barcode_content" in box:
            if box['barcode_content'] == 'Source Content':
                barcode_content_flag = 1
            else:
                barcode_content_flag = 0
        else:
            barcode_content_flag = 0           

        if "barcode_content_position" in box:
            if box['barcode_content_position'] == 'Text at Bottom':
                barcode_content_position_flag = 1
            else:
                barcode_content_position_flag = 0
        else:
            barcode_content_position_flag = 0
        
        
        fontColor = box['font_color']
        search_word = "#"

        if search_word in fontColor:

            placer_font_color = hex_to_rgb(box['font_color'])

            # print(f"The word '{word}' is in the text.")
        else:
            if box['font_color'] == '':
                placer_font_color = black
            else:
                placer_font_color = box['font_color']
                placer_font_color=fitz.utils.getColor(placer_font_color)
        


        if placer_display == '':
            placer_align = int(0)
        else:
            placer_align = int(placer_display)         
            
        if placer_type == 'Invisible':
            placer_color = yellow
        else:
            placer_color = black
        
        placer_degree_angle = box['degree_angle']        
        placer_opacity = box['opacity_val']
        
        
        

        # print(placer_type)     
        if placer_type == 'QR Default': 
            if not os.path.exists(dirName):
                os.makedirs(dirName)
            now = datetime.datetime.now()            
            #dt_string = now.strftime("%Y-%m-%d %H:%M:%S")+str(cnt)
            dt_string = dt_string+str(cnt)
            result = hashlib.md5(dt_string.encode()) 
            barcode_en=result.hexdigest()           
            qr_txt=barcode_en
            if qr_show_flag==1:
                qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=4, border=0,)
                qr.add_data(qr_txt)
                qr.make(fit=True)
                img = qr.make_image()  # fill_color="black", back_color="white"
                img.save(dirName+"/"+"qr_"+str(barcode_en)+".png")
                qrcode_file=dirName+"/"+"qr_"+str(barcode_en)+".png"            
                page.insertImage(prect, qrcode_file, overlay=True) 
        elif placer_type == 'QR Dynamic': 
            if not os.path.exists(dirName):
                os.makedirs(dirName) 
            qr_txt=''
            now = datetime.datetime.now()          
            get_text=[]

            
            if box['source'] == '' or box['source'] == 'Current DateTime':
                #dt_string = now.strftime("%Y-%m-%d %H:%M:%S")+str(cnt)
                dt_string = dt_string+str(cnt)
                result = hashlib.md5(dt_string.encode()) 
                barcode_en=result.hexdigest()
            else:
                #dt_string = otxt.replace("/", "")+now.strftime("%Y-%m-%d %H:%M:%S")+str(cnt)                
                #dt_string = otxt.replace("/", "")
                #dt_string = otxt
                col_inx=keys.index(box['source'])
                otxt_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                get_text.append(otxt_string.replace("\n", ""))
                dt_string = get_text[0].replace("/", "").replace("\\", "").replace("-", "").replace(" ", "").strip()
                result = hashlib.md5(dt_string.encode()) 
                barcode_en=result.hexdigest()         
                
            qr_details = box['qr_details'].replace("{", "").replace("}", "")
            qr_list=list(filter(bool, qr_details.splitlines()))
            
            for n in qr_list:
                str_val = n.split("^")

                # col_inx=keys.index(str_val)
                # print(str_val[1])
                col_inx=keys.index(str_val[1])
                otxt_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                qr_txt += otxt_string +"\n"
                
            
            qr_txt += "\n" + barcode_en
            # qr_txt =  qr_txt + "\n"+barcode_en
            #print(qr_txt)
            if qr_show_flag==1:
                qr = qrcode.QRCode(version=1,error_correction=qrcode.constants.ERROR_CORRECT_L,box_size=4,border=0,)
                qr.add_data(qr_txt)
                qr.make(fit=True)
                img = qr.make_image()  # fill_color="black", back_color="white"
                img.save(dirName+"/"+"qr_"+str(barcode_en)+".png")
                qrcode_file=dirName+"/"+"qr_"+str(barcode_en)+".png"            
                page.insertImage(prect, qrcode_file, overlay=True)
        
        elif placer_type == 'Barcode': 
            #get_text=[]
            #get_text.append(page.getTextbox(srect).replace("\n", ""))
            #dt_strings = get_text[0].strip()
            if connection.is_connected():
                today = date.today()
                current_year=get_financial_year(str(today))
                current_year='PN/'+current_year+'/'
                cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
                record = cursor.fetchone() 
                next_print=record[0]+1
                next_print_serial=current_year+str(next_print)

            if barcode_content_flag==1:
                get_text=[]
                col_inx=keys.index(box['source'])
                otxt_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                get_text.append(otxt_string.replace("\n", ""))

                # get_text.append(page.getTextbox(srect).replace("\n", ""))
                temp = get_text[0].replace(" ", "").strip()
            else:
                temp = next_print_serial

            if barcode_content_position_flag==1:
                write_text_flag = True
            else:
                write_text_flag = False

            EAN = barcode.get_barcode_class('code128')
            ean = EAN(temp, writer=ImageWriter())
            if isinstance(prect.width, float)==True:
                bcwidth = float(prect.width)
            else:
                bcwidth = int(prect.width)
            
            if isinstance(prect.height, float)==True:
                bcheight = float(prect.height)
            else:
                bcheight = int(prect.height)
            
            rect = fitz.Rect(0, 0.85*bcheight, bcwidth, bcheight)
            #print(w,h) 300 dpi=667 mil
            options = {
                'dpi': 300,
                'write_text': write_text_flag,
                'module_width': bcwidth/667, #0.40
                'module_height': rect.height, #10
                'quiet_zone': 0,
                'text_distance': 1,
                'text_line_distance': 1,
                'font_size': placer_font_sizev1,
				'center_text':True
            }            
            barcode_file = ean.save(dirName, options = options)               
            #page.insertTextbox(prect, str(rect.width))           
            #page.drawRect( prect, color = green, fill = green)           
            im = Image.open(barcode_file)            
            imwidth, imheight = im.size            
            rectwidth = bcwidth*3.7795275591            
            if imwidth>rectwidth:
                page.insertImage(prect, barcode_file, overlay=True)
            else:
                width_in_mm = imwidth*0.2645833333
                newwidth = rectwidth-imwidth
                right_pos = newwidth*0.2645833333
                left_pos=int(float(pcoords[0]))
                top_pos=int(float(pcoords[1]))
                right_pos=int(float(pcoords[2])) - right_pos
                bottom_pos=int(float(pcoords[3]))               
                prects = fitz.Rect(left_pos,top_pos,right_pos,bottom_pos)          
                page.insertImage(prects, barcode_file, overlay=True)             
        

        elif placer_type == 'Static Text':
            otxt_string = str(box['qr_details'])
            chk_otxt_string = str(box['qr_details'])

            prects,prects_coords = text_tl_position(pcoords[0], pcoords[1], pcoords[2], pcoords[3])
            m = fitz.Matrix(0)
            points = fitz.Point(prects.x0, prects.y0)        
            

            if isinstance(placer_font_size, float)==True:
                placer_font_size = float(placer_font_size)
            else:
                placer_font_size = int(placer_font_size)

            # print(type(placer_font_size))
            if type(chk_otxt_string) is float:      
                page.insertTextbox(prect, otxt_string, fontfile=str(placer_font_name), fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)
            else:
                # print(placer_font_size)
                wr = fitz.TextWriter(page.rect)
                
                wr.fillTextbox(prect, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True, morph=(points,m))

            if placer_font_underline=="underline":
                rl = page.searchFor(otxt_string)  
                output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")                    
                ucoords = output.split(",")
                urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                shape = page.newShape()
                shape.drawLine(urect.bl, urect.br)
                shape.finish(color=placer_font_color, stroke_opacity=float(placer_opacity))
                shape.commit()


            # otxt_string = box['qr_details']
            # m = fitz.Matrix(0) #placer_degree_angle
            # points = fitz.Point(prect.x0, prect.y0)
            # wr = fitz.TextWriter(page.rect)
            # wr.fillTextbox(prect, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
            # wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))
            # if placer_font_underline=="underline":
            #     rl = page.searchFor(otxt_string)  
            #     output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")                    
            #     ucoords = output.split(",")
            #     urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
            #     shape = page.newShape()
            #     shape.drawLine(urect.bl, urect.br)
            #     shape.finish(color=placer_font_color, stroke_opacity=float(placer_opacity))
            #     shape.commit()



        elif placer_type == 'Plain Text':
                
            
            col_inx=keys.index(box['source'])
            
            otxt_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
            chk_otxt_string = sheet_obj.cell(row = cnt+1, column = col_inx+1).value
            
            # Check for None or invalid values and convert accordingly
            if otxt_string is None:
                otxt_string = ""
            



            # Load the font
            # font_path = dirFont+box['placer_font_name']
            if box['placer_font_name'] == '':  
                font_path = dirFont+"arial.ttf" 
            else:
                font_path = dirFont+box['placer_font_name']
            nxt_prt_width, nxt_prt_height = getTextlength_custom(otxt_string, font_path, int(placer_font_size))


            if(int(nxt_prt_width) > int(box['width'])):
                # calculate_coordinates
                
                nxt_prt_width_v1 = int(pcoords[0]) + int(nxt_prt_width)
                nxt_prt_height_v1 = int(pcoords[1])+ int(nxt_prt_height)
                
                prects,prects_coords = calculate_coordinates(pcoords[0], pcoords[1], nxt_prt_width_v1, pcoords[3])
            else:
                prects,prects_coords = text_tl_position(pcoords[0], pcoords[1], pcoords[2], pcoords[3])
            
            

            if placer_degree_angle:
                # Create an identity matrix and apply rotation
                m = fitz.Matrix(1, 0, 0, 1, 0, 0).preRotate(placer_degree_angle)
            else:
                # Fallback in case the angle is not defined (0 degrees rotation)
                m = fitz.Matrix(1, 0, 0, 1, 0, 0)
            # if placer_degree_angle:
            # m = fitz.Matrix(placer_degree_angle)
                
            points = fitz.Point(prects.x0, prects.y0)
            
            if otxt_string != 'None':

                if type(chk_otxt_string) is float:      
                    
                    page.insertTextbox(prects, otxt_string, fontfile=str(placer_font_name), fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)
                else:
                    
                    wr = fitz.TextWriter(page.rect)
                    wr.fillTextbox(prects, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True, morph=(points,m))

        elif placer_type == 'Micro Line': 
            mi_txt=''            
            if box['qr_details']=='':
                get_text=[]

                col_inx=keys.index(box['source'])
                otxt_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)

                chk_otxt_string = sheet_obj.cell(row = cnt+1, column = col_inx+1).value

                get_text.append(otxt_string.replace("\n", ""))
                # get_text.append(page.getTextbox(srect).replace("\n", ""))
                #text_len=len(temp)                
                temp = get_text[0].replace(" ", "").strip()
                search_text = get_text[0].strip()
                
                if placer_font_underline=="underline":
                    rl = doc.loadPage(cnt-1).searchFor(search_text, hit_max=1)
                    #print(rl[0])
                    #print(rl)
                    
                    if page_count==1:
                        output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                    else:
                        if len(rl) <=2:
                            output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                        else:
                            output = str(rl[len(rl)-1]).replace("Rect", "").replace("(", "").replace(")", "")                    
                    
                    #output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")                   
                    ucoords = output.split(",")
                    urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                    #print (urect) 
                    #print (urect.height) 
                    #urect.bl, urect.br
                    #prect_coords = pcoords[0],pcoords[1],pcoords[2],pcoords[3]
                    urect_coords = int(float(ucoords[1]))+int(float(urect.height))-1                    
                    new_urect=fitz.Rect(ucoords[0],urect_coords,ucoords[2],ucoords[3])
                    text_len=len(temp)
                    if isinstance(placer_font_size, float)==True:
                        textwidth = fitz.getTextlength(search_text, fontsize=float(placer_font_size))    
                    else:
                        textwidth = fitz.getTextlength(search_text, fontsize=int(placer_font_size))   
                                
                    #print(textwidth)
                    chrPerLine=int(float(new_urect.width))/textwidth
                    Totalchrs=str(chrPerLine).split(".")[0]
                    #print ("chr = ", chrPerLine)
                    #print (Totalchrs)
                    repeat_txt=temp * int(Totalchrs)
                   
                    if isinstance(placer_font_size, float)==True:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                    else:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                    
                    #print(int(float(new_urect.width)),"|",int(textwidth),"|",int(float(repeat_textwidth)))
                    remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))        
                    #print("rs: ",remain_space,"|",text_len)
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = temp[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''
                     
                    #print(remain_chrs)

                    # Load the font
                    # font_path = dirFont+box['placer_font_name']
                    if box['placer_font_name'] == '':  
                        font_path = dirFont+"arial.ttf" 
                    else:
                        font_path = dirFont+box['placer_font_name']

                    nxt_prt_width, nxt_prt_height = getTextlength_custom((temp * int(Totalchrs))+remain_chrs, font_path, int(placer_font_size))


                    if(int(nxt_prt_width) > int(box['width'])):
                        # calculate_coordinates
                        
                        nxt_prt_width_v1 = int(pcoords[0]) + int(nxt_prt_width)
                        nxt_prt_height_v1 = int(pcoords[1])+ int(nxt_prt_height)
                        
                        prects,prects_coords = calculate_coordinates(pcoords[0], pcoords[1], nxt_prt_width_v1, pcoords[3])
                    else:
                        prects,prects_coords = text_tl_position(pcoords[0], pcoords[1], pcoords[2], pcoords[3])


                    if 'is_repeat' in box:
                        if box['is_repeat'] == '1':
                            microTextStr = (temp * int(Totalchrs))+remain_chrs
                        else:
                            microTextStr = otxt_string
                    else:
                        # Handle the case where 'is_repeat' is missing
                        microTextStr = (temp * int(Totalchrs))+remain_chrs

                    wr = fitz.TextWriter(page.rect)
                    wr.fillTextbox(prects, microTextStr, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True) 
                    #page.drawRect(new_urect)
                else:

                    
                    text_len=len(temp)
                    if isinstance(placer_font_size, float)==True:
                        textwidth = fitz.getTextlength(temp, fontsize=float(placer_font_size))    
                    else:
                        textwidth = fitz.getTextlength(temp, fontsize=int(placer_font_size))   
                           
                    # print(placer_font_size)
                    chrPerLine=int(float(prect.width))/textwidth
                    Totalchrs=str(chrPerLine).split(".")[0]
                    #print ("chr = ", chrPerLine)
                    #print (Totalchrs)
                    repeat_txt=temp * int(Totalchrs)
                    if isinstance(placer_font_size, float)==True:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                    else:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                    
                    #print(int(float(prect.width)),"|",int(textwidth),"|",int(float(repeat_textwidth)))
                    remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3        
                    #print("rs: ",remain_space,"|",text_len)
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = temp[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''
                    

                    # Load the font
                    # font_path = dirFont+box['placer_font_name']
                    if box['placer_font_name'] == '':  
                        font_path = dirFont+"arial.ttf" 
                    else:
                        font_path = dirFont+box['placer_font_name']

                    nxt_prt_width, nxt_prt_height = getTextlength_custom((temp * int(Totalchrs))+remain_chrs, font_path, int(placer_font_size))


                    if(int(nxt_prt_width) > int(box['width'])):
                        # calculate_coordinates
                        
                        nxt_prt_width_v1 = int(pcoords[0]) + int(nxt_prt_width)
                        nxt_prt_height_v1 = int(pcoords[1])+ int(nxt_prt_height)
                        
                        prects,prects_coords = calculate_coordinates(pcoords[0], pcoords[1], nxt_prt_width_v1, pcoords[3])
                    else:
                        prects,prects_coords = text_tl_position(pcoords[0], pcoords[1], pcoords[2], pcoords[3])

                    #print(remain_chrs) # 0 = left, 1 = center, 2 = right
                    #page.insertTextbox(prect, (temp * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)
                    

                    if 'is_repeat' in box:
                        if box['is_repeat'] == '1':
                            microTextStr = (temp * int(Totalchrs))+remain_chrs
                        else:
                            microTextStr = otxt_string
                    else:
                        # Handle the case where 'is_repeat' is missing
                        microTextStr = (temp * int(Totalchrs))+remain_chrs

                    if type(chk_otxt_string) is float:      
                        page.insertTextbox(prects, microTextStr, fontfile=str(placer_font_name), fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)
                    else:
                        
                        wr = fitz.TextWriter(page.rect)
                        wr.fillTextbox(prects, microTextStr, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                        wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True, morph=(points,m))

                    # wr.fillTextbox(prect, (temp * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    # wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)    
            else:

                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                
                for n in qr_list:
                    str_val = n.split("^")

                    # col_inx=keys.index(str_val)
                    # print(str_val[0])
                    col_inx=keys.index(str_val[1])
                    mi_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                    mi_txt += mi_string +"\n"

                    # mi_txt = mi_txt + extract_microline_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                    
                #print(mi_txt)      
                
                #print(placer_font_name)
                if "Kruti" in box['placer_font_name']:
                    temp=mi_txt.replace(" ", "")
                    get_text=[]
                    get_text.append(page.getTextbox(srect).replace("\n", ""))
                    search_text = get_text[0].strip()
                    text_len=len(temp)                                   
                    text_to_cal=temp                                   
                else:
                    get_text=[]
                    col_inx=keys.index(box['source'])
                    otxt_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                    get_text.append(otxt_string.replace("\n", ""))
                    # get_text.append(page.getTextbox(srect).replace("\n", ""))
                    temp = get_text[0].replace(" ", "").strip()              
                    search_text = get_text[0].strip()
                    text_len=len(temp)
                    text_to_cal=search_text

                mi_txt = mi_txt.replace(" ", "")
                
                if placer_font_underline=="underline":
                    #rl = page.searchFor(search_text)  
                    rl = doc.loadPage(cnt-1).searchFor(search_text, hit_max=1) 
                    
                    if page_count==1:
                        output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                    else:
                        if len(rl) > 2:
                            output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                        else:
                            output = str(rl[len(rl)-1]).replace("Rect", "").replace("(", "").replace(")", "")
                    
                    #output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")                  
                    ucoords = output.split(",")
                    urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                    #print (urect, prect) 
                    #urect.bl, urect.br
                    #prect_coords = pcoords[0],pcoords[1],pcoords[2],pcoords[3]
                    urect_coords = int(float(ucoords[1]))+int(float(urect.height))-2                    
                    new_urect=fitz.Rect(ucoords[0],urect_coords,ucoords[2],ucoords[3])
                    
                    if isinstance(placer_font_size, float)==True:
                        textwidth = fitz.getTextlength(text_to_cal, fontsize=float(placer_font_size))    
                    else:
                        textwidth = fitz.getTextlength(text_to_cal, fontsize=int(placer_font_size))
                                
                    #print(textwidth,mi_txt)
                    chrPerLine=(int(float(new_urect.width))/textwidth)
                    Totalchrs=str(chrPerLine).split(".")[0]
                    #print ("chr = ", chrPerLine)
                    #print (Totalchrs)                    
                    repeat_txt=mi_txt * int(Totalchrs)
                    if isinstance(placer_font_size, float)==True:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))-10   
                    else:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                    
                    #print(int(float(new_urect.width)),"|",int(textwidth),"|",int(float(repeat_textwidth)))
                    remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))-3
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = mi_txt[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''                    
                    #exit()  
                    
                    wl_lst = fitz.getTextlength((mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size)+1
                    wl = sum([wl_lst])
                    if wl>new_urect.width:      
                        remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))-10     
                    else:       
                        remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))-3
                    
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = mi_txt[0:int(remain_chrs_count)]
                        #print("space|", text_len, remain_space, remain_chrs_count)
                    else:
                        remain_chrs = ''
                    
                    #print((mi_txt * int(Totalchrs))+remain_chrs)
                    if 'is_repeat' in box:
                        if box['is_repeat'] == '1':
                            microTextStr = (mi_txt * int(Totalchrs))+remain_chrs
                        else:
                            microTextStr = otxt_string
                    else:
                        # Handle the case where 'is_repeat' is missing
                        microTextStr = (mi_txt * int(Totalchrs))+remain_chrs

                    wr = fitz.TextWriter(page.rect)
                    if "Kruti" in box['placer_font_name']:
                        wr.fillTextbox(new_urect, microTextStr, fontsize=placer_font_size, align=placer_align)
                    else:
                        wr.fillTextbox(new_urect, microTextStr, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True) 
                    #page.drawRect(new_urect)
                else:
                    text_len=len(mi_txt)
                    if isinstance(placer_font_size, float)==True:
                        textwidth = fitz.getTextlength(mi_txt, fontsize=float(placer_font_size))    
                    else:
                        textwidth = fitz.getTextlength(mi_txt, fontsize=int(placer_font_size))   
                                
                    chrPerLine=(int(float(prect.width))/textwidth)
                    Totalchrs=str(chrPerLine).split(".")[0]
                    repeat_txt=mi_txt * int(Totalchrs)
                    if isinstance(placer_font_size, float)==True:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                    else:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                    
                          
                    remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3 
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = mi_txt[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''                    
                    #exit()                    
                    wl_lst = fitz.getTextlength((mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size)+1
                    wl = sum([wl_lst])
                    if wl>prect.width:      
                        remain_space=int(float(prect.width))-int(float(repeat_textwidth))-10     
                    else:       
                        remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3
                    
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = mi_txt[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''


                    if 'is_repeat' in box:
                        if box['is_repeat'] == '1':
                            microTextStr = (mi_txt * int(Totalchrs))+remain_chrs
                        else:
                            microTextStr = otxt_string
                    else:
                        # Handle the case where 'is_repeat' is missing
                        microTextStr = (mi_txt * int(Totalchrs))+remain_chrs

                    #page.insertTextbox(prect, (mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)
                    wr = fitz.TextWriter(page.rect)
                    wr.fillTextbox(prect, microTextStr, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)
        


        elif placer_type == 'Invisible':                        
            invisible_font_color=fitz.utils.getColor("YELLOW")
            inv_txt=''
            get_text=[]
            if box['qr_details']=='':

                col_inx=keys.index(box['source'])
                invtxt_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                chk_oinvtxt_string = sheet_obj.cell(row = cnt+1, column = col_inx+1).value

                get_text.append(invtxt_string.replace("\n", ""))
                otxt=get_text[0].strip()
                #print(otxt)
                #page.insertTextbox(prect, otxt, fontsize=placer_font_size, color=placer_color, align=placer_align, overlay=True)
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prect, otxt,  font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                #wr.writeText(page, color=placer_color, opacity=float(1.0), overlay=True)
                wr.writeText(page, color=invisible_font_color, opacity=float(1.0), overlay=True) 
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)

                    str_val = inv.split("^")
                    str_val_count=len(str_val)

                    col_inx=keys.index(str_val[1])
                    inv_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                    inv_txt += inv_string

                    # inv_txt = inv_txt + extract_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                          
                     
                #page.insertTextbox(prect, inv_txt, fontsize=placer_font_size, color=placer_color, align=placer_align, overlay=True)
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prect, inv_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                #wr.writeText(page, color=placer_color, opacity=float(1.0), overlay=True)
                wr.writeText(page, color=invisible_font_color, opacity=float(1.0), overlay=True) 
        
        elif placer_type == 'Image':
            #image_path=rootDir+sys.argv[7]+"/upload_images/" + box['image_path']
            # image_path=rootDir+sys.argv[7]+"/backend/templates/excel2pdf_images/" + box['image_path']
            image_path=rootDir+sys.argv[7]+"/backend/templates/excel2pdf/"+sys.argv[1]+'/images/' + box['image_path']
            #print(image_path)
            if os.path.exists(image_path):
                page.insertImage(prect, image_path,overlay=True)

        elif placer_type == 'Dynamic Image': 
            # temp_img = [block for block in page_dict['blocks'] if (fitz.Rect(block['bbox']) in srect and block['type'] == 1)]  
            
            col_inx=keys.index(box['source'])
            temp_img = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)


            # ImagedirName = rootDir+sys.argv[7]+'/'+"backend/templates/100/"
            # ImagedirName=rootDir+sys.argv[7]+"/backend/templates/excel2pdf_images/"
            ImagedirName=rootDir+sys.argv[7]+"/backend/templates/excel2pdf/"+sys.argv[1]+'/'

           
            # print(ImagedirName + temp_img)
            if temp_img != 'None':
                # print(temp_img)
                pix = fitz.Pixmap(ImagedirName + temp_img)
                if not os.path.exists(ImagedirName + temp_img):
                    os.makedirs(ImagedirName + temp_img)
                file_path = ImagedirName + temp_img
                # file_path2 = ImagedirName + "/"+str(int(round(time.time() * 1000))) + "-" + str(temp_img) + ".png"                        
                # pix.writeImage(file_path)
                # imgs = Image.open(file_path).convert("L") 
                # img1 = ImageOps.colorize(imgs, black ="white", white ="yellow")             
                # img1.save(file_path2, 'png')
                page.insertImage(prect, file_path, overlay=True)
                
        elif placer_type == 'Invisible Image': 
            # temp_img = [block for block in page_dict['blocks'] if (fitz.Rect(block['bbox']) in srect and block['type'] == 1)]  
            
            # col_inx=keys.index(box['source'])
            # temp_img = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)

            if box['source'] not in keys or not box['source']:
                temp_img = box['image_path']
            else: 
                col_inx=keys.index(box['source'])
                temp_img = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)

            # ImagedirName = rootDir+sys.argv[7]+'/'+"backend/templates/100/"
            # ImagedirName=rootDir+sys.argv[7]+"/backend/templates/excel2pdf_images/"
            ImagedirName=rootDir+sys.argv[7]+"/backend/templates/excel2pdf/"+sys.argv[1]+"/"

           
            # print(ImagedirName + temp_img)
            if temp_img != 'None':
                # print(temp_img)
                pix = fitz.Pixmap(ImagedirName + temp_img)
                if not os.path.exists(ImagedirName + temp_img):
                    os.makedirs(ImagedirName + temp_img)
                file_path = ImagedirName + temp_img
                file_path2 = ImagedirName + "/"+str(int(round(time.time() * 1000))) + "-" + str(temp_img) + ".png"                        
                pix.writeImage(file_path)
                imgs = Image.open(file_path).convert("L") 
                img1 = ImageOps.colorize(imgs, black ="white", white ="yellow")             
                img1.save(file_path2, 'png')
                page.insertImage(prect, file_path2, overlay=True) 

        elif placer_type == 'Watermark Text':           
            watermark_txt=''
            get_text=[]
            if box['qr_details']=='':   
                col_inx=keys.index(box['source'])
                otxt_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                get_text.append(otxt_string.replace("\n", ""))

                # get_text.append(page.getTextbox(srect).replace("\n", ""))
                otxt_string = get_text[0].strip()
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(prect_coords)
                new_rect=ir * m         
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(prect.x0, prect.y0)                 
                wr.fillTextbox(prect, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))                
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    col_inx=keys.index(str_val[1])
                    watermark_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                    watermark_txt += watermark_string
                
                # print(f"watermark_txt: {watermark_txt}, type: {type(watermark_txt)}")

                watermark_string = str(watermark_txt)
                chk_watermark_string = watermark_txt

                # print(f"watermark_txt: {watermark_txt}, type: {type(watermark_txt)}")
                
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(prect_coords)
                new_rect = ir * m
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(prect.x0, prect.y0)
                # try:
                #     wr.fillTextbox(prect, watermark_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                # except Exception as e:
                #     print(f"Error: {e}")
                #     # print(f"Error: {e}")
                
                # wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points, m))  
            
                page.insertTextbox(prect, watermark_string, fontfile=str(placer_font_name), fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)

                # wr.fillTextbox(prect, watermark_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                # wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))         
        
        elif placer_type == 'Watermark Multi Lines':  
            mi_txt=''            
            if box['qr_details']=='':
                get_text=[]

                col_inx=keys.index(box['source'])
                otxt_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                get_text.append(otxt_string.replace("\n", ""))

                # get_text.append(page.getTextbox(srect).replace("\n", ""))
                text_len=len(temp)
                temp = get_text[0].strip()+' '
                #chrPerLine=int(float(page_rect.width)+int(float(page_rect.height)))
                chrPerLine=int(390)
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=temp * int(Totalchrs)       
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(page_rect)
                new_page_rect=ir * m 
                #new_page_rect.y0=new_page_rect.y0-int(10)
                #new_page_rect.x0=new_page_rect.x0+int(200)
                new_page_rect.x1=new_page_rect.x1+int(100) 
                new_page_rect.y1=new_page_rect.y1+int(100)     
                wr = fitz.TextWriter(page.rect) #wr = fitz.TextWriter(page.rect)
                points = fitz.Point(0, 0)
                # wr.fillTextbox(new_page_rect, repeat_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                # wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))
                
                wr.fillTextbox(new_page_rect, repeat_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align, lineheight=placer_lineHeight)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=False, morph=(points,m))

                     
                   # wC=textsize(self, 'ABCD', font=None, *args, **kwargs)
                    #print(wC)
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)

                    str_val = inv.split("^")
                    str_val_count=len(str_val)

                    col_inx=keys.index(str_val[1])
                    watermark_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
                    mi_txt += watermark_string

                    # mi_txt = mi_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                    
                # print(mi_txt)      
                text_len=len(mi_txt)

                
                if isinstance(placer_font_size, float)==True:
                    textwidth = fitz.getTextlength(mi_txt, fontsize=float(placer_font_size))    
                else:
                    textwidth = fitz.getTextlength(mi_txt, fontsize=int(placer_font_size))   
                
                chrPerLine=(int(float(page_rect.width))/textwidth)
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=mi_txt * int(Totalchrs)
                if isinstance(placer_font_size, float)==True:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                else:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                
                      
                remain_space=int(float(page_rect.width))-int(float(repeat_textwidth))-3 
                if remain_space>0:                
                    RchrPerLine=int(remain_space)/textwidth
                    remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                    remain_chrs = mi_txt[0:int(remain_chrs_count)]
                else:
                    remain_chrs = ''                    
                #exit()                    
                wl_lst = fitz.getTextlength((mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size)+1
                wl = sum([wl_lst])

                # placer_lineHeight = 4

                # lines = repeat_txt.splitlines()

                if placer_lineHeight<3:
                    yOffset=600
                else:
                    yOffset=100
                
                pageW=float(page_rect.width)+int(60)
                pageH=float(page_rect.height)+int(yOffset)
                areaPage=int(pageW)*int(pageH)
                diagonal = round(math.sqrt((pageW**2) + (pageH**2)), 4)
                font = ImageFont.truetype('arial.ttf', placer_font_size)
                

                # m = fitz.Matrix(placer_degree_angle)
                # ir = fitz.IRect(page_rect)
                # new_page_rect = ir * m

                # Prepare the text writer
                # wr = fitz.TextWriter(page.rect)
                # points = fitz.Point(0, 0)  


                dummy_image = Image.new('RGB', (1, 1))
                draw = ImageDraw.Draw(dummy_image)

                # Calculate the size of the text
                bbox = draw.textbbox((0, 0), mi_txt, font=font)
                size = (bbox[2] - bbox[0], bbox[3] - bbox[1])

                # print(mi_txt)
                # size = font.getsize(mi_txt)
                diagonal=float(diagonal) 
                sizeWInPoints=float(size[0])*float(1.3333)
                charPerSingleLine=int(math.ceil(diagonal/sizeWInPoints))
                sizeHInPoints=float(size[1])*float(1.3333)
                areaPerLine=int(diagonal)* int(math.ceil(sizeHInPoints))
                totalRepeatText=int(math.ceil(areaPage/areaPerLine))
                
                # areaPage=int(float(page_rect.width)*int(float(page_rect.height)))
                # #font = ImageFont.truetype('arial.ttf', placer_font_size)
                # font = ImageFont.truetype('arial.ttf', 10)
                # size = font.getsize(mi_txt)
                # mmWidth= int(math.ceil(float(size[0])*float(0.26458333))) 
                # #charPerSingleLine=int(math.ceil(364/mmWidth))
                # charPerSingleLine=int(math.ceil(wl/mmWidth))
                # areaPerLine=int(mmWidth)*int(charPerSingleLine)*int(size[1])
                # totalRepeatText=int(math.ceil(areaPage/areaPerLine))

                #chrPerLine=int(364)
                #Totalchrs=str(chrPerLine).split(".")[0]
                #repeat_txt=mi_txt * int(Totalchrs)
                no_of_repeat=int(totalRepeatText) * int(charPerSingleLine)
                #no_of_repeat= int(no_of_repeat) + int(10) 
                repeat_txt=mi_txt * int(no_of_repeat)
                repeat_txt=repeat_txt
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(page_rect)
                new_page_rect=ir * m
                new_page_rect.x1=new_page_rect.x1 #+int(100) 
                new_page_rect.y1=new_page_rect.y1 #+int(100) 
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(0, 0)  #prect.x0, prect.y0

                # y_offset = 0
                # # Manually handle line height and draw text line by line
                # for line in lines:
                #     # Define the rectangle for each line
                #     line_rect = fitz.Rect(new_page_rect.x0, new_page_rect.y0 + y_offset, new_page_rect.x1, new_page_rect.y1 + y_offset)
                    
                #     # Fill the textbox with the current line
                #     wr.fillTextbox(line_rect, line, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    
                #     # Move down by the font size multiplied by the line height factor
                #     y_offset += placer_font_size * placer_lineHeight

                # # Write the text on the page
                # wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=False, morph=(points, m))

                # wr.fillTextbox(new_page_rect, repeat_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align, lineheight=placer_lineHeight)
                # wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=False, morph=(points,m))

                wr.fillTextbox(new_page_rect, repeat_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align, lineheight=placer_lineHeight)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=False, morph=(points,m))
                          
                #if sys.argv[7]=='demo':
                    #print(mi_txt)
                    #print(totalRepeatText)
                    #print(Totalchrs)
                #     font = ImageFont.truetype('arial.ttf', 12)
                #     size = font.getsize(mi_txt)
                #     print(mi_txt)
        

        elif placer_type == 'Ghost Image': 
            if placer_degree_angle==90:
                half_width=int(float(prect.height))/2
                top_pos_minus=(int(float(pcoords[0])) - int(float(half_width)))  
                add_right_pos=int(float(pcoords[0])) - top_pos_minus  
                left_pos=int(float(pcoords[0]))-add_right_pos
                top_pos=int(float(pcoords[1]))
                right_pos=int(float(pcoords[2])) + add_right_pos
                bottom_pos=int(float(pcoords[3]))               
                prects = fitz.Rect(left_pos,top_pos,right_pos,bottom_pos) 
                #print(prects)
                #print(prects.width,"| ",prects.height)                         
            else:
                prects=prect         

            get_text=[]

            col_inx=keys.index(box['source'])
            
            otxt_string = str(sheet_obj.cell(row = cnt+1, column = col_inx+1).value)
            chk_otxt_string = sheet_obj.cell(row = cnt+1, column = col_inx+1).value

            get_text.append(otxt_string.replace("\n", ""))
            # get_text.append(page.getTextbox(srect).replace("\n", ""))
            #temp = ''.join(c for c in get_text[0].strip().replace("/", "").replace(".", "") if c.isalnum())
            extracted_string = re.sub(r'\W+', '', get_text[0])
            temp = ''.join(c for c in extracted_string if c.isalnum())
            
            ghost_width = round(float(box['width']) * 3.7795275591)  # Ensure box['width'] is a float
            ghost_height = round(float(box['height']) * 3.7795275591)  # Ensure box['height'] is a float
            # ghost_width = round(box['width'] * 3.7795275591)  # Millimeter to Pixel, 1 mm = 3.7795275591 pixel
            # ghost_height = round(box['height'] * 3.7795275591)  
           
            ghost_words = box['ghost_words']  
            ghost_words = int(ghost_words)
            PrintableChars=temp[ 0 : ghost_words ] #extract first chars
            
            # print(PrintableChars)
            try:
                placer_degree_angle = int(placer_degree_angle)
            except ValueError:
                # print(f"Invalid value for placer_degree_angle: {placer_degree_angle}")
                placer_degree_angle = 0  # Assign a default value or handle the error as needed

            if not os.path.exists(dirName):
                os.makedirs(dirName)            
            ghostImg=CreateGhostImage(dirName, PrintableChars, placer_font_size, ghost_width, ghost_height)
            page.insertImage(prects, ghostImg,overlay=True, rotate=placer_degree_angle)

        elif placer_type == 'Box':  
            page.drawRect(prect, color=black, fill=None, dashes=None, width=0.5, lineCap=0, lineJoin=0, morph=None, overlay=True, stroke_opacity=1, fill_opacity=1)
        
        # if verification_setbg=='Yes':
        #     # for dpage in doc_new:
        #     page.insertImage(page_rect, verification_bg_file,overlay=False)

        

    
    single_page_doc.insertPDF(doc, from_page=i, to_page=i)
    output_file_path = rootDir + "/backend/pdf_file/" + dt_string + ".pdf"

    # Ensure there are pages before saving
    if len(single_page_doc) > 0:
        
        if verification_setbg=='Yes':
            for dpage in single_page_doc:
                dpage.insertImage(page_rect, verification_bg_file,overlay=False)  

        # Save the single-page PDF
        single_page_doc.save(output_file_path, garbage=4, deflate=True)

        # Optionally, move the file if needed
        final_path = rootDir + sys.argv[7] + "/backend/pdf_file/" + dt_string + ".pdf"
        shutil.move(output_file_path, final_path)

    
    # Close the single-page document
    single_page_doc.close()
    

    if print_setbg=='Yes':
        page.insertImage(page_rect, print_bg_file,overlay=False)      
    if connection.is_connected():
        sql = "INSERT INTO individual_records (file_records_id, template_id, template_name, pdf_page, page_no, encoded_id, unique_no, qr_details, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

        val = (file_records_next_id, template_id, template_name, pdf_page, cnt, barcode_en, dt_string, qr_txt, userid, record_unique_id)
        cursor.execute(sql, val)

    current_date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S');
    if connection.is_connected():
        cursor.execute("select * from student_table where serial_no = '%s' and status=1 order by id desc" % (dt_string))
        records = cursor.fetchall()  
        row_count = cursor.rowcount
        if row_count > 0: 
            sql = "UPDATE student_table SET status = %s, updated_by = %s, updated_at = %s WHERE serial_no = %s"
            val = ("0", userid, current_date_time, dt_string)
            cursor.execute(sql, val)   


        path="path"
        key="key"
        certificate_filename=dt_string+'.pdf'
        qr_path='qr/'+barcode_en+'.png'
        sql2 = """INSERT INTO student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        val2 = (dt_string, certificate_filename, template_id, barcode_en, qr_path, userid, userid, 1, 1, current_date_time,sys.argv[13],1)
        cursor.execute(sql2, val2) 
        student_table_id = cursor.lastrowid

        cursor.execute("select printer_name from system_config where site_id = '%s'" % (sys.argv[13]))
        recordSystem = cursor.fetchone()  
        printer_name=recordSystem[0]

        cursor.execute("SELECT * FROM printing_details WHERE sr_no = '%s' " % (dt_string))
        records = cursor.fetchall()  
        print_count = cursor.rowcount

        today = date.today()
        current_year=get_financial_year(str(today))
        current_year='PN/'+current_year+'/'
        cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
        record = cursor.fetchone() 
        next_print=record[0]+1
        next_print_serial=current_year+str(next_print)
        sql3 = """INSERT INTO printing_details (`username`, `print_datetime`, `printer_name`, `print_count`, `print_serial_no`, `sr_no`, `template_name`,`created_at`, `created_by`, `updated_at`, `updated_by`,`status`, `site_id`, `publish`, `student_table_id`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        val3 = (sys.argv[14], current_date_time, printer_name, print_count, next_print_serial, dt_string, template_name, current_date_time, userid, current_date_time,userid, 1, sys.argv[13], 1,student_table_id)
        cursor.execute(sql3, val3)



    datetime_IND = datetime.datetime.now(tz_IND)
    ending_time = datetime_IND.strftime("%H:%M:%S")
    end_time=time.time() - start_time
    seconds_to_hhmmss=time.strftime('%H:%M:%S', time.gmtime(end_time))
    page_seconds_to_hhmmss=time.strftime('%H:%M:%S', time.gmtime(end_time/cnt))
    arr_content['percent'] = int(cnt/page_count * 100)
    arr_content['message'] = "Generating "+str(cnt)+"/"+str(page_count)+" PDF(s)"
    arr_content['beginning_time'] = beginning_time
    arr_content['ending_time'] = ending_time
    arr_content['exec_time'] = end_time
    arr_content['hms_time'] = seconds_to_hhmmss
    arr_content['page_time'] = page_seconds_to_hhmmss
    arr_content['pages_processed'] = cnt
    json_object = json.dumps(arr_content, indent = 4)

    f = open(rootDir+sys.argv[7]+'/'+"processed_pdfs/"+sys.argv[5], "w")
    f.write(json_object)

    column += 1
    cnt += 1

if connection.is_connected():
    sql = "INSERT INTO file_records (template_id, template_name, pdf_page, total_records, source_file, userid, record_unique_id,map_type) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
    val = (template_id, template_name, pdf_page, cnt-1, sys.argv[2], userid, record_unique_id,1)
    cursor.execute(sql, val)

    connection.commit()
# print(path_pdf_moved)

removePath=rootDir+sys.argv[7]+'/'
#print(removePath)
rewisePath = path_pdf_moved.replace(removePath,'')

doc.save(path_pdf_moved, garbage=4, deflate=True)
#doc.save(path_pdf_moved, garbage=4, deflate=True, owner_pw="owner", encryption=encrypt_meth, permissions=perm)
#doc.save(path_pdf_moved, incremental=True)
#shutil.copyfile(output_file, path_pdf_moved)
# print(rewisePath)
#shutil.make_archive(folder+"/"+template_name+"-"+str(record_unique_id), "zip", inner_folder)
#print(folder+"/"+template_name+"-"+str(record_unique_id)+".zip")
total_pages=cnt-1
#print("Total Records:"+str(total_pages))
shutil.rmtree(dirName, ignore_errors=True)

fileName = sys.argv[2].replace(".xlsx", "")



doc.save(inner_folder+"/"+fileName+".pdf", garbage=4, deflate=True)

pdfPath = inner_folder+"/"+fileName+".pdf"


rewisePath = pdfPath.replace(removePath,'')

print(rewisePath)
exit()

