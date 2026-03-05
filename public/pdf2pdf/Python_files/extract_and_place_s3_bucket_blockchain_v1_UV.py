#!C:/Users/Administrator/AppData/Local/Programs/Python/Python38/python.exe
from __future__ import print_function
from operator import itemgetter
from itertools import groupby
import sys
import os
import shutil
from shutil import make_archive
from turtle import update
import simplejson as json
import fitz
import mysql.connector
import pprint
import time
from mysql.connector import Error
import textwrap
from PIL import Image, ImageOps 
import numpy as np
#import cv2.cv2 as cv2
import barcode
from barcode.writer import ImageWriter
import qrcode
import hashlib 
#from datetime import datetime
from xml.dom.minidom import parseString
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font 
from openpyxl.styles import Alignment
#from datetime import datetime
import uuid
import pytz
#from PyPDF2 import PdfFileWriter, PdfFileReader
import socket
import requests
import datetime
from datetime import date
from PIL import ImageDraw, ImageFont
import math
import subprocess
import tempfile
import base64
from urllib.parse import unquote
from boto3.s3.transfer import S3Transfer
import boto3
import re
import hmac

tz_IND = pytz.timezone('Asia/Calcutta') 
#sys.argv[1] = template id
#sys.argv[2] = data file
#sys.argv[3] = session user id
#sys.argv[4] = entry type (Fresh/Proceed)
#sys.argv[5] = progress file
#sys.argv[6] = dbName
#sys.argv[7] = subdomain
#sys.argv[8] = directoryUrlForward
#sys.argv[9] = directoryUrlBackward
#sys.argv[10] = servername
#sys.argv[11] = username
#sys.argv[12] = password
#sys.argv[13] = siteid
#sys.argv[14] = username
#sys.argv[15] = printer_name
# python extract_and_place.py 1 FYBAF_1619953609.pdf 1
#print(sys.argv[2])
sys.argv[12]=unquote(unquote(sys.argv[12]))
try:
    directory=sys.argv[8]
    rootDir= directory.replace('pdf2pdf', '')
    connection = mysql.connector.connect(host=sys.argv[10],
                                         database=sys.argv[6],
                                         user=sys.argv[11],
                                         password=sys.argv[12])
    if connection.is_connected():
        db_Info = connection.get_server_info()
        cursor = connection.cursor(buffered=True)
        cursor.execute("select ep_details, id, extractor_details, template_name, pdf_page, print_bg_file, print_bg_status, verification_bg_file, verification_bg_status, IFNULL(bc_contract_address, '') AS bc_contract_address, bc_document_description, bc_document_type from uploaded_pdfs where id = '%s'" % (sys.argv[1]))
        record = cursor.fetchone()        
        boxes = json.loads(record[0], strict=False)        
        template_id=record[1]
        extractor_details = json.loads(record[2], strict=False)
        template_name=record[3]
        pdf_page=record[4]
        pbg_file=record[5]
        print_bg_status=record[6]
        vbg_file=record[7]
        verification_bg_status=record[8]
        bc_contract_address=record[9]
        bc_document_description=record[10]
        bc_document_type=record[11]
        cur=connection.cursor(buffered=True)
        print_setbg=''
        verification_setbg=''
        if pbg_file !=0 and print_bg_status == 'Yes':
            sql_bg="SELECT image_path FROM background_template_master where id= '%s'" % (pbg_file)
            cur.execute(sql_bg)
            precord = cur.fetchone()
            #print_bg_file=sys.argv[8]+"upload_bgs/"+precord[0] 
            print_bg_file=rootDir+sys.argv[7]+"/backend/canvas/bg_images/" + precord[0]
            #print_bg_file="C:/wamp/www/demo/upload_bgs/"+precord[0] 
            print_setbg='Yes'
            #print(print_bg_file)
        if vbg_file !=0 and verification_bg_status == 'Yes':
            sql_bg="SELECT image_path FROM background_template_master where id= '%s'" % (vbg_file)
            cur.execute(sql_bg)
            vrecord = cur.fetchone()
            #verification_bg_file=sys.argv[8]+"upload_bgs/"+vrecord[0]
            verification_bg_file=rootDir+sys.argv[7]+"/backend/canvas/bg_images/" + vrecord[0]
            #verification_bg_file="C:/wamp/www/demo/upload_bgs/"+vrecord[0]
            verification_setbg='Yes'
            #print(verification_bg_file)

        record_unique_id = datetime.datetime.now().strftime('%Y%m%d%H%M%S-') + str(uuid.uuid4()).split('-')[-1]
        
        #sqli="SELECT id FROM file_records ORDER BY id DESC LIMIT 1"
        # Update rohit 01/09/2023
        #print(sys.argv[17],sys.argv[18])
        if sys.argv[18] == '1':
            sqli="SELECT id FROM sb_file_records ORDER BY id DESC LIMIT 1"
        else:
            sqli="SELECT id FROM file_records ORDER BY id DESC LIMIT 1"
        #Update Rohit 01/09/2023
        cur.execute(sqli)
        last_id = cur.fetchone()
        if last_id == None:
            file_records_next_id=1
        else:
            file_records_next_id=last_id[0]+1    
        # get last record's columns
        #for lastID in last_id:
            #print(lastID)     

except Error as e:
    print("Error while connecting to MySQL", e)
"""
finally:
    if (connection.is_connected()):
        #cursor.close()
        #connection.close()
        print("MySQL connection is closed")
"""
#exit()
#print(bc_contract_address)
userid=sys.argv[3]


bc_sc_id = None
#function to link student to bc_smart_contract id 
def update_student(student_id, bc_sc_id):
    try:
        # Update the student's bc_sc_id
        update_student_query = """
        UPDATE student_table
        SET bc_sc_id = %s
        WHERE id = %s;
        """
        cursor.execute(update_student_query, (bc_sc_id, student_id))

        # Increment the count in bc_smart_contracts
        update_bc_smart_contracts_query = """
        UPDATE bc_smart_contracts
        SET count = count + 1
        WHERE id = %s;
        """
        cursor.execute(update_bc_smart_contracts_query, (bc_sc_id,))

        # Commit the transaction
        connection_v4.commit()
        print("Student and smart contract records updated successfully.")

    except mysql.connector.Error as err:
        print(f"An error occurred: {err}")


#function to create new contract
def deploy_contract_old(mode='0'):
    responseArr = {}
    
    try:
        # Determine the request URL based on the mode (Testnet or Live)
        # request_url = (
        #     'https://veraciousapis.herokuapp.com/v1/deployContract'
        #     if mode == '0'
        #     else 'https://mainnet-apis.herokuapp.com/v1/mainnet/deployContract'
        # )

        if(mode==0):
            #Testnet Url
            request_url='https://veraciousapis.herokuapp.com/v1/deployContract';
        else:
            #Live Url
            request_url='https://mainnet-apis.herokuapp.com/v1/mainnet/deployContract';
        


        # Send the GET request to the API
        res = requests.get(request_url)

        # Check if the response status code is 200 (OK)
        if res.status_code == 200:
            response = res.json()
            if response.get('success') == 200:
                contract_address = response.get('contractAddress')
                responseArr['status'] = 200
                responseArr['message'] = "Contract successfully deployed."
                responseArr['contractAddress'] = contract_address
            else:
                responseArr['status'] = response.get('success', 400)
                responseArr['message'] = response.get('message', "Something went wrong.")
        else:
            responseArr['status'] = res.status_code
            responseArr['message'] = "Failed to deploy contract."

    except Exception as e:
        responseArr['message'] = str(e)
        responseArr['status'] = 400

    return responseArr


def deploy_contract(mode='0'):
    try:
        # Determine the request URL based on the mode
        request_url = (
            'http://node.seqrdoc.com:9090/testnet/deploy_contract'
            if mode == '0'
            else 'http://node.seqrdoc.com:9090/mainnet/deploy_contract'
        )

        # Prepare the JSON data
        payload = {
            'contract_name': 'scube',
            'contract_symbol': 'string',
        }

        # Headers
        headers = {
            'Accept': 'application/json',
            'Content-Type': 'application/json',
            'x-api-key': '123456789',
        }

        # Send POST request
        res = requests.post(request_url, json=payload, headers=headers)

        # Prepare response structure
        responseArr = {}

        # Handle response
        if res.status_code == 200:
            response = res.json()
            if response.get('status') == 200:
                responseArr['status'] = 200
                responseArr['message'] = response.get('message', 'Success')
                responseArr['contractAddress'] = response.get('contractAddress')
            else:
                responseArr['status'] = response.get('status', 400)
                responseArr['message'] = response.get('message', 'Something went wrong.')
        else:
            responseArr['status'] = res.status_code
            responseArr['message'] = "Failed to deploy contract."

    except Exception as e:
        responseArr = {
            'status': 400,
            'message': str(e)
        }

    return responseArr

#function to check smart contract with count if count exceed then new smart contract will create
def checkContract(wallet_address,template_id, limit=100):
    try:
        # # Establish connection using command-line arguments
        # connection = mysql.connector.connect(
        #     host=sys.argv[10],
        #     database=sys.argv[6],
        #     user=sys.argv[11],
        #     password=sys.argv[12]
        # )
        # cursor = connection.cursor()
        # Establish connection using command-line arguments
        connection3 = mysql.connector.connect(
            host=sys.argv[10],
            database=sys.argv[6],
            user=sys.argv[11],
            password=sys.argv[12]
        )
        cursor_bc = connection3.cursor()

        # Check the current count of the smart contract
        check_count_query = """
        SELECT count, wallet_address,smart_contract_address,id FROM bc_smart_contracts 
        WHERE template_type = 1 AND template_id = %s AND is_active = 1;
        """
        cursor_bc.execute(check_count_query, (template_id,))
        connection3.commit()
        current_contract = cursor_bc.fetchone()


        # print('limit ' + str(limit))

        if current_contract is not None and (current_contract[0] < limit):
            # print(current_contract[0])
            # print('Contarct Exist')
            # Retrieve the wallet address from the current active contract
            wallet_address = current_contract[1]
            contract_address = current_contract[2]
            bc_sc_id = current_contract[3]

            return {
                'contract_address': contract_address,
                'wallet_address': wallet_address,
                'bc_sc_id': bc_sc_id
            }
        
        else:
            
            # print('Deplot Contarct')
            # Deploy the new contract via the blockchain API
            blockchain_response = deploy_contract('1')
            # blockchain_response = {'status': 200, 'message': 'Contract Successfully deployed', 'contractAddress': '0xE8db7cABE2D28F9AfFE3a062583108a0Fa901A10'}
            if blockchain_response.get('status') == 200:
                contract_address = blockchain_response.get('contractAddress')

                # Mark existing contracts for this template as inactive
                deactivate_existing_contracts_query = """
                UPDATE bc_smart_contracts
                SET is_active = 0
                WHERE template_id = %s AND template_type = 1;
                """
                cursor_bc.execute(deactivate_existing_contracts_query, (template_id,))

                # Create a new entry for the smart contract in bc_smart_contracts table
                # insert_new_contract_query = """
                # INSERT INTO bc_smart_contracts 
                # (count, template_id, smart_contract_address, wallet_address, template_type, is_active, is_live)
                # VALUES (%s, %s, %s, %s, %s, %s, %s);
                # """
                # cursor.execute(
                #     insert_new_contract_query,(0, template_id, contract_address, wallet_address, 1, 1, 1)
                # )

                insert_new_contract_query = """
                    INSERT INTO bc_smart_contracts 
                    (count, template_id, smart_contract_address, wallet_address, template_type, is_active, is_live)
                    VALUES (%s, %s, %s, %s, %s, %s, %s);
                    """
                cursor_bc.execute(insert_new_contract_query, (0, template_id, contract_address, bc_wallet_address, 1, 1, 1))
                connection3.commit()


                # Check if there is an active smart contract for the given template_id
                check_contract_query = """
                SELECT smart_contract_address, wallet_address ,id
                FROM bc_smart_contracts
                WHERE template_type = 1 AND template_id = %s AND is_active = 1;
                """
                cursor_bc.execute(check_contract_query, (template_id,))
                fetchContract = cursor_bc.fetchone()

                if fetchContract:
                    # If there's an active contract, extract the details
                    contract_address, wallet_address, bc_sc_id = fetchContract

                    return {
                        'contract_address': contract_address,
                        'wallet_address': wallet_address,
                        'bc_sc_id': bc_sc_id
                    }
                else:
                    return {
                        'contract_address': '',
                        'wallet_address': '',
                        'bc_sc_id': ''
                    }
                # Commit the transaction to save the changes
                # connection3.commit()
                # print("New contract deployed, added to the database, and existing contracts marked as inactive.")
            else:
                # return {
                #     'contract_address': '',
                #     'wallet_address': '',
                #     'bc_sc_id': ''
                # }
                # print(f"Failed to deploy new contract. Error: {blockchain_response.get('message')}")
                return None

    except mysql.connector.Error as err:
        print(f"An error occurred: {err}")

#function to return active smart contract address and wallet id 
def checkActiveContract(template_id):

    # print('Check Active Contract')

    try:
        # Check if there is an active smart contract for the given template_id
        check_contract_query = """
        SELECT smart_contract_address, wallet_address,id 
        FROM bc_smart_contracts
        WHERE template_type = 1 AND template_id = %s AND is_active = 1;
        """
        cursor.execute(check_contract_query, (template_id,))
        activeContract = cursor.fetchone()

        
        if activeContract:
            # If there's an active contract, extract the details

            # print(activeContract)
            contract_address, wallet_address,bc_sc_id = activeContract
            # print(f"Active contract found: Contract Address: {contract_address}, Wallet Address: {wallet_address}")
            return {
                'contract_address': contract_address,
                'wallet_address': wallet_address,
                'bc_sc_id': bc_sc_id
            }
        # else:
            # print(f"No active contract found for template_id: {template_id}")
            # return None

    except mysql.connector.Error as err:
        print(f"An error occurred: {err}")
        return None


def createNewContract(bc_wallet_address,template_id):


    try:
        # Establish connection using command-line arguments
        connection3 = mysql.connector.connect(
            host=sys.argv[10],
            database=sys.argv[6],
            user=sys.argv[11],
            password=sys.argv[12]
        )
        cursor_bc = connection3.cursor()
        
        # print('create new contract')
        blockchain_response = deploy_contract('1')
        # blockchain_response = {'status': 200, 'message': 'Contract Successfully deployed', 'contractAddress': '0xcF254CAbDc050600c56b77290Ebc8FBA23d1F64B'} 
        if blockchain_response.get('status') == 200:
            
            contract_address = blockchain_response.get('contractAddress')
            # Create a new entry for the smart contract in bc_smart_contracts table

            try:
                insert_new_contract_query = """
                    INSERT INTO bc_smart_contracts 
                    (count, template_id, smart_contract_address, wallet_address, template_type, is_active, is_live)
                    VALUES (%s, %s, %s, %s, %s, %s, %s);
                    """
                cursor_bc.execute(insert_new_contract_query, (0, template_id, contract_address, bc_wallet_address, 1, 1, 1))
                connection3.commit()  # Ensure changes are saved
            except Exception as e:
                print("DB Error:", e)
            # insert_new_contract_query = """
            # INSERT INTO bc_smart_contracts 
            # (count, template_id, smart_contract_address_, wallet_address, template_type, is_active, is_live)
            # VALUES (%s, %s, %s, %s, %s, %s, %s);
            # """
            # cursor_bc.execute(
            #     insert_new_contract_query,
            #     (0, template_id, contract_address, bc_wallet_address, 1, 1, 1)
            # )
            
            # print(blockchain_response)
            # exit()
            bc_sc_id = cursor_bc.lastrowid
            # Check if there is an active smart contract for the given template_id
            # check_contract_query = """ SELECT smart_contract_address, wallet_address ,id FROM bc_smart_contracts WHERE template_type = 1 AND template_id = %s AND is_active = 1;
            # """


            get_bc_query ="SELECT smart_contract_address, wallet_address ,id FROM bc_smart_contracts where template_type = 1 AND is_active = 1 AND template_id= '%s'" % (template_id)
            cursor_bc.execute(get_bc_query)
            fetchContract = cursor_bc.fetchone()
            
            # cursor.execute(check_contract_query, (template_id,))
            # fetchContract = cursor.fetchone()
            
            if fetchContract:
                # If there's an active contract, extract the details
                contract_address = fetchContract[0] 
                wallet_address = fetchContract[1]
                bc_sc_id = fetchContract[2]

            return {
                'contract_address': contract_address,
                'wallet_address': wallet_address,
                'bc_sc_id': bc_sc_id
            }

            # Commit the transaction to save the changes
            connection3.commit()
            # print('200')
            # print("New contract deployed, added to the database, and existing contracts marked as inactive.")
        else:
            print('blockchain not working') 
    except mysql.connector.Error as err:
        print(f"An error occurred: {err}")
        if connection3.is_connected():
            connection3.rollback()

    finally:
        # Close the connection3
        if connection3.is_connected():
            cursor_bc.close()
            connection3.close()

#function to check smart contract with count if count exceed then new smart contract will create
def getIdOfContract(template_id):
    try:
        # Establish connection using command-line arguments
        connection3 = mysql.connector.connect(
            host=sys.argv[10],
            database=sys.argv[6],
            user=sys.argv[11],
            password=sys.argv[12]
        )
        cursor_bc = connection3.cursor()

        # print(sys.argv[10])
        # print(sys.argv[6])
        # print(sys.argv[11])
        # print(sys.argv[12])
        # sql_bg="SELECT id FROM bc_smart_contracts where template_type = 1 AND is_active = 1 AND template_id= '%s'" % (template_id)
        # cursor.execute(sql_bg)
        # current_contract = cur.fetchone()

        get_bc_query ="SELECT id FROM bc_smart_contracts where template_type = 1 AND is_active = 1 AND template_id= '%s'" % (template_id)
        cursor_bc.execute(get_bc_query)
        current_contract = cursor_bc.fetchone()


        # print(current_contract)
        if current_contract:
            
            # Retrieve the wallet address from the current active contract
            bc_sc_id = current_contract[0]
        
        else:
            bc_sc_id = 0
    except mysql.connector.Error as err:
        print(f"An error occurred: {err}")
        if connection3.is_connected():
            connection3.rollback()

    finally:
        # Close the connection3
        if connection3.is_connected():
            cursor_bc.close()
            connection3.close()



def extract_microline_info(extractor_detail,search_value,extra_line_first='',extra_line_second='',page=''):
    for keyval in extractor_detail:
        if search_value.lower() == keyval['name'].lower():
            source_coord = keyval['coords']
            scoord = source_coord.split(",")
            s_rect = fitz.Rect(scoord[0],scoord[1],scoord[2],scoord[3])
            get_text=[]            
            get_text.append(page.getTextbox(s_rect).replace("\n", ""))
            otxt=get_text[0].replace(" ", "")
            if extra_line_first == '' and extra_line_second == '':
                return otxt
            elif extra_line_first != '' and extra_line_second != '':
                return extra_line_first + otxt + extra_line_second 
            elif extra_line_first != '' and extra_line_second == '':
                return extra_line_first + otxt 
            elif extra_line_first == '' and extra_line_second != '':
                return otxt + extra_line_second

def extract_plainText(extractor_detail,search_value,extra_line_first='',extra_line_second='',page=''):
    for keyval in extractor_detail:
        if search_value.lower() == keyval['name'].lower():
            source_coord = keyval['coords']
            scoord = source_coord.split(",")
            s_rect = fitz.Rect(scoord[0],scoord[1],scoord[2],scoord[3]) 
            get_text=[]            
            get_text.append(page.getTextbox(s_rect).replace("\n", ""))
            otxt=get_text[0].strip()
            if extra_line_first == '' and extra_line_second == '':
                return otxt +" "
            elif extra_line_first != '' and extra_line_second != '':
                return extra_line_first + otxt + extra_line_second +" " 
            elif extra_line_first != '' and extra_line_second == '':
                return extra_line_first + otxt +" " 
            elif extra_line_first == '' and extra_line_second != '':
                return otxt + extra_line_second +" "                 
                
def extract_info(extractor_detail,search_value,extra_line_first='',extra_line_second='',page=''):
    for keyval in extractor_detail:
        if search_value.lower() == keyval['name'].lower():
            source_coord = keyval['coords']
            scoord = source_coord.split(",")
            s_rect = fitz.Rect(scoord[0],scoord[1],scoord[2],scoord[3])
            get_text=[]            
            get_text.append(page.getTextbox(s_rect).replace("\n", ""))
            otxt=get_text[0].strip()
            if extra_line_first == '' and extra_line_second == '':
                return otxt +"\n"
            elif extra_line_first != '' and extra_line_second != '':
                return extra_line_first + otxt + extra_line_second +"\n" 
            elif extra_line_first != '' and extra_line_second == '':
                return extra_line_first + otxt +"\n" 
            elif extra_line_first == '' and extra_line_second != '':
                return otxt + extra_line_second +"\n" 

def CreateGhostImage(dirName, name, p_font_size, ghost_width, ghost_height):
    #dirChars = "F:/projects/flask-app-env/marksheet/grade_card_reader/defence_secure_docs/chars/"+str(p_font_size)
    #dirChars = "C:/Program Files/Python38/projects/demo/chars/"+str(p_font_size)
    dirChars=sys.argv[8]+'Python_files/chars/'+str(p_font_size);
    name=name.upper()    
    single_char=split(name)       
    my_list = list()
    for c in single_char:
        my_list.append(dirChars +"/"+ c +".png")    
    #print(my_list)
    images = [Image.open(x) for x in my_list]
    widths, heights = zip(*(i.size for i in images))
    total_width = sum(widths)
    max_height = max(heights)    
    new_im = Image.new('RGB', (total_width, max_height))
    x_offset = 0
    for im in images:
      new_im.paste(im, (x_offset,0))
      x_offset += im.size[0]

    new_im.save(dirName +"/"+ name +".png")  
    isize=ghost_width,ghost_height            
    im = Image.open(dirName +"/"+ name +".png")
    im.thumbnail((isize), Image.ANTIALIAS)
    im.save(dirName +"/"+ name +str(p_font_size)+"_th.png", quality=100)    
    #return dirName +"/"+ name +".png" 
    return dirName +"/"+ name +str(p_font_size)+"_th.png"     
    
def split(word): 
    return [char for char in word]

def get_financial_year(datestring):
            date = datetime.datetime.strptime(datestring, "%Y-%m-%d").date()
            #initialize the current year
            year_of_date=date.year
            #initialize the current financial year start date
            financial_year_start_date = datetime.datetime.strptime(str(year_of_date)+"-04-01","%Y-%m-%d").date()
            if date<financial_year_start_date:
                    return str(financial_year_start_date.year-1)[2:]+'-'+ str(financial_year_start_date.year)[2:]
            else:
                    return str(financial_year_start_date.year)[2:]+'-'+ str(financial_year_start_date.year+1)[2:]    

def repeat_to_length(string_to_expand, length):
    return (string_to_expand * (int(length/len(string_to_expand))+1))[:length]

def get_pil_text_size(text, font_size, font_name):
    font = ImageFont.truetype(font_name, font_size)
    size = font.getsize(text)
    return size


def set_background(input_pdf, output, watermark):
    watermark_obj = PdfFileReader(watermark)
    #watermark_obj.decrypt("owner")
    watermark_page = watermark_obj.getPage(0)

    pdf_reader = PdfFileReader(input_pdf)    
    pdf_writer = PdfFileWriter()

    for page in range(pdf_reader.getNumPages()):        
        page = pdf_reader.getPage(page)
        page.mergePage(watermark_page)
        #page.compressContentStreams()
        pdf_writer.addPage(page)

    with open(output, 'wb') as out:
        pdf_writer.write(out)


def generate_file_hash(file_url, salt):
    # Download file from URL
    response = requests.get(file_url)
    if response.status_code != 200:
        raise Exception(f"Failed to download file: {file_url}")

    file_content = response.content

    hmac_hash = hmac.new(
        key=salt.encode('utf-8'),
        msg=file_content,
        digestmod=hashlib.sha3_256
    ).hexdigest()

    return hmac_hash


encrypt_meth = fitz.PDF_ENCRYPT_AES_256  # strongest algorithm
perm = int(
fitz.PDF_PERM_PRINT  # permit printing
)

dirFont = sys.argv[8]+"Python_files/fonts/"
#directory="C:/wamp/www/demo/"
#dirFont = "C:/Program Files/Python38/projects/demo/fonts/"

#doc = fitz.open(rootDir+sys.argv[7]+'/uploads/data/'+sys.argv[2])#directory+"uploads/data/"+
#Updated rohit 01/09/2023
if sys.argv[18] == '1':
    doc = fitz.open(rootDir+sys.argv[7]+'/uploads/data/sandbox/'+sys.argv[2])#directory+"uploads/data/sandbox/"+
else:
    doc = fitz.open(rootDir+sys.argv[7]+'/uploads/data/'+sys.argv[2])#directory+"uploads/data/"+
    
page_count=doc.pageCount
arr_content = {} #The array for storing the progress.

#Check print limit 
if connection.is_connected():
    #cursor.execute("select count(*) AS ts from student_table where site_id = '%s'" % (sys.argv[13]))
    # Update Rohit 01/09/2023
    if sys.argv[18] == '1':
        cursor.execute("select count(*) AS ts from sb_student_table where site_id = '%s'" % (sys.argv[13]))
    else:
        cursor.execute("select count(*) AS ts from student_table where site_id = '%s'" % (sys.argv[13]))
    # Update Rohit 01/09/2023
    rsStudent = cursor.fetchone()
    studentTableCounts=rsStudent[0]

connection2 = mysql.connector.connect(host=sys.argv[10],
                                     database='seqr_demo',
                                     user=sys.argv[11],
                                     password=sys.argv[12])
if connection2.is_connected():
    cursor2 = connection2.cursor(buffered=True)
    cursor2.execute ("UPDATE super_admin SET current_value='%s' WHERE site_id=%s " % (studentTableCounts, sys.argv[13]))    
    cursor2.execute("select value, current_value from super_admin where site_id = '%s'" % (sys.argv[13]))
    rsGenerated = cursor2.fetchone()  
    printLimit=int(rsGenerated[0])
    currentValue=int(rsGenerated[1])
    recordGenerated= currentValue + int(page_count)
    noOfCertificateCanGenerate=printLimit-currentValue

    siteurl_param=sys.argv[7]+".seqrdoc.com"
    #siteurl_param="demo.seqrdoc.com"
    cursor2.execute("select IFNULL(bc_wallet_address, '') AS bc_wallet_address from sites where site_url = '%s'" % (siteurl_param))
    rsSite = cursor2.fetchone()
    bc_wallet_address=rsSite[0]

    connection2.commit()

if currentValue == printLimit:
    print("Over Limit")
    print("You have reached a limit for generating PDF.")
    exit()

if recordGenerated > printLimit:
    print("Over Limit")
    if noOfCertificateCanGenerate > 1:
        print("You can generate "+str(noOfCertificateCanGenerate)+" PDFs.")
    else:
        print("You can generate "+str(noOfCertificateCanGenerate)+" PDF.")
    exit()

if sys.argv[7]!='mpkv':
    get_blank = []
    get_extractor_name = []
    pno=0
    for pagechk in doc:
        pno += 1
        get_extractor_name.append('<br><b style="color:red">Page '+str(pno)+':</b>')
        for box_chk in extractor_details:
            source_coords_chk = box_chk['coords']
            scoords = source_coords_chk.split(",")
            srect = fitz.Rect(scoords[0],scoords[1],scoords[2],scoords[3])
            blchk=pagechk.getTextbox(srect).replace("\n", "").strip()
            if blchk == '':
                extractor_name=box_chk['name']
                get_blank.append(extractor_name)
                get_extractor_name.append(extractor_name+', ')
            
    if get_blank:
        print("Empty Extractor")
        get_blank_string = ' '.join(get_extractor_name)
        print('<b style="color:blue">Empty Extractor</b>'+get_blank_string) #Blank Source
        exit()


#Check duplicate ids
if sys.argv[4]=='Fresh':
    get_pdffiles = []
    for pname in doc:
        for boks in boxes:
            if boks['placer_type'] == 'QR Dynamic':
                get_text=[]
                if boks['source'] != '' and boks['source'] != 'Current DateTime':
                    scoords = boks['source_coords'].split(",")
                    srect = fitz.Rect(scoords[0],scoords[1],scoords[2],scoords[3])                
                    get_text.append(pname.getTextbox(srect).replace("\n", ""))
                    extracted_id = get_text[0].replace("/", "").replace("\\", "").replace("-", "").strip()
                    if connection.is_connected():
                        #cursor.execute("select * from individual_records where unique_no = '%s' and publish=1 order by id desc" % (extracted_id))
                        # Update Rohit 01/09/2023
                        if sys.argv[18]=='1':
                            cursor.execute("select * from sb_individual_records where unique_no = '%s' and publish=1 order by id desc" % (extracted_id))
                        else:
                            cursor.execute("select * from individual_records where unique_no = '%s' and publish=1 order by id desc" % (extracted_id))
                        # Update Rohit 01/09/2023
                        records = cursor.fetchall()  
                        row_count = cursor.rowcount
                        if row_count > 0:    
                            get_pdffiles.append(records[0][8])  

    if not get_pdffiles:
        print("No Duplicates")    #List is empty
    else:
        print("Duplicates")
        print(len(get_pdffiles))
        get_pdffiles_string = ','.join(get_pdffiles)
        print(get_pdffiles_string) #Duplicate unique ids
        """
        for unid in get_pdffiles:
            if connection.is_connected():
                sql = "UPDATE individual_records SET publish = 2 WHERE unique_no = '%s'" % (unid)
                cursor.execute(sql)
                connection.commit()
        """
        exit()

  

final_data = {'data':[]}
cnt = 1
pp = pprint.PrettyPrinter(indent=4)

#output_file = directory+sys.argv[7]+'/'+"processed_pdfs/"+sys.argv[2]
#Update Rohit 01/09/2023
# if sys.argv[18] == '1':
#     output_folder = rootDir+sys.argv[7]+'/'+"processed_pdfs/sandbox"
# else:
output_folder = rootDir+sys.argv[7]+'/'+"processed_pdfs"

if not os.path.exists(output_folder):
    os.makedirs(output_folder)


# if sys.argv[18] == '1':
#     output_file = directory+sys.argv[7]+'/'+"processed_pdfs/sandbox/"+sys.argv[2]
# else:
output_file = directory+sys.argv[7]+'/'+"processed_pdfs/"+sys.argv[2]
#Update Rohit 01/09/2023

white = (0, 0, 0, 0)
black = (0, 0, 0, 1)
red = (0, 1, 1, 0)
blue = (1, 1, 0, 0)
pink = (0, 1, 0, 0)
aqua = (1, 0, 0, 0)
green = (1, 0, 1, 0)
purple = (1, 1, 0, 0)
yellow = (0, 0, 1, 0)

#folder=rootDir+sys.argv[7]+'/'+"documents/" + template_name
#Update Rohit 01/09/2023
if sys.argv[18] == '1':
    folder=rootDir+sys.argv[7]+'/'+"documents/sandbox/" + template_name
else:
    folder=rootDir+sys.argv[7]+'/'+"documents/" + template_name
#Update Rohit 01/09/2023
if not os.path.exists(folder):
    os.makedirs(folder)

inner_folder=folder +"/"+str(record_unique_id)
if not os.path.exists(inner_folder):
    os.makedirs(inner_folder)
    
pdf_folder=inner_folder +"/pdfs"
if not os.path.exists(pdf_folder):
    os.makedirs(pdf_folder)
    
path_pdf_moved= inner_folder+"/" +sys.argv[2]  #
#print(path_pdf_moved)
#Create excel file to save unique id and QR details
workbook_name=inner_folder+"/" + template_name+".xlsx"
wbc = openpyxl.Workbook()            
Sheet_name = wbc.sheetnames
wbc.save(filename=workbook_name)
wbs = load_workbook(workbook_name)
page_sheet = wbs.active
page_sheet.column_dimensions['A'].width = 30
page_sheet.column_dimensions['B'].width = 50
page_sheet.column_dimensions['C'].width = 120
wrap_alignment = Alignment(wrap_text=True)
page_sheet.append(['Original ID','Unique ID','QR Details'])  
page_sheet.cell(row = 1, column = 1).font = Font(bold = True) 
page_sheet.cell(row = 1, column = 2).font = Font(bold = True) 
page_sheet.cell(row = 1, column = 3).font = Font(bold = True)
datetime_IND = datetime.datetime.now(tz_IND) 
beginning_time = datetime_IND.strftime("%H:%M:%S")
start_time = time.time()
#convert aws string to array
awsS3InstancesArr = sys.argv[17].split('#')
test_pdffiles=[]
for i in doc:
    page = i
    if not(page._isWrapped):
        page._wrapContents()
    page_data = {cnt:[]}
    words = page.getTextWords()
    page_dict = page.getText('dict')    
    page_rect=page.MediaBox
    #print(page_rect)
    #exit()
    #dirName = directory+sys.argv[7]+'/'+"documents/" + str(sys.argv[1])
    # Update Rohit 01/09/2023
    if sys.argv[18] == '1':
        dirName = directory+sys.argv[7]+'/'+"documents/sandbox/" + str(sys.argv[1])
    else:
        dirName = directory+sys.argv[7]+'/'+"documents/" + str(sys.argv[1])
    # Update Rohit 01/09/2023
        
   # print(dirName)
    mintData={}
    mintData["documentType"] = bc_document_type
    mintData["description"] = bc_document_description
    mcount=1
    use_count=0
    for box_blockchain in extractor_details:
        if "blockchain_flag" in box_blockchain:
            if box_blockchain['blockchain_flag'] == 'use':
                box_blockchain_show_flag = 1
                use_count +=1
            else:
                box_blockchain_show_flag = 0
        else:
            box_blockchain_show_flag = 0  

        if box_blockchain_show_flag==1:
            store_metalabel=box_blockchain['metadata_label']
            store_metavalue=box_blockchain['metadata_value']
            store_coords=box_blockchain['coords']
            meta_value=''
            metadata_label = box_blockchain['metadata_label']
            metadata_value = box_blockchain['metadata_value'].replace("{", "").replace("}", "")
            meta_list=list(filter(bool, metadata_value.splitlines()))
            for mdv in meta_list:
                str_val = mdv.split("^")
                meta_value = meta_value + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)  
            mintData['metadata'+str(mcount)]=json.dumps(dict(label = metadata_label, value = meta_value))
            mcount +=1

    if use_count > 0 and use_count < 5:
        for uc in range(use_count+1, 6):
            mintData['metadata'+str(uc)]=json.dumps(dict({}))

    for box in boxes:
        temp = ''
        otxt = ''
        file_path = ''
        #print()
        if box['source'] == '' or box['source'] == 'Current DateTime':
            srect = fitz.Rect(0,0,0,0)            
        else:
            source_coords = box['source_coords']
            scoords = source_coords.split(",")
            srect = fitz.Rect(scoords[0],scoords[1],scoords[2],scoords[3])        
        
        placer_coords = box['placer_coords']
        pcoords = placer_coords.split(",")
        prect = fitz.Rect(pcoords[0],pcoords[1],pcoords[2],pcoords[3])
        prect_coords = pcoords[0],pcoords[1],pcoords[2],pcoords[3]
        #print(placer_coords)  
        if box['placer_font_name'] == '':  
            placer_font_name = fitz.Font(fontfile=dirFont+"arial.ttf")   
        else:
            placer_font_name = fitz.Font(fontfile=dirFont+box['placer_font_name'])
            
        placer_font_underline = box['placer_font_underline'] 
        placer_font_size = box['placer_font_size']                   
        placer_type = box['placer_type']      
        placer_display = box['placer_display']
        if placer_display == '':
            placer_align = int(0)
        else:
            placer_align = int(placer_display)         
            
        if placer_type == 'Invisible':
            placer_color = yellow
        else:
            placer_color = black
        
        placer_degree_angle = box['degree_angle']
        if placer_degree_angle == '':
            placer_degree_angle = int(0)
        else:
            placer_degree_angle = int(placer_degree_angle)        

        placer_opacity = box['opacity_val']
        if box['line_height'] == '':
            placer_lineHeight = 1
        else:
            placer_lineHeight = box['line_height'] 
        
        if "qr_place" in box:
            if box['qr_place'] == 'show':
                qr_show_flag = 1
            else:
                qr_show_flag = 0
        else:
            qr_show_flag = 1         
        
        if "blockchain_flag" in box:
            if box['blockchain_flag'] == 'use':
                blockchain_show_flag = 1
            else:
                blockchain_show_flag = 0
        else:
            blockchain_show_flag = 0         

        if "barcode_content" in box:
            if box['barcode_content'] == 'Source Content':
                barcode_content_flag = 1
            else:
                barcode_content_flag = 0
        else:
            barcode_content_flag = 0           

        if "barcode_content_position" in box:
            if box['barcode_content_position'] == 'Text at Bottom':
                barcode_content_position_flag = 1
            else:
                barcode_content_position_flag = 0
        else:
            barcode_content_position_flag = 0
        
        if box['font_color'] == '':
            placer_font_color = black
        else:
            placer_font_color = box['font_color']
            placer_font_color=fitz.utils.getColor(placer_font_color) 
        
        if placer_type == 'QR Default': 
            if not os.path.exists(dirName):
                os.makedirs(dirName) 
            now = datetime.datetime.now()            
            dt_string = now.strftime("%Y%m%d%H%M%S")+str(cnt)            
            if bc_contract_address == '':  
                result = hashlib.md5(dt_string.encode("latin-1")) 
                barcode_en=result.hexdigest()    
                qr_txt=barcode_en
            else:
                result = hashlib.md5(dt_string.encode("latin-1")) 
                barcode_en=result.hexdigest()  
                arr = bytes(barcode_en, 'latin-1')    
                encryptedData=base64.b64encode(arr)
                qr_txt="https://"+sys.argv[7]+".seqrdoc.com/bverify/" + encryptedData.decode() + "\n\n" + barcode_en

            if qr_show_flag==1:
                qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=4, border=0,)
                qr.add_data(qr_txt)
                qr.make(fit=True)
                img = qr.make_image()  # fill_color="black", back_color="white"
                img.save(dirName+"/"+"qr_"+str(barcode_en)+".png")
                qrcode_file=dirName+"/"+"qr_"+str(barcode_en)+".png"            
                page.insertImage(prect, qrcode_file, overlay=True)   
        elif placer_type == 'QR Dynamic': 
            if not os.path.exists(dirName):
                os.makedirs(dirName) 
            qr_txt=''
            now = datetime.datetime.now()          
            get_text=[]
            if box['source'] == '' or box['source'] == 'Current DateTime':
                #dt_string = now.strftime("%Y-%m-%d %H:%M:%S")+str(cnt)
                dt_string = now.strftime("%Y%m%d%H%M%S")+str(cnt)
                result = hashlib.md5(dt_string.encode('latin-1')) 
                barcode_en=result.hexdigest()
            else:
                #dt_string = otxt.replace("/", "")+now.strftime("%Y-%m-%d %H:%M:%S")+str(cnt)                
                #dt_string = otxt.replace("/", "")
                #dt_string = otxt
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                dt_string = get_text[0].replace("/", "").replace("\\", "").replace("-", "").replace(" ", "").strip()
                result = hashlib.md5(dt_string.encode('latin-1')) 
                barcode_en=result.hexdigest()         

            if bc_contract_address == '':
                barcode_enc=barcode_en
            else:
                arr = bytes(barcode_en, 'latin-1')
                encryptedData=base64.b64encode(arr)
                barcode_enc="https://"+sys.argv[7]+".seqrdoc.com/bverify/" + encryptedData.decode() + "\n\n" + barcode_en
                
            qr_details = box['qr_details'].replace("{", "").replace("}", "")
            qr_list=list(filter(bool, qr_details.splitlines()))
            for n in qr_list:
                str_val = n.split("^")
                str_val_count=len(str_val)
                qr_txt = qr_txt + extract_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)
            qr_txt =  qr_txt + "\n"+barcode_enc
            #print(qr_txt)
            if qr_show_flag==1:
                qr = qrcode.QRCode(version=1,error_correction=qrcode.constants.ERROR_CORRECT_L,box_size=4,border=0,)
                qr.add_data(qr_txt)
                qr.make(fit=True)
                img = qr.make_image()  # fill_color="black", back_color="white"
                img.save(dirName+"/"+"qr_"+str(barcode_en)+".png")
                qrcode_file=dirName+"/"+"qr_"+str(barcode_en)+".png"            
                page.insertImage(prect, qrcode_file, overlay=True)
        elif placer_type == 'Barcode': 
            #get_text=[]
            #get_text.append(page.getTextbox(srect).replace("\n", ""))
            #dt_strings = get_text[0].strip()
            if connection.is_connected():
                today = date.today()
                current_year=get_financial_year(str(today))
                current_year='PN/'+current_year+'/'
                #cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
                # Update Rohit 01/09/2023
                if sys.argv[18] == '1':
                    cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM sb_printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
                else:
                    cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
                # Update Rohit 01/09/2023    
                record = cursor.fetchone() 
                next_print=record[0]+1
                next_print_serial=current_year+str(next_print)

            if barcode_content_flag==1:
                get_text=[]
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                temp = get_text[0].replace(" ", "").strip()
            else:
                temp = next_print_serial

            if barcode_content_position_flag==1:
                write_text_flag = True
            else:
                write_text_flag = False

            EAN = barcode.get_barcode_class('code128')
            ean = EAN(temp, writer=ImageWriter())
            if isinstance(prect.width, float)==True:
                bcwidth = float(prect.width)
            else:
                bcwidth = int(prect.width)
            
            if isinstance(prect.height, float)==True:
                bcheight = float(prect.height)
            else:
                bcheight = int(prect.height)
            
            rect = fitz.Rect(0, 0.85*bcheight, bcwidth, bcheight)
            #print(w,h) 300 dpi=667 mil
            options = {
                'dpi': 300,
                'write_text': write_text_flag,
                'module_width': bcwidth/667, #0.40
                'module_height': rect.height, #10
                'quiet_zone': 0,
                'text_distance': 1,
                'text_line_distance': 1,
                'font_size': placer_font_size,
				'center_text':True
            }            
            barcode_file = ean.save(dirName, options = options)               
            #page.insertTextbox(prect, str(rect.width))           
            #page.drawRect( prect, color = green, fill = green)           
            im = Image.open(barcode_file)            
            imwidth, imheight = im.size            
            rectwidth = bcwidth*3.7795275591            
            if imwidth>rectwidth:
                page.insertImage(prect, barcode_file, overlay=True)
            else:
                width_in_mm = imwidth*0.2645833333
                newwidth = rectwidth-imwidth
                right_pos = newwidth*0.2645833333
                left_pos=int(float(pcoords[0]))
                top_pos=int(float(pcoords[1]))
                right_pos=int(float(pcoords[2])) - right_pos
                bottom_pos=int(float(pcoords[3]))               
                prects = fitz.Rect(left_pos,top_pos,right_pos,bottom_pos)          
                page.insertImage(prects, barcode_file, overlay=True)             
        elif placer_type == 'Micro Line': 
            mi_txt=''            
            if box['qr_details']=='':
                get_text=[]
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                #text_len=len(temp)                
                temp = get_text[0].replace(" ", "").strip()
                search_text = get_text[0].strip()
                #print(type(search_text))
                if placer_font_underline=="underline":
                    rl = doc.loadPage(cnt-1).searchFor(search_text, hit_max=1)
                    #print(rl[0])
                    #print(rl)
                    
                    if page_count==1:
                        output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                    else:
                        if len(rl) <=2:
                            output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                        else:
                            output = str(rl[len(rl)-1]).replace("Rect", "").replace("(", "").replace(")", "")                    
                    
                    #output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")                   
                    ucoords = output.split(",")
                    urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                    #print (urect) 
                    #print (urect.height) 
                    #urect.bl, urect.br
                    #prect_coords = pcoords[0],pcoords[1],pcoords[2],pcoords[3]
                    urect_coords = int(float(ucoords[1]))+int(float(urect.height))-1                    
                    new_urect=fitz.Rect(ucoords[0],urect_coords,ucoords[2],ucoords[3])
                    text_len=len(temp)
                    if isinstance(placer_font_size, float)==True:
                        textwidth = fitz.getTextlength(search_text, fontsize=float(placer_font_size))    
                    else:
                        textwidth = fitz.getTextlength(search_text, fontsize=int(placer_font_size))   
                                
                    #print(textwidth)
                    chrPerLine=int(float(new_urect.width))/textwidth
                    Totalchrs=str(chrPerLine).split(".")[0]
                    #print ("chr = ", chrPerLine)
                    #print (Totalchrs)
                    repeat_txt=temp * int(Totalchrs)
                    #print (repeat_txt)
                    if isinstance(placer_font_size, float)==True:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                    else:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                    
                    #print(int(float(new_urect.width)),"|",int(textwidth),"|",int(float(repeat_textwidth)))
                    remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))        
                    #print("rs: ",remain_space,"|",text_len)
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = temp[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''
                     
                    #print(remain_chrs)
                    wr = fitz.TextWriter(page.rect)
                    wr.fillTextbox(new_urect, (temp * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True) 
                    #page.drawRect(new_urect)
                else:
                    text_len=len(temp)
                    if isinstance(placer_font_size, float)==True:
                        textwidth = fitz.getTextlength(temp, fontsize=float(placer_font_size))    
                    else:
                        textwidth = fitz.getTextlength(temp, fontsize=int(placer_font_size))   
                                
                    #print(textwidth)
                    chrPerLine=int(float(prect.width))/textwidth
                    Totalchrs=str(chrPerLine).split(".")[0]
                    #print ("chr = ", chrPerLine)
                    #print (Totalchrs)
                    repeat_txt=temp * int(Totalchrs)
                    if isinstance(placer_font_size, float)==True:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                    else:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                    
                    #print(int(float(prect.width)),"|",int(textwidth),"|",int(float(repeat_textwidth)))
                    remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3        
                    #print("rs: ",remain_space,"|",text_len)
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = temp[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''
                        
                    #print(remain_chrs) # 0 = left, 1 = center, 2 = right
                    #page.insertTextbox(prect, (temp * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)
                    wr = fitz.TextWriter(page.rect)
                    wr.fillTextbox(prect, (temp * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)    
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    #mi_txt = mi_txt + extract_microline_info(extractor_details,inv,words)  
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    mi_txt = mi_txt + extract_microline_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                    
                #print(mi_txt)      
                
                #print(placer_font_name)
                if "Kruti" in box['placer_font_name']:
                    temp=mi_txt.replace(" ", "")
                    get_text=[]
                    get_text.append(page.getTextbox(srect).replace("\n", ""))
                    search_text = get_text[0].strip()
                    text_len=len(temp)                                   
                    text_to_cal=temp                                   
                else:
                    get_text=[]
                    get_text.append(page.getTextbox(srect).replace("\n", ""))
                    temp = get_text[0].replace(" ", "").strip()              
                    search_text = get_text[0].strip()
                    text_len=len(temp)
                    text_to_cal=search_text

                mi_txt = mi_txt.replace(" ", "")
                
                if placer_font_underline=="underline":
                    #rl = page.searchFor(search_text)  
                    rl = doc.loadPage(cnt-1).searchFor(search_text, hit_max=1) 
                    
                    if page_count==1:
                        output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                    else:
                        if len(rl) > 2:
                            output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")
                        else:
                            output = str(rl[len(rl)-1]).replace("Rect", "").replace("(", "").replace(")", "")
                    
                    #output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")                  
                    ucoords = output.split(",")
                    urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                    #print (urect, prect) 
                    #urect.bl, urect.br
                    #prect_coords = pcoords[0],pcoords[1],pcoords[2],pcoords[3]
                    urect_coords = int(float(ucoords[1]))+int(float(urect.height))-2                    
                    new_urect=fitz.Rect(ucoords[0],urect_coords,ucoords[2],ucoords[3])
                    
                    if isinstance(placer_font_size, float)==True:
                        textwidth = fitz.getTextlength(text_to_cal, fontsize=float(placer_font_size))    
                    else:
                        textwidth = fitz.getTextlength(text_to_cal, fontsize=int(placer_font_size))
                                
                    #print(textwidth,mi_txt)
                    chrPerLine=(int(float(new_urect.width))/textwidth)
                    Totalchrs=str(chrPerLine).split(".")[0]
                    #print ("chr = ", chrPerLine)
                    #print (Totalchrs)                    
                    repeat_txt=mi_txt * int(Totalchrs)
                    if isinstance(placer_font_size, float)==True:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))-10   
                    else:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                    
                    #print(int(float(new_urect.width)),"|",int(textwidth),"|",int(float(repeat_textwidth)))
                    remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))-3
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = mi_txt[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''                    
                    #exit()  
                    
                    wl_lst = fitz.getTextlength((mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size)+1
                    wl = sum([wl_lst])
                    if wl>new_urect.width:      
                        remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))-10     
                    else:       
                        remain_space=int(float(new_urect.width))-int(float(repeat_textwidth))-3
                    
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = mi_txt[0:int(remain_chrs_count)]
                        #print("space|", text_len, remain_space, remain_chrs_count)
                    else:
                        remain_chrs = ''
                    
                    #print((mi_txt * int(Totalchrs))+remain_chrs)
                    wr = fitz.TextWriter(page.rect)
                    if "Kruti" in box['placer_font_name']:
                        wr.fillTextbox(new_urect, (mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align)
                    else:
                        wr.fillTextbox(new_urect, (mi_txt * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True) 
                    #page.drawRect(new_urect)
                else:
                    text_len=len(mi_txt)
                    if isinstance(placer_font_size, float)==True:
                        textwidth = fitz.getTextlength(mi_txt, fontsize=float(placer_font_size))    
                    else:
                        textwidth = fitz.getTextlength(mi_txt, fontsize=int(placer_font_size))   
                                
                    chrPerLine=(int(float(prect.width))/textwidth)
                    Totalchrs=str(chrPerLine).split(".")[0]
                    repeat_txt=mi_txt * int(Totalchrs)
                    if isinstance(placer_font_size, float)==True:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                    else:
                        repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                    
                          
                    remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3 
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = mi_txt[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''                    
                    #exit()                    
                    wl_lst = fitz.getTextlength((mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size)+1
                    wl = sum([wl_lst])
                    if wl>prect.width:      
                        remain_space=int(float(prect.width))-int(float(repeat_textwidth))-10     
                    else:       
                        remain_space=int(float(prect.width))-int(float(repeat_textwidth))-3
                    
                    if remain_space>0:                
                        RchrPerLine=int(remain_space)/textwidth
                        remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                        remain_chrs = mi_txt[0:int(remain_chrs_count)]
                    else:
                        remain_chrs = ''
                        
                    #page.insertTextbox(prect, (mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size, align=placer_align, color=placer_font_color, overlay=True)
                    wr = fitz.TextWriter(page.rect)
                    wr.fillTextbox(prect, (mi_txt * int(Totalchrs))+remain_chrs, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                    wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)
        elif placer_type == 'Invisible':                        
            invisible_font_color=fitz.utils.getColor("YELLOW")
            inv_txt=''
            get_text=[]
            if box['qr_details']=='':  
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                otxt=get_text[0].strip()
                #print(otxt)
                #page.insertTextbox(prect, otxt, fontsize=placer_font_size, color=placer_color, align=placer_align, overlay=True)
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prect, otxt,  font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                #wr.writeText(page, color=placer_color, opacity=float(1.0), overlay=True)
                wr.writeText(page, color=invisible_font_color, opacity=float(1.0), overlay=True) 
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    inv_txt = inv_txt + extract_info(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                          
                     
                #page.insertTextbox(prect, inv_txt, fontsize=placer_font_size, color=placer_color, align=placer_align, overlay=True)
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prect, inv_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                #wr.writeText(page, color=placer_color, opacity=float(1.0), overlay=True)
                wr.writeText(page, color=invisible_font_color, opacity=float(1.0), overlay=True) 
        elif placer_type == 'Invisible Image': 
            temp_img = [block for block in page_dict['blocks'] if (fitz.Rect(block['bbox']) in srect and block['type'] == 1)]  
            if len(temp_img) > 0:
                pix = fitz.Pixmap(temp_img[0]['image'])
                if not os.path.exists(dirName):
                    os.makedirs(dirName)
                file_path = dirName + "/"+str(int(round(time.time() * 1000))) + "-" + str(cnt) + "." + temp_img[0]['ext']
                file_path2 = dirName + "/"+str(int(round(time.time() * 1000))) + "-" + str(cnt) + ".png"                        
                pix.writeImage(file_path)
                imgs = Image.open(file_path).convert("L") 
                img1 = ImageOps.colorize(imgs, black ="white", white ="yellow")             
                img1.save(file_path2, 'png')
                page.insertImage(prect, file_path2, overlay=True)                       
        elif placer_type == 'Ghost Image': 
            #print(prect.width,"| ",prect.height)
            if placer_degree_angle==90:
                half_width=int(float(prect.height))/2
                top_pos_minus=(int(float(pcoords[0])) - int(float(half_width)))  
                add_right_pos=int(float(pcoords[0])) - top_pos_minus  
                left_pos=int(float(pcoords[0]))-add_right_pos
                top_pos=int(float(pcoords[1]))
                right_pos=int(float(pcoords[2])) + add_right_pos
                bottom_pos=int(float(pcoords[3]))               
                prects = fitz.Rect(left_pos,top_pos,right_pos,bottom_pos) 
                #print(prects)
                #print(prects.width,"| ",prects.height)                         
            else:
                prects=prect         

            get_text=[]
            get_text.append(page.getTextbox(srect).replace("\n", ""))
            #temp = ''.join(c for c in get_text[0].strip().replace("/", "").replace(".", "") if c.isalnum())
            extracted_string = re.sub(r'\W+', '', get_text[0])
            temp = ''.join(c for c in extracted_string if c.isalnum())
            ghost_width = round(box['width'] * 3.7795275591)  # Millimeter to Pixel, 1 mm = 3.7795275591 pixel
            ghost_height = round(box['height'] * 3.7795275591)  
            ghost_words = box['ghost_words']  
            PrintableChars=temp[ 0 : ghost_words ] #extract first chars
            #print(ghost_words)
            if not os.path.exists(dirName):
                os.makedirs(dirName)            
            ghostImg=CreateGhostImage(dirName, PrintableChars, placer_font_size, ghost_width, ghost_height)
            page.insertImage(prects, ghostImg,overlay=True, rotate=placer_degree_angle)        
        elif placer_type == 'Image':
            #image_path=rootDir+sys.argv[7]+"/upload_images/" + box['image_path']
            image_path=rootDir+sys.argv[7]+"/backend/templates/pdf2pdf_images/" + box['image_path']
            #print(image_path)
            page.insertImage(prect, image_path,overlay=True)
        elif placer_type == 'Plain Text':           
            inv_txt=''
            get_text=[]
            if box['qr_details']=='':                
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                otxt_string = get_text[0].strip()
                #otxt_string = otxt.replace("/", "")
                #page.insertTextbox(prect, otxt_string, fontsize=placer_font_size, color=placer_font_color, align=placer_align, overlay=True)
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prect, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)                 
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    inv_txt = inv_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                          
                     
                #page.insertTextbox(prect, inv_txt, fontsize=placer_font_size, color=placer_font_color, align=placer_align, overlay=True) 
                wr = fitz.TextWriter(page.rect)
                wr.fillTextbox(prect, inv_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(1.0), overlay=True)               
        elif placer_type == 'Static Text':     
            otxt_string = box['qr_details']
            m = fitz.Matrix(0) #placer_degree_angle
            points = fitz.Point(prect.x0, prect.y0)
            wr = fitz.TextWriter(page.rect)
            wr.fillTextbox(prect, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
            wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))
            if placer_font_underline=="underline":
                rl = page.searchFor(otxt_string)  
                output = str(rl[0]).replace("Rect", "").replace("(", "").replace(")", "")                    
                ucoords = output.split(",")
                urect=fitz.Rect(ucoords[0],ucoords[1],ucoords[2],ucoords[3])
                shape = page.newShape()
                shape.drawLine(urect.bl, urect.br)
                shape.finish(color=placer_font_color, stroke_opacity=float(placer_opacity))
                shape.commit()            
        elif placer_type == 'Watermark Text':           
            inv_txt=''
            get_text=[]
            if box['qr_details']=='':                
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                otxt_string = get_text[0].strip()
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(prect_coords)
                new_rect=ir * m         
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(prect.x0, prect.y0)                 
                wr.fillTextbox(prect, otxt_string, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))                
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    inv_txt = inv_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                          
                    
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(prect_coords)
                new_rect=ir * m         
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(prect.x0, prect.y0)                 
                wr.fillTextbox(prect, inv_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=True, morph=(points,m))        
        elif placer_type == 'Watermark Multi Lines':  
            mi_txt=''            
            if box['qr_details']=='':
                get_text=[]
                get_text.append(page.getTextbox(srect).replace("\n", ""))
                text_len=len(temp)
                temp = get_text[0].strip()+' '
                #chrPerLine=int(float(page_rect.width)+int(float(page_rect.height)))
                chrPerLine=int(390)
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=temp * int(Totalchrs)       
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(page_rect)
                new_page_rect=ir * m 
                #new_page_rect.y0=new_page_rect.y0-int(10)
                #new_page_rect.x0=new_page_rect.x0+int(200)
                new_page_rect.x1=new_page_rect.x1+int(100) 
                new_page_rect.y1=new_page_rect.y1+int(100)     
                wr = fitz.TextWriter(page.rect) #wr = fitz.TextWriter(page.rect)
                points = fitz.Point(0, 0)
                wr.fillTextbox(new_page_rect, repeat_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align, lineheight=placer_lineHeight)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=False, morph=(points,m))

                     
                   # wC=textsize(self, 'ABCD', font=None, *args, **kwargs)
                    #print(wC)
            else:
                qr_details = box['qr_details'].replace("{", "").replace("}", "")
                qr_list=list(filter(bool, qr_details.splitlines()))
                for inv in qr_list:
                    str_val = inv.split("^")
                    str_val_count=len(str_val)
                    mi_txt = mi_txt + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)                    
                #print(mi_txt)      
                text_len=len(mi_txt)

                
                if isinstance(placer_font_size, float)==True:
                    textwidth = fitz.getTextlength(mi_txt, fontsize=float(placer_font_size))    
                else:
                    textwidth = fitz.getTextlength(mi_txt, fontsize=int(placer_font_size))   
                
                chrPerLine=(int(float(page_rect.width))/textwidth)
                Totalchrs=str(chrPerLine).split(".")[0]
                repeat_txt=mi_txt * int(Totalchrs)
                if isinstance(placer_font_size, float)==True:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=float(placer_font_size))    
                else:
                    repeat_textwidth = fitz.getTextlength(repeat_txt, fontsize=int(placer_font_size))   
                
                      
                remain_space=int(float(page_rect.width))-int(float(repeat_textwidth))-3 
                if remain_space>0:                
                    RchrPerLine=int(remain_space)/textwidth
                    remain_chrs_count=int(text_len) * int(remain_space)/int(textwidth)
                    remain_chrs = mi_txt[0:int(remain_chrs_count)]
                else:
                    remain_chrs = ''                    
                #exit()                    
                wl_lst = fitz.getTextlength((mi_txt * int(Totalchrs))+remain_chrs, fontsize=placer_font_size)+1
                wl = sum([wl_lst])

                if placer_lineHeight<3:
                    yOffset=600
                else:
                    yOffset=100

                pageW=float(page_rect.width)+int(60)
                pageH=float(page_rect.height)+int(yOffset)
                areaPage=int(pageW)*int(pageH)
                diagonal = round(math.sqrt((pageW**2) + (pageH**2)), 4)
                font = ImageFont.truetype('arial.ttf', placer_font_size)
                size = font.getsize(mi_txt)
                diagonal=float(diagonal) 
                sizeWInPoints=float(size[0])*float(1.3333)
                charPerSingleLine=int(math.ceil(diagonal/sizeWInPoints))
                sizeHInPoints=float(size[1])*float(1.3333)
                areaPerLine=int(diagonal)* int(math.ceil(sizeHInPoints))
                totalRepeatText=int(math.ceil(areaPage/areaPerLine))
                
                # areaPage=int(float(page_rect.width)*int(float(page_rect.height)))
                # #font = ImageFont.truetype('arial.ttf', placer_font_size)
                # font = ImageFont.truetype('arial.ttf', 10)
                # size = font.getsize(mi_txt)
                # mmWidth= int(math.ceil(float(size[0])*float(0.26458333))) 
                # #charPerSingleLine=int(math.ceil(364/mmWidth))
                # charPerSingleLine=int(math.ceil(wl/mmWidth))
                # areaPerLine=int(mmWidth)*int(charPerSingleLine)*int(size[1])
                # totalRepeatText=int(math.ceil(areaPage/areaPerLine))

                #chrPerLine=int(364)
                #Totalchrs=str(chrPerLine).split(".")[0]
                #repeat_txt=mi_txt * int(Totalchrs)
                no_of_repeat=int(totalRepeatText) * int(charPerSingleLine)
                #no_of_repeat= int(no_of_repeat) + int(10) 
                repeat_txt=mi_txt * int(no_of_repeat)
                repeat_txt=repeat_txt
                m = fitz.Matrix(placer_degree_angle)
                ir = fitz.IRect(page_rect)
                new_page_rect=ir * m
                new_page_rect.x1=new_page_rect.x1 #+int(100) 
                new_page_rect.y1=new_page_rect.y1 #+int(100) 
                wr = fitz.TextWriter(page.rect)
                points = fitz.Point(0, 0)  #prect.x0, prect.y0
                wr.fillTextbox(new_page_rect, repeat_txt, font=placer_font_name, fontsize=placer_font_size, align=placer_align, lineheight=placer_lineHeight)
                wr.writeText(page, color=placer_font_color, opacity=float(placer_opacity), overlay=False, morph=(points,m))
                          
                #if sys.argv[7]=='demo':
                    #print(mi_txt)
                    #print(totalRepeatText)
                    #print(Totalchrs)
                #     font = ImageFont.truetype('arial.ttf', 12)
                #     size = font.getsize(mi_txt)
                #     print(mi_txt)
        """
        if blockchain_show_flag==1 and bc_contract_address != '':
            meta_value=''
            metadata_label = box['metadata_label']
            metadata_value = box['metadata_value'].replace("{", "").replace("}", "")
            meta_list=list(filter(bool, metadata_value.splitlines()))
            for mdv in meta_list:
                str_val = mdv.split("^")
                meta_value = meta_value + extract_plainText(extractor_details,str_val[1],str(str_val[0]),str(str_val[2]),page)  
            mintData['metadata'+str(mcount)]=json.dumps(dict(label = metadata_label, value = meta_value))
            mcount +=1
        """            
        #page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
        #page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_REMOVE)
        doc_new = fitz.open()
        doc_new.insertPDF(doc, from_page=cnt-1, to_page=cnt-1)
        if verification_setbg=='Yes':
            for dpage in doc_new:
                dpage.insertImage(page_rect, verification_bg_file,overlay=False)                
        #doc_new.save(pdf_folder+"/"+dt_string+".pdf", garbage=4, deflate=True)
        #doc_new.save(rootDir+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf", garbage=4, deflate=True)
        #Upload file to secdoc directory
        """
        if sys.argv[13] :
            URL = "https://"+sys.argv[7]+"."+sys.argv[10]+"/admin/store-file"
            pdf_path=pdf_folder+"/"+dt_string+".pdf"
            PARAMS = {'pdf_path':pdf_path,'site_id':sys.argv[13]}
            response =requests.get(url = URL, params = PARAMS, timeout=None)        
            response.close()     
        #doc_new.save(pdf_folder+"/"+dt_string+".pdf", garbage=4, deflate=True, owner_pw="owner", encryption=encrypt_meth, permissions=perm) 
        """ 
        
        # doc_new.save(rootDir+"/backend/pdf_file/" + dt_string+".pdf", garbage=4, deflate=True)
        # pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
        # shutil.move(rootDir+"/backend/pdf_file/" + dt_string+".pdf", pdf_file_path)
        # Update Rohit 01/09/2023
        if sys.argv[18] == '1':
            if not os.path.exists(rootDir+"/backend/pdf_file/sandbox"):
                os.makedirs(rootDir+"/backend/pdf_file/sandbox") 

            doc_new.save(rootDir+"/backend/pdf_file/sandbox/" + dt_string+".pdf", garbage=4, deflate=True)
            pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/sandbox/" + dt_string+".pdf"
            shutil.move(rootDir+"/backend/pdf_file/sandbox/" + dt_string+".pdf", pdf_file_path)
        else:
            if not os.path.exists(rootDir+"/backend/pdf_file"):
                os.makedirs(rootDir+"/backend/pdf_file") 
            doc_new.save(rootDir+"/backend/pdf_file/" + dt_string+".pdf", garbage=4, deflate=True)
            pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
            shutil.move(rootDir+"/backend/pdf_file/" + dt_string+".pdf", pdf_file_path)
        # Update Rohit 01/09/2023

        """
        doc_new.save(pdf_folder+"/"+dt_string+".pdf", garbage=4, deflate=True)
        pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
        
        #shutil.move(pdf_folder+"/"+dt_string+".pdf", pdf_file_path)
        arg2= pdf_folder+"/"+dt_string+".pdf"
        arg1= '-sOutputFile='+pdf_file_path
        #on local
        #subprocess.Popen(['C:/Program Files/gs/gs9.53.3/bin/gswin64c.exe', '-sDEVICE=pdfwrite', '-dCompatibilityLevel=1.4', '-dPDFSETTINGS=/ebook', '-dNOPAUSE', '-dBATCH',  '-dQUIET', str(arg1), arg2], stdout=subprocess.PIPE)
        #on server
        subprocess.Popen(['C:/Program Files/gs/gs9.56.1/bin/gswin64c.exe', '-sDEVICE=pdfwrite', '-dCompatibilityLevel=1.4', '-dPDFSETTINGS=/ebook', '-dNOPAUSE', '-dBATCH',  '-dQUIET', str(arg1), arg2], stdout=subprocess.PIPE)
        """
        """ 
        set_background(
            input_pdf=verification_bg_file, 
            output=pdf_folder+"/"+dt_string+".pdf",
            watermark=pdf_folder+"/"+dt_string+".pdf")   
        """       


    # Update Blockchain by rohit 17-06-2025
    if(bc_wallet_address): 
            
        checkActiveContractExist =  checkActiveContract(template_id)

        if not checkActiveContractExist:

            # print(checkActiveContractExist)
            createNewContractResponse = createNewContract(bc_wallet_address,template_id)
            if createNewContractResponse:
                bc_contract_address = createNewContractResponse['contract_address'] 
                bc_wallet_address = createNewContractResponse['wallet_address']
                bc_sc_id = createNewContractResponse['bc_sc_id']
                # checkContract
        else:
            bc_contract_address = checkActiveContractExist['contract_address'] 
            bc_wallet_address = checkActiveContractExist['wallet_address']
            bc_sc_id = checkActiveContractExist['bc_sc_id']
            # print(checkActiveContractExist)
        

        
        

            # print(bc_wallet_address)
            # print(bc_contract_address)
            # print("An active contract exists.")

        # if(checkActiveContractExist == 'None'):

        #     createNewContractResponse = createNewContract(bc_wallet_address,template_id)
        checkContractExist =checkContract(bc_wallet_address,template_id,5000)

        # print(checkContractExist)
        # exit()
        if checkContractExist:
            bc_contract_address = checkContractExist['contract_address'] 
            bc_wallet_address = checkContractExist['wallet_address']
            bc_sc_id = checkContractExist['bc_sc_id']

    

    # sys.argv[18] == sandboxing param  -- [0 -> off || 1 -> On]
    # Update Rohit 01/09/2023
    bc_start_time = time.time()
    #if use_count > 0 and bc_contract_address != '':
    if use_count > 0 and bc_contract_address != '' and sys.argv[18] != '1':      
        pdf_file_path=rootDir+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
        pdf_file_path = pdf_file_path.replace("//", "/")
        pdf_file_path = pdf_file_path.replace("/", "\\")
        # print(pdf_file_path)
        # exit()
        """
        temp_dir=tempfile.gettempdir()
        r = requests.post("https://"+sys.argv[7]+".seqrdoc.com/"+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf", allow_redirects=True)
        open(temp_dir+'\/'+dt_string+'.pdf', 'wb').write(r.content)
        files=[
            ('document',(dt_string+'.pdf',open(temp_dir+'\/'+dt_string+'.pdf','rb'),'application/pdf'))
        ]
        """        
        # files=[
        #     ('document',(dt_string+'.pdf',open(pdf_file_path,'rb'),'application/pdf'))
        # ]
        
        files=[
            ('file',(dt_string+'.pdf',open(pdf_file_path,'rb'),'application/pdf'))
        ]
        
        if sys.argv[7]=='mpkv':
            mintData["description"]=mintData["description"] + " "+dt_string
        


        mintData["walletID"] = bc_wallet_address 
        mintData["smartContractAddress"] = bc_contract_address 
        mintData["uniqueHash"] = barcode_en
        mintData["pdf_file"] = "https://"+sys.argv[7]+".seqrdoc.com/"+sys.argv[7]+"/backend/pdf_file/" + dt_string+".pdf"
        mintData["storageIdentifier"] = 'pinata'
        blockchain_url = 'http://node.seqrdoc.com:9090/mainnet/mint'
        
        headers = {
            'accept': 'application/json',
            'x-api-key': '123456789'
        }
        blockchain_response = requests.post(blockchain_url, headers=headers, data=mintData, files=files)
    else:
        
        blockchain_response = requests.post('http://node.seqrdoc.com:9090/mainnet/mint') #"<Response [201]>"
        
    #print(blockchain_response)
    bc_response_time = round(time.time() - bc_start_time, 3)
    if print_setbg=='Yes':
        page.insertImage(page_rect, print_bg_file,overlay=False)      
    if connection.is_connected():
        #sql = "INSERT INTO individual_records (file_records_id, template_id, template_name, pdf_page, page_no, encoded_id, unique_no, qr_details, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        # Update Rohit 01/09/2023
        if sys.argv[18]=='1':
            sql = "INSERT INTO sb_individual_records (file_records_id, template_id, template_name, pdf_page, page_no, encoded_id, unique_no, qr_details, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        else:
            sql = "INSERT INTO individual_records (file_records_id, template_id, template_name, pdf_page, page_no, encoded_id, unique_no, qr_details, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        # Update Rohit 01/09/2023
        val = (file_records_next_id, template_id, template_name, pdf_page, cnt, barcode_en, dt_string, qr_txt, userid, record_unique_id)
        cursor.execute(sql, val)         
        
    current_date_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S');
    if connection.is_connected():

        #cursor.execute("select * from student_table where serial_no = '%s' and status=1 order by id desc" % (dt_string))
        # Update Rohit 01/09/2023
        if sys.argv[18] == '1':
            cursor.execute("select * from sb_student_table where serial_no = '%s' and status=1 order by id desc" % (dt_string))
        else:
            cursor.execute("select * from student_table where serial_no = '%s' and status=1 order by id desc" % (dt_string))
        # Update Rohit 01/09/2023
        records = cursor.fetchall()  
        row_count = cursor.rowcount
        if row_count > 0: 
            #sql = "UPDATE student_table SET status = %s, updated_by = %s, updated_at = %s WHERE serial_no = %s"
            # Update Rohit 01/09/2023
            if sys.argv[18] == '1':
                sql = "UPDATE sb_student_table SET status = %s, updated_by = %s, updated_at = %s WHERE serial_no = %s"
            else:
                sql = "UPDATE student_table SET status = %s, updated_by = %s, updated_at = %s WHERE serial_no = %s"
            # Update Rohit 01/09/2023

            val = ("0", userid, current_date_time, dt_string)
            cursor.execute(sql, val)   

            #cursor.execute ("UPDATE student_table SET status = '%s', updated_by = '%s', updated_at = '%s' WHERE serial_no = '%s' ", ('0', userid, current_date_time, dt_string))

    # site_url='demo.seqrdoc.com'
    # if connection.is_connected():
    #     cursor.execute("SELECT site_id FROM sites WHERE site_url = '%s' " % (site_url))
    #     siteData = cursor.fetchone()
        path="path"
        key="key"
        certificate_filename=dt_string+'.pdf'
        qr_path='qr/'+barcode_en+'.png'

        # print(blockchain_response.status_code)
        salt = 'AJITNATH'
        pdf_url = f"https://{sys.argv[7]}.seqrdoc.com/{sys.argv[7]}/backend/pdf_file/{dt_string}.pdf"
        hash_result = generate_file_hash(pdf_url, salt)
        # print
        if blockchain_response.status_code == 200:
            mint_json = json.loads(blockchain_response.text)
            #sql2 = """INSERT INTO student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`, `bc_txn_hash`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            # Update Rohit 01/09/2023
            if sys.argv[18] == '1':
                sql2 = """INSERT INTO sb_student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`, `bc_txn_hash`,`bc_ipfs_hash`,`pinata_ipfs_hash`,bc_file_hash) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s , %s)"""
            else:
                sql2 = """INSERT INTO student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`, `bc_txn_hash`,`bc_ipfs_hash`,`pinata_ipfs_hash`,bc_file_hash) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s , %s)"""
            # Update Rohit 01/09/2023
            # val2 = (dt_string, certificate_filename, template_id, barcode_en, qr_path, userid, userid, 1, 1, current_date_time,sys.argv[13],1,mint_json['txnHash'],mint_json['ipfsHash'],mint_json['pinataIpfsHash'])
            val2 = (dt_string, certificate_filename, template_id, barcode_en, qr_path, userid, userid, 1, 1, current_date_time,sys.argv[13],1,mint_json.get('txnHash'),mint_json.get('ipfsHash'),mint_json.get('pinataIpfsHash'),hash_result)
        else:
            
            #sql2 = """INSERT INTO student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            # Update Rohit 01/09/2023
            if sys.argv[18] == '1':
                sql2 = """INSERT INTO sb_student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            else:
                sql2 = """INSERT INTO student_table (`serial_no`, `certificate_filename`, `template_id`, `key`, `path`, `created_by`, `updated_by`,`status`, `publish`, `created_at`, `site_id`,`template_type`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            # Update Rohit 01/09/2023    
            val2 = (dt_string, certificate_filename, template_id, barcode_en, qr_path, userid, userid, 1, 1, current_date_time,sys.argv[13],1)
        cursor.execute(sql2, val2) 
        student_table_id = cursor.lastrowid

        if blockchain_response.status_code == 200:
            mint_json = json.loads(blockchain_response.text)
            #print(mint_json['txnHash'], mint_json['tokenID'], mint_json['gasPrice'])
            #sql = "UPDATE student_table SET bc_txn_hash = %s WHERE serial_no = %s AND status = %s"
            # Update Rohit 01/09/2023
            if sys.argv[18] == '1':
                sql = "UPDATE sb_student_table SET bc_txn_hash = %s ,bc_ipfs_hash = %s , pinata_ipfs_hash = %s WHERE serial_no = %s AND status = %s"
            else:
                sql = "UPDATE student_table SET bc_txn_hash = %s ,bc_ipfs_hash = %s , pinata_ipfs_hash = %s WHERE serial_no = %s AND status = %s"
            val = (mint_json.get('txnHash'),mint_json.get('ipfsHash'),mint_json.get('pinataIpfsHash'), dt_string, '1')
            # Update Rohit 01/09/2023
            cursor.execute(sql, val)
            # Update Rohit 01/09/2023
            if sys.argv[18] != '1':
            # Update Rohit 01/09/2023
                sql2 = """INSERT INTO bc_mint_data (`txn_hash`, `gas_fees`, `token_id`, `key`, `created_at`) VALUES (%s, %s, %s, %s, %s)"""
                val2 = (mint_json['txnHash'], mint_json['gasPrice'], mint_json['tokenID'], barcode_en, current_date_time)
                cursor.execute(sql2, val2) 
                
                # bc_api_tracker  # Update Rohit 28/07/2025
                header_parameters = []
                client_ip = requests.get('https://api.ipify.org').text
                current_datetime_bc = now.strftime('%Y-%m-%d %H:%M:%S')

                # current_datetime_bc = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                # $requestUrl,$requestMethod,$requestParameters,$response,$status,$responseTime,"mintDataV1"
                
                # print(('mintDataV1', 'POST', blockchain_url, header_parameters, mintData, blockchain_response, 'success', client_ip, '',current_datetime_bc))
                
                try:
                    sql2313 = "INSERT INTO bc_api_tracker (api_name, request_method, request_url, header_parameters, request_parameters, response, status, client_ip,response_time,created_at) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                   
                    val2313 = (
                        'mintDataV1',
                        'POST',
                        blockchain_url,
                        json.dumps(header_parameters),     # convert to JSON string
                        json.dumps(mintData),              # convert to JSON string
                        blockchain_response.text,          # convert response to string content
                        'success',
                        client_ip,
                        str(bc_response_time),                # response_time (optional or calculate duration)
                        current_datetime_bc
                    )
                    cursor.execute(sql2313, val2313)
                    connection.commit() # Ensure changes are saved
                except Exception as e:
                    print("DB Error:", e)


                # Blockchain other data
                cursor.execute("SELECT id FROM student_table WHERE serial_no = %s AND status = 1", (dt_string,))
                student = cursor.fetchone()
                if student:
                    student_id = student[0]

                    # Step 2: Insert or update in blockchain_other_data
                    sql = """
                    INSERT INTO blockchain_other_data (student_table_id, vendor_identifier)
                    VALUES (%s, %s)
                    ON DUPLICATE KEY UPDATE vendor_identifier = VALUES(vendor_identifier)
                    """
                    cursor.execute(sql, (student_id, 1))
                    connection.commit()

            # Update Blockchain By Rohit 17/06/2025
            if bc_sc_id:
                # print('in thus condition')
                sql2308 = "UPDATE student_table SET bc_sc_id = %s WHERE serial_no = %s AND status = %s"
                val2308 = (bc_sc_id, dt_string, '1')
                cursor.execute(sql2308, val2308)

                sql2308V1 = "UPDATE bc_smart_contracts SET count = count+1 WHERE id = %s"
                val2308V1 = (bc_sc_id,)
                cursor.execute(sql2308V1, val2308V1)
        cursor.execute("select printer_name from system_config where site_id = '%s'" % (sys.argv[13]))
        recordSystem = cursor.fetchone()  
        printer_name=recordSystem[0]

        # cursor.execute("SELECT * FROM printing_details WHERE sr_no = '%s' " % (dt_string))
        # Update Rohit 01/09/2023
        if sys.argv[18] == '1':
            cursor.execute("SELECT * FROM sb_printing_details WHERE sr_no = '%s' " % (dt_string))
        else:
            cursor.execute("SELECT * FROM printing_details WHERE sr_no = '%s' " % (dt_string))
        # Update Rohit 01/09/2023
        records = cursor.fetchall()  
        print_count = cursor.rowcount

        today = date.today()
        current_year=get_financial_year(str(today))
        current_year='PN/'+current_year+'/'
        #cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
        # Update Rohit 01/09/2023
        if sys.argv[18] == '1':
            cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM sb_printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
        else:
            cursor.execute("SELECT COALESCE(MAX(CONVERT(SUBSTR(print_serial_no, 10), UNSIGNED)), 0) AS next_num FROM printing_details WHERE SUBSTR(print_serial_no, 1, 9) = '%s'" % (current_year))
        # Update Rohit 01/09/2023
        record = cursor.fetchone() 
        next_print=record[0]+1
        next_print_serial=current_year+str(next_print)
        #sql3 = """INSERT INTO printing_details (`username`, `print_datetime`, `printer_name`, `print_count`, `print_serial_no`, `sr_no`, `template_name`,`created_at`, `created_by`, `updated_at`, `updated_by`,`status`, `site_id`, `publish`, `student_table_id`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        # Update Rohit 01/09/2023
        if sys.argv[18] == '1':
            sql3 = """INSERT INTO sb_printing_details (`username`, `print_datetime`, `printer_name`, `print_count`, `print_serial_no`, `sr_no`, `template_name`,`created_at`, `created_by`, `updated_at`, `updated_by`,`status`, `site_id`, `publish`, `student_table_id`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        else:
            sql3 = """INSERT INTO printing_details (`username`, `print_datetime`, `printer_name`, `print_count`, `print_serial_no`, `sr_no`, `template_name`,`created_at`, `created_by`, `updated_at`, `updated_by`,`status`, `site_id`, `publish`, `student_table_id`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        # Update Rohit 01/09/2023
        val3 = (sys.argv[14], current_date_time, printer_name, print_count, next_print_serial, dt_string, template_name, current_date_time, userid, current_date_time,userid, 1, sys.argv[13], 1,student_table_id)
        cursor.execute(sql3, val3) 
    if (sys.argv[7] in awsS3InstancesArr):    
        source_file=pdf_file_path
        test_pdffiles.append(source_file)
    page_sheet.append([dt_string,barcode_en,qr_txt])
    page_sheet.cell(row = cnt+1, column = 1).alignment = Alignment(vertical='top')
    page_sheet.cell(row = cnt+1, column = 2).alignment = Alignment(vertical='top')
    page_sheet.cell(row = cnt+1, column = 3).alignment = Alignment(wrapText=True,vertical='top')
    wbs.save(filename=workbook_name) 
    datetime_IND = datetime.datetime.now(tz_IND)
    ending_time = datetime_IND.strftime("%H:%M:%S")
    end_time=time.time() - start_time
    seconds_to_hhmmss=time.strftime('%H:%M:%S', time.gmtime(end_time))
    page_seconds_to_hhmmss=time.strftime('%H:%M:%S', time.gmtime(end_time/cnt))
    arr_content['percent'] = int(cnt/page_count * 100)
    arr_content['message'] = "Generating "+str(cnt)+"/"+str(page_count)+" PDF(s)"
    arr_content['beginning_time'] = beginning_time
    arr_content['ending_time'] = ending_time
    arr_content['exec_time'] = end_time
    arr_content['hms_time'] = seconds_to_hhmmss
    arr_content['page_time'] = page_seconds_to_hhmmss
    arr_content['pages_processed'] = cnt
    json_object = json.dumps(arr_content, indent = 4)

    #f = open(rootDir+sys.argv[7]+'/'+"processed_pdfs/"+sys.argv[5], "w")
    # Update Rohit 01/09/2023
    # if sys.argv[18] == '1':
    #     f = open(rootDir+sys.argv[7]+'/'+"processed_pdfs/sandbox/"+sys.argv[5], "w")
    # else:
    f = open(rootDir+sys.argv[7]+'/'+"processed_pdfs/"+sys.argv[5], "w")
    # Update Rohit 01/09/2023
        
    f.write(json_object)
    #time.sleep(1)
    cnt += 1

if connection.is_connected():
    #sql = "INSERT INTO file_records (template_id, template_name, pdf_page, total_records, source_file, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    # Update Rohit 01/09/2023
    if sys.argv[18]=='1':
        sql = "INSERT INTO sb_file_records (template_id, template_name, pdf_page, total_records, source_file, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    else:
        sql = "INSERT INTO file_records (template_id, template_name, pdf_page, total_records, source_file, userid, record_unique_id) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    # Update Rohit 01/09/2023
    val = (template_id, template_name, pdf_page, cnt-1, sys.argv[2], userid, record_unique_id)
    cursor.execute(sql, val)

    connection.commit()
print(path_pdf_moved)

#Upload pdf to AWS s3
if (sys.argv[7] in awsS3InstancesArr):
            URL = "https://"+sys.argv[7]+"."+sys.argv[10]+"/admin/upload-file-s3"
            test_pdffilesStr = ', '.join(test_pdffiles)
            PARAMS = {'pdf_paths':test_pdffilesStr}
            response =requests.get(url = URL, params = PARAMS, timeout=None)        
            response.close()  
            print("mmk"+test_pdffilesStr)

removePath=rootDir+sys.argv[7]+'/'
#print(removePath)
rewisePath = path_pdf_moved.replace(removePath,'')

doc.save(path_pdf_moved, garbage=4, deflate=True)
#doc.save(path_pdf_moved, garbage=4, deflate=True, owner_pw="owner", encryption=encrypt_meth, permissions=perm)
#doc.save(path_pdf_moved, incremental=True)
#shutil.copyfile(output_file, path_pdf_moved)
print(rewisePath)
#shutil.make_archive(folder+"/"+template_name+"-"+str(record_unique_id), "zip", inner_folder)
#print(folder+"/"+template_name+"-"+str(record_unique_id)+".zip")
total_pages=cnt-1
#print("Total Records:"+str(total_pages))
shutil.rmtree(dirName, ignore_errors=True)
#shutil.rmtree(inner_folder, ignore_errors=True)
#print("documents/"+template_name+"/"+template_name+"-"+str(record_unique_id)+".zip")
#print(pdf_folder+"/"+dt_string+".pdf");
#print(rewisePath)
#Upload file to secdoc directory
"""
if sys.argv[13] :
    URL = "https://"+sys.argv[7]+"."+sys.argv[10]+"/admin/store-file"
    pdf_path=pdf_folder+"/"+dt_string+".pdf"
    PARAMS = {'pdf_path':pdf_path,'site_id':sys.argv[13]}
    response =requests.get(url = URL, params = PARAMS)
"""


